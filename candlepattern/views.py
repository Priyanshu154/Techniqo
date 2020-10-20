from django.shortcuts import render
import openpyxl as xl
from openpyxl.utils import column_index_from_string
import os
import datetime
import traceback

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip
def assign_cell(cp):
    if(cp == 'Hammer'): return 'M'
    elif(cp == 'Dragonfly Doji'): return 'L'
    elif(cp == 'White Marubozu'): return 'Q'
    elif(cp == 'Bullish Engulfing'): return 'AE'
    elif(cp == 'Bullish Harami'):   return 'AC'
    elif(cp == 'Rising Sun'):       return 'AG'
    elif(cp == 'Morning Star'):     return 'AU'
    elif(cp == 'Three White Soldiers'): return 'AW'
    elif(cp == 'Inverted Hammer'):       return 'O'
    elif(cp == 'Gravestone Doji'):      return 'N'
    elif(cp == 'Black Marubozu'):       return 'R'
    elif(cp == 'Bearish Engulfing'):    return 'AF'
    elif(cp == 'Bearish Harami'):       return 'AD'
    elif(cp == 'Dark Cloud'):           return 'AH'
    elif(cp == 'Evening Star'):         return 'AV'
    elif(cp == 'Three Black Crows'):    return 'AX'
    elif(cp == 'Doji'):                 return 'P'
    return 'A'
# Create your views here.

def lele(cp):
    if cp == 'Hammer' or 'Dragonfly Doji' or 'White Marubozu' or 'Bullish Engulfing' or 'Bullish Harami' or 'Rising Sun' or 'Morning Star' or 'Three White Soldiers':
        return "Bullish"
    elif cp == 'Inverted Hammer' or 'Gravestone Doji' or 'Black Marubozu' or 'Bearish Engulfing' or 'Bearish Harami' or 'Dark Cloud' or 'Evening Star' or 'Three Black Crows':
        return "Bearish"
    elif (cp == 'Doji'):
        return "Neutral"

def assign_para(cp):
    if(cp == 'Hammer'): return "A hammer is a bullish price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    elif(cp == 'Dragonfly Doji'): return "A Dragonfly Doji is a bullish type of candlestick pattern that can signal a potential reversal in price to the downside or upside, depending on past price action. It's formed when the asset's high, open, and close prices are the same. The long lower shadow suggests that there was aggressive selling during the period of the candle, but since the price closed near the open it shows that buyers were able to absorb the selling and push the price back up."
    elif(cp == 'White Marubozu'): return "White Marubozu is a bullish reversal/continuation pattern. It is a large white candlestick with no wicks on either end. It is considered to be an extremely bullish candle. The candle can lead to a continuation of the current uptrend or start of a bearish reversal. This candle often occurs with high volume."
    elif(cp == 'Bullish Engulfing'): return "The bullish engulfing pattern is a two-candle reversal pattern. The second candle completely ‘engulfs’ the real body of the first one, without regard to the length of the tail shadows. The Bullish Engulfing pattern appears in a downtrend and is a combination of one dark candle followed by a larger hollow candle. On the second day of the pattern, price opens lower than the previous low, yet buying pressure pushes the price up to a higher level than the previous high, culminating in an obvious win for the buyers. It is advisable to enter a long position when the price moves higher than the high of the second engulfing candle—in other words when the downtrend reversal is confirmed."
    elif(cp == 'Bullish Harami'):   return "A bullish harami is a candlestick chart indicator suggesting that a bearish trend may be coming to end. Some investors may look at a bullish harami as a good sign that they should enter a long position on an asset."
    elif(cp == 'Rising Sun'):   return "A rising sun pattern is a two-day bullish candlestick price pattern that marks a potential short-term reversal from a downward trend to an upward trend. The pattern includes the first day opening near the high and closing near the low with an average or larger-sized trading range. It also includes a gap down after the first day where the second day begins trading, opening near the low and closing near the high. The close should also be a candlestick that covers at least half of the upward length of the previous day's red candlestick body."
    elif(cp == 'Morning Star'): return "A morning star is a visual bullish pattern consisting of three candlesticks that is interpreted as a bullish sign by technical analysts. A morning star forms following a downward trend and it indicates the start of an upward climb. It is a sign of a reversal in the previous price trend. Traders watch for the formation of a morning star and then seek confirmation that a reversal is indeed occurring using additional indicators."
    elif(cp == 'Three White Soldiers'): return "Typically occurring at the end of a downtrend, the bullish three white soldiers consists of three large bullish candles, each closing higher than the last. However, there should be no gaps between candles – each candle opens within the body of the one preceding it."
    elif(cp == 'Inverted Hammer'):      return "The inverted hammer candle is bearish and has a small red body, an extended upper wick and little or no lower wick. It indicates seller pressure is way far more than buyer's pressure"
    elif(cp == 'Gravestone Doji'):  return "A gravestone doji is a bearish reversal candlestick pattern that is formed when the open, low, and closing prices are all near each other with a long upper shadow. The long upper shadow suggests that the bullish advance in the beginning of the session was overcome by bears by the end of the session, which often comes just before a longer term bearish downtrend."
    elif(cp == 'Black Marubozu'):   return "The black marubozu is simply a bearish long black (down, or red on the charts below) candle, with little to no upper or lower shadows. The pattern shows that sellers controlled the trading day from open to close, and is therefore a bearish pattern. The candlestick can provide a trade signal or analytical insight into the future direction of a stock price."
    elif(cp == 'Bearish Engulfing'):  return '''A bearish engulfing pattern is a technical chart pattern that signals lower prices to come. The pattern consists of an up (white or green) candlestick followed by a large down (black or red) candlestick that eclipses or "engulfs" the smaller up candle. The pattern can be important because it shows sellers have overtaken the buyers and are pushing the price more aggressively down (down candle) than the buyers were able to push it up (up candle).'''
    elif(cp == 'Bearish Harami'):   return "A bearish harami is a two bar Japanese candlestick pattern that suggests prices may soon reverse to the downside. The pattern consists of a long white candle followed by a small black candle. The opening and closing prices of the second candle must be contained within the body of the first candle. An uptrend precedes the formation of a bearish harami."
    elif(cp == 'Dark Cloud'):   return "Dark Cloud Cover is a bearish reversal candlestick pattern where a down candle (typically black or red) opens above the close of the prior up candle (typically white or green), and then closes below the midpoint of the up candle.<br>The pattern is significant as it shows a shift in the momentum from the upside to the downside. The pattern is created by an up candle followed by a down candle. Traders look for the price to continue lower on the next (third) candle. This is called confirmation."
    elif(cp == 'Evening Star'): return "An Evening Star is a bearish stock-price chart pattern used by technical analysts to detect when a trend is about to reverse. It is a bearish candlestick pattern consisting of three candles: a large white candlestick, a small-bodied candle, and a red candle."
    elif(cp == 'Three Black Crows'): return "Three black crows are a visual bearish pattern, meaning that there are no particular calculations to worry about when identifying this indicator. The three black crows pattern occurs when bears overtake the bulls during three consecutive trading sessions. The pattern shows on the pricing charts as three bearish long-bodied candlesticks with short or no shadows or wicks."
    elif(cp == 'Doji'): return "The long-legged doji is a neutral candlestick that consists of long upper and lower shadows and has approximately the same opening and closing price. The candlestick signals indecision about the future direction of the underlying security."
def index(request):
    try:
        return render(request,'candlepattern/candlepatternh.html')
    except Exception as e:
        workpath = os.path.dirname( os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, "errors.xlsx")  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row , 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")

def candle(request):
    try:
        ni = request.GET.get("nifty", "nifty_500")
        cp = request.GET.get("candle", "Hammer")
        print(cp)
        col_name = assign_cell(cp)
        type_kevi = lele(cp)
        nifty = ni + ".xlsx"
        #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
        workpath = os.path.dirname(os.path.abspath(__file__))
        xx = os.path.join(workpath, nifty) #yaha tak
        wb = xl.load_workbook(xx, data_only=True)
        sheet = wb['Sheet1']
        cell = sheet['a1']
        stocks = []
        opens = []
        high = []
        low = []
        close = []
        num = 0
        for i in range(2, sheet.max_row + 1):

            cell = sheet.cell(i, column_index_from_string(col_name))
            c = cell.value
            if c == "YES":
                num = 1
                stocks.append(sheet.cell(i, 1).value)
                opens.append(sheet.cell(i, 2).value)
                high.append(sheet.cell(i, 3).value)
                low.append(sheet.cell(i, 4).value)
                close.append(sheet.cell(i, 5).value)
        signal = "Please select index by default result shows Nifty 500 stocks"
        para = assign_para(cp)
        zipp = zip(stocks, opens, high, low, close)
        dictt = {'zips': zipp, 'candle': cp, 'number': num, 'paras': para, 'signals': signal, 'nifty': ni, 'typekevi': type_kevi}
        return render(request, 'result.html', dictt)
    except Exception as e:
        workpath = os.path.dirname( os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, f"errors.xlsx")  # yaha tak

        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row , 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")
