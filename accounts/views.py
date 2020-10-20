from django.shortcuts import render,redirect
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm,AuthenticationForm
from django.contrib.auth import login,logout, authenticate
from django.core.mail import send_mail
import random
import traceback
import os
import openpyxl as xl
import datetime
# Create your views here.
def signup_view(request):
    try:
        form = ""
        if(request.method == "POST"):
            form = UserCreationForm( request.POST )
            if(form.is_valid()):
                new_user = form.save()
                new_user = authenticate(username=form.cleaned_data['username'],
                                        password=form.cleaned_data['password1'],
                                        )
                login(request, new_user)
                return redirect('/')
        else:
            form = UserCreationForm()
        return render(request, "signup2.html", {"form": form} )
    except Exception as e:
        workpath = os.path.dirname( os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, "errors.xlsx")  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row, 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")

def login_view(request):
    try:
        form = ""
        if(request.method == "POST"):
            form = AuthenticationForm(data= request.POST)
            if(form.is_valid()):
                login(request, form.get_user())
                if('next' in request.GET):
                    return redirect(request.GET.get("next"))
                else:
                    return redirect('/')
        else:
            form = AuthenticationForm()
        return  render(request, "login2.html", {"form": form})
    except Exception as e:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, "errors.xlsx")  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row + 1, 1).value = str(e)
        sheet1.cell(sheet1.max_row, 2).value = request.path_info
        sheet1.cell(sheet1.max_row, 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row, 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")

def logout_view(request):
    try:
        if(request.method == "POST"):
            logout(request)
        return redirect("/")
    except Exception as e:
        workpath = os.path.dirname( os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, "errors.xlsx")  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row, 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")

def change_view(request):
    try:
        if(request.method == "POST"):
            if( request.user.is_authenticated):
                u = User.objects.get(username__exact= request.user.username)
                u.set_password(request.POST.get("password1"))
                u.save()
                return redirect("/")
        return render(request, "change.html")
    except Exception as e:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, "errors.xlsx")  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row + 1, 1).value = str(e)
        sheet1.cell(sheet1.max_row, 2).value = request.path_info
        sheet1.cell(sheet1.max_row, 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row, 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")

def generate_otp():
    otp = ""
    for i in range(0, 6):
        n = random.randint(0, 9)
        otp += str(n)
    return otp

def forgot_view(request):
    try:
        if (request.method == "POST"):
            try:
                otp = generate_otp()
                email = request.POST.get("email")
                u = User.objects.get(username__exact=email)
                u.set_password( otp )
                u.save()
                print(otp)
                subject = 'Techniqo reset password'
                message = \
                f"""
                Hi, {email} your new password is: {otp}
                    Instructions:
                    Use this password for next login and remember to change it 
                """
                email_from = "techniqo2020@gmail.com"
                recipient_list = [email]
                send_mail(subject, message, email_from, recipient_list)

                return render(request, "mailed.html")
            except:
                return render(request, "forgot.html", {"error": "user does not exist"})
        return render(request, "forgot.html")
    except Exception as e:
        workpath = os.path.dirname( os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, "errors.xlsx")  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row, 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")
