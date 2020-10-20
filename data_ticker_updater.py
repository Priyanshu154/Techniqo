import openpyxl as xl
from openpyxl.utils import column_index_from_string
import os
import matplotlib.dates

tickers = ["3MINDIA", "ABB", "ACC", "AIAENG", "APLAPOLLO", "AUBANK", "AARTIIND", "AAVAS", "ABBOTINDIA", "ADANIGAS",
           "ADANIGREEN", "ADANIPORTS", "ADANIPOWER", "ADANITRANS", "ABCAPITAL", "ABFRL", "ADVENZYMES", "AEGISCHEM",
           "AFFLE", "AJANTPHARM", "AKZOINDIA", "APLLTD", "ALKEM", "ALKYLAMINE", "ALLCARGO", "AMARAJABAT", "AMBER",
           "AMBUJACEM", "APOLLOHOSP", "APOLLOTYRE", "ARVINDFASN", "ASAHIINDIA", "ASHOKLEY", "ASHOKA", "ASIANPAINT",
           "ASTERDM", "ASTRAZEN", "ASTRAL", "ATUL", "AUROPHARMA", "AVANTIFEED", "DMART", "AXISBANK", "BASF", "BEML",
           "BSE", "BAJAJ-AUTO", "BAJAJCON", "BAJAJELEC", "BAJFINANCE", "BAJAJFINSV", "BAJAJHLDNG", "BALKRISIND",
           "BALMLAWRIE", "BALRAMCHIN", "BANDHANBNK", "BANKBARODA", "BANKINDIA", "MAHABANK", "BATAINDIA", "BAYERCROP",
           "BERGEPAINT", "BDL", "BEL", "BHARATFORG", "BHEL", "BPCL", "BHARATRAS", "BHARTIARTL", "INFRATEL", "BIOCON",
           "BIRLACORPN", "BSOFT", "BLISSGVS", "BLUEDART", "BLUESTARCO", "BBTC", "BOMDYEING", "BOSCHLTD", "BRIGADE",
           "BRITANNIA", "CARERATING", "CCL", "CESC", "CRISIL", "CSBBANK", "CADILAHC", "CANFINHOME", "CANBK",
           "CAPLIPOINT", "CGCL", "CARBORUNIV", "CASTROLIND", "CEATLTD", "CENTRALBK", "CDSL", "CENTURYPLY", "CENTURYTEX",
           "CERA", "CHAMBLFERT", "CHENNPETRO", "CHOLAHLDNG", "CHOLAFIN", "CIPLA", "CUB", "COALINDIA", "COCHINSHIP",
           "COFORGE", "COLPAL", "CONCOR", "COROMANDEL", "CREDITACC", "CROMPTON", "CUMMINSIND", "CYIENT", "DBCORP",
           "DCBBANK", "DCMSHRIRAM", "DLF", "DABUR", "DALBHARAT", "DEEPAKNTR", "DELTACORP", "DHANUKA", "DBL", "DISHTV",
           "DCAL", "DIVISLAB", "DIXON", "LALPATHLAB", "DRREDDY", "EIDPARRY", "EIHOTEL", "ESABINDIA", "EDELWEISS",
           "EICHERMOT", "ELGIEQUIP", "EMAMILTD", "ENDURANCE", "ENGINERSIN", "EQUITAS", "ERIS", "ESCORTS", "ESSELPACK",
           "EXIDEIND", "FDC", "FEDERALBNK", "FINEORG", "FINCABLES", "FINPIPE", "FSL", "FORTIS", "FCONSUMER", "FRETAIL",
           "GAIL", "GEPIL", "GET&D", "GHCL", "GMMPFAUDLR", "GMRINFRA", "GALAXYSURF", "GRSE", "GARFIBRES", "GICRE",
           "GILLETTE", "GLAXO", "GLENMARK", "GODFRYPHLP", "GODREJAGRO", "GODREJCP", "GODREJIND", "GODREJPROP",
           "GRANULES", "GRAPHITE", "GRASIM", "GESHIP", "GREAVESCOT", "GRINDWELL", "GUJALKALI", "FLUOROCHEM",
           "GUJGASLTD", "GMDCLTD", "GNFC", "GPPL", "GSFC", "GSPL", "GULFOILLUB", "HEG", "HCLTECH", "HDFCAMC",
           "HDFCBANK", "HDFCLIFE", "HFCL", "HATHWAY", "HATSUN", "HAVELLS", "HEIDELBERG", "HERITGFOOD", "HEROMOTOCO",
           "HEXAWARE", "HSCL", "HIMATSEIDE", "HINDALCO", "HAL", "HINDCOPPER", "HINDPETRO", "HINDUNILVR", "HINDZINC",
           "HONAUT", "HUDCO", "HDFC", "ICICIBANK", "ICICIGI", "ICICIPRULI", "ISEC", "ICRA", "IDBI", "IDFCFIRSTB",
           "IDFC", "IFBIND", "IFCI", "IIFL", "IIFLWAM", "IRB", "IRCON", "ITC", "ITI", "INDIACEM", "ITDC", "IBULHSGFIN",
           "IBREALEST", "IBVENTURES", "INDIAMART", "INDIANB", "IEX", "INDHOTEL", "IOC", "IOB", "IRCTC", "INDOSTAR",
           "INDOCO", "IGL", "INDUSINDBK", "INFIBEAM", "NAUKRI", "INFY", "INGERRAND", "INOXLEISUR", "INTELLECT",
           "INDIGO", "IPCALAB", "JBCHEPHARM", "JKCEMENT", "JKLAKSHMI", "JKPAPER", "JKTYRE", "JMFINANCIL", "JSWENERGY",
           "JSWSTEEL", "JAGRAN", "JAICORPLTD", "J&KBANK", "JAMNAAUTO", "JINDALSAW", "JSLHISAR", "JSL", "JINDALSTEL",
           "JCHAC", "JUBLFOOD", "JUBILANT", "JUSTDIAL", "JYOTHYLAB", "KPRMILL", "KEI", "KNRCON", "KPITTECH", "KRBL",
           "KSB", "KAJARIACER", "KALPATPOWR", "KANSAINER", "KTKBANK", "KARURVYSYA", "KSCL", "KEC", "KOLTEPATIL",
           "KOTAKBANK", "L&TFH", "LTTS", "LICHSGFIN", "LAOPALA", "LAXMIMACH", "LTI", "LT", "LAURUSLABS", "LEMONTREE",
           "LINDEINDIA", "LUPIN", "LUXIND", "MASFIN", "MMTC", "MOIL", "MRF", "MGL", "MAHSCOOTER", "MAHSEAMLES",
           "M&MFIN", "M&M", "MAHINDCIE", "MHRIL", "MAHLOG", "MANAPPURAM", "MRPL", "MARICO", "MARUTI", "MFSL",
           "METROPOLIS", "MINDTREE", "MINDACORP", "MINDAIND", "MIDHANI", "MOTHERSUMI", "MOTILALOFS", "MPHASIS", "MCX",
           "MUTHOOTFIN", "NATCOPHARM", "NBCC", "NCC", "NESCO", "NHPC", "NLCINDIA", "NMDC", "NTPC", "NH", "NATIONALUM",
           "NFL", "NBVENTURES", "NAVINFLUOR", "NESTLEIND", "NILKAMAL", "NAM-INDIA", "OBEROIRLTY", "ONGC", "OIL",
           "OMAXE", "OFSS", "ORIENTCEM", "ORIENTELEC", "ORIENTREF", "PIIND", "PNBHOUSING", "PNCINFRA", "PTC", "PVR",
           "PAGEIND", "PERSISTENT", "PETRONET", "PFIZER", "PHILIPCARB", "PHOENIXLTD", "PIDILITIND", "PEL", "POLYMED",
           "POLYCAB", "PFC", "POWERGRID", "PRAJIND", "PRESTIGE", "PRSMJOHNSN", "PGHL", "PGHH", "PNB", "QUESS",
           "RBLBANK", "RECLTD", "RITES", "RADICO", "RVNL", "RAIN", "RAJESHEXPO", "RALLIS", "RCF", "RATNAMANI",
           "RAYMOND", "REDINGTON", "RELAXO", "RELIANCE", "REPCOHOME", "SBICARD", "SBILIFE", "SJVN", "SKFINDIA", "SRF",
           "SADBHAV", "SANOFI", "SCHAEFFLER", "SCHNEIDER", "SIS", "SEQUENT", "SFL", "SCI", "SHOPERSTOP", "SHREECEM",
           "RENUKA", "SHRIRAMCIT", "SRTRANSFIN", "SIEMENS", "SOBHA", "SOLARINDS", "SONATSOFTW", "SOUTHBANK", "SPANDANA",
           "SPICEJET", "STARCEMENT", "SBIN", "SAIL", "SWSOLAR", "STRTECH", "STAR", "SUDARSCHEM", "SUMICHEM", "SPARC",
           "SUNPHARMA", "SUNTV", "SUNDARMFIN", "SUNDRMFAST", "SUNTECK", "SUPRAJIT", "SUPREMEIND", "SUZLON",
           "SWANENERGY", "SYMPHONY", "SYNGENE", "TCIEXP", "TCNSBRANDS", "TTKPRESTIG", "TVTODAY", "TV18BRDCST",
           "TVSMOTOR", "TAKE", "TASTYBITE", "TATACOMM", "TCS", "TATACONSUM", "TATAELXSI", "TATAINVEST", "TATAMTRDVR",
           "TATAMOTORS", "TATAPOWER", "TATASTLBSL", "TATASTEEL", "TEAMLEASE", "TECHM", "NIACL", "RAMCOCEM", "THERMAX",
           "THYROCARE", "TIMETECHNO", "TIMKEN", "TITAN", "TORNTPHARM", "TORNTPOWER", "TRENT", "TRIDENT", "TIINDIA",
           "UCOBANK", "UFLEX", "UPL", "UJJIVAN", "UJJIVANSFB", "ULTRACEMCO", "UNIONBANK", "UBL", "MCDOWELL-N", "VGUARD",
           "VMART", "VIPIND", "VRLLOG", "VSTIND", "VAIBHAVGBL", "VAKRANGEE", "VTL", "VARROC", "VBL", "VENKEYS",
           "VESUVIUS", "VINATIORGA", "IDEA", "VOLTAS", "WABCOINDIA", "WELCORP", "WELSPUNIND", "WESTLIFE", "WHIRLPOOL",
           "WIPRO", "WOCKPHARMA", "ZEEL", "ZENSARTECH", "ZYDUSWELL", "ECLERX"]

workpath = os.path.dirname(os.path.abspath(__file__))
count = 1
xx2 = os.path.join(workpath, f'F:/My Projects/techniqo/bhavcopynew.xlsx')  # yaha tak
wb2 = xl.load_workbook(xx2, data_only=True)
sheet2 = wb2["sheet"]
for ticker in tickers:
    try:

        xx = os.path.join(workpath, f'F:/My Projects/techniqo/data_new_ticker/{ticker}.xlsx')  # yaha tak
        wb = xl.load_workbook(xx, data_only=True)
        sheet = wb["Sheet1"]

        row_to_append = sheet.max_row+1

        sheet.cell(sheet.max_row+1, 1).value = "19-10-2020"  # Date manually nakhvi padse date month year
        for row in range(1, sheet2.max_row + 1):
            if sheet2.cell(row, 1).value == ticker and sheet2.cell(row, 2).value == "EQ":
                sheet.cell(row_to_append, 2).value = sheet2.cell(row, 3).value
                sheet.cell(row_to_append, 3).value = sheet2.cell(
                    row, 4).value
                sheet.cell(row_to_append, 4).value = sheet2.cell(row, 5).value
                sheet.cell(row_to_append, 5).value = sheet2.cell(row, 6).value
                sheet.cell(row_to_append, 6).value = sheet2.cell(row, 9).value
                sheet.cell(row_to_append, 7).value = matplotlib.dates.datestr2num("2020-10-19")  # ymd

        wb.save(f'F:/My Projects/techniqo/data_new_ticker/{ticker}.xlsx')
        wb2.save(f'F:/My Projects/techniqo/bhavcopynew.xlsx')
        print(count)
        count = count + 1
    except:
        print(ticker)
