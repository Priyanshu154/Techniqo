# I have created
from urllib import request
from django.shortcuts import render
from django.shortcuts import redirect
import openpyxl as xl
import os
from . import settings
import datetime
import traceback
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip
# Create your views here.

def index(request):
    try:
        return render(request,'homepage.html')
    except Exception as e:
        workpath = os.path.dirname( os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, "errors.xlsx")  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row, 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")
def redirect_previous(request, str):
    for app in settings.INSTALLED_APPS:
        ind = str.find( app )
        if(app == "stock"):
            return "http://" + request.get_host() + "/"
        if(ind != -1):
            return "http://" + request.get_host() + "/" + app
    return "invalid"

def suggest(request):
    try:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, 'FeedbackData.xlsx')  # yaha tak
        wb = xl.load_workbook(xx)
        sheet = wb["Sheet1"]
        feedbackdata = request.POST.get('feedback')
        sheet.cell(sheet.max_row+1,1).value = feedbackdata
        wb.save(xx)
        curr_url =  request.POST.get("curr_url")
        curr_url = redirect_previous(request, curr_url)
        if(curr_url != "invalid"):
            response = redirect(curr_url)
            return response
        else:
            response = redirect("http://"+request.get_host())
            return response
    except Exception as e:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, 'errors.xlsx')  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row , 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")
