from django.shortcuts import render
from django.shortcuts import redirect
import pandas_datareader.data as web
from datetime import datetime
from matplotlib.dates import date2num
import statistics
from . import data_indic
import openpyxl as xl
import os
import pandas as pd
import xlrd

# Create your views here.
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def backdata(request):
    try:
        dataa = request.POST.get('backe', 'default')
        print(dataa)
        inter = dataa.split(",||")
        long_short = inter[5].strip()
        yahoo_st = inter[4].split("|")[1].strip()
        print(yahoo_st)
        #start = datetime(2017, 6, 6)
        #end = datetime(2020, 8, 17)
        workpath = os.path.dirname( os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, f"data_new_ticker/{yahoo_st}.xlsx")  # yaha tak
        wb = xl.load_workbook(xx, data_only=True)
        sheet = wb["Sheet1"]
        data_reset = pd.read_excel(xx)
        wb.save(xx)
        close = data_reset['Close'].to_list()
        high = data_reset['High'].to_list()
        low = data_reset['Low'].to_list()
        openn = data_reset['Open'].to_list()
        date = data_reset['Date'].to_list()
        dt = data_reset['date_ax'].to_list()
        volume = data_reset['Volume'].to_list()

        col = 'a'
        i = 2
        macd_line = []
        signal_line = []
        stoploss = 1000
        if inter[3] == "":
            stoploss = 1000
        else:
            stoploss = int(inter[3])

        inter_indi = inter[0].split(",!!")
        inter_para = inter[1].split(",!!")
        inter_val = inter[2].split(",!!")

        indicator_entry_str = inter_indi[0]
        parameter_entry_str = inter_para[0]
        value_entry_str = inter_val[0]
        indicator_exit_str = inter_indi[1]
        parameter_exit_str = inter_para[1]
        value_exit_str = inter_val[1]

        parameter_entry_str = parameter_entry_str[1:]

        indicator_entry = indicator_entry_str.split(",")
        parameter_entry = parameter_entry_str.split(",")
        value_entry = value_entry_str.split(",")
        indicator_exit = indicator_exit_str.split(",")
        parameter_exit = parameter_exit_str.split(",")
        value_exit = value_exit_str.split(",")

        count_entry = len(indicator_entry)

        count_exit = len(indicator_exit)

        rsi_arr = []
        bb_arr = []
        macd_arr = []
        mfi_arr = []
        roc_arr = []
        srsi_arr = []
        wil_arr = []
        sma_arr = []
        ema_arr = []
        pp_arr = []
        ic_arr = []

        for i in range(len(indicator_entry)):
            rsi_arr = []
            bb_arr = []
            macd_arr = []
            mfi_arr = []
            roc_arr = []
            srsi_arr = []
            wil_arr = []
            sma_arr = []
            ema_arr = []
            pp_arr = []
            ic_arr = []
            if indicator_entry[i] == "rsi":
                val = data_indic.RSI(close, 14)  # aiyan t
                if parameter_entry[i] == "crossover":
                    for c in range(14):
                        rsi_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] <= int(value_entry[i]) < val[j] != 0 and val[j - 1] != 0:
                            rsi_arr.append("Yes")
                        else:
                            rsi_arr.append("No")
                    data_reset[col] = rsi_arr
                    col += 'a'

                elif parameter_entry[i] == "crossunder":
                    for c in range(14):
                        rsi_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] >= int(value_entry[i]) > val[j] != 0 and val[j - 1] != 0:
                            rsi_arr.append("Yes")
                        else:
                            rsi_arr.append("No")
                    data_reset[col] = rsi_arr
                    col += 'a'

                elif parameter_entry[i] == "above":
                    for c in range(14):
                        rsi_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] > int(value_entry[i]):
                            rsi_arr.append("Yes")
                        else:
                            rsi_arr.append("No")
                    data_reset[col] = rsi_arr
                    col += 'a'

                elif parameter_entry[i] == "below":
                    for c in range(14):
                        rsi_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] < int(value_entry[i]):
                            rsi_arr.append("Yes")
                        else:
                            rsi_arr.append("No")
                    data_reset[col] = rsi_arr
                    col += 'a'

            elif indicator_entry[i] == "clo" and (value_entry[i] == "lb" or value_entry[i] == "mb" or value_entry[i] == "up"):
                upper, lower, middle = data_indic.bollinger_band(close, 20, 2)
                if parameter_entry[i] == "crossover":
                    for c in range(20):
                        bb_arr.append("No")
                    for j in range(20, len(close)):
                        if value_entry[i] == "lb":
                            if close[j - 1] <= lower[j] < close[j] and lower[j] != 0:
                                bb_arr.append("Yes")
                            else:
                                bb_arr.append("No")
                        elif value_entry[i] == "up":
                            if close[j - 1] <= upper[j] < close[j] and upper[j] != 0:
                                bb_arr.append("Yes")
                            else:
                                bb_arr.append("No")
                        elif value_entry[i] == "mb":
                            if close[j - 1] <= middle[j] < close[j] and middle[j] != 0:
                                bb_arr.append("Yes")
                            else:
                                bb_arr.append("No")
                    data_reset[col] = bb_arr
                    col += 'a'
                elif parameter_entry[i] == "crossunder":
                    for c in range(20):
                        bb_arr.append("No")
                    for j in range(20, len(close)):
                        if value_entry[i] == "lb":
                            if close[j - 1] >= lower[j] > close[j] and lower[j] != 0:
                                bb_arr.append("Yes")
                            else:
                                bb_arr.append("No")
                        elif value_entry[i] == "up":
                            if close[j - 1] >= upper[j] > close[j] and upper[j] != 0:
                                bb_arr.append("Yes")
                            else:
                                bb_arr.append("No")
                        elif value_entry[i] == "mb":
                            if close[j - 1] >= middle[j] > close[j] and middle[j] != 0:
                                bb_arr.append("Yes")
                            else:
                                bb_arr.append("No")
                    data_reset[col] = bb_arr
                    col += 'a'

                elif parameter_entry[i] == "above":
                    for c in range(20):
                        bb_arr.append("No")
                    for j in range(20, len(close)):
                        if value_entry[i] == "lb":
                            if close[j] > lower[j] != 0:
                                bb_arr.append("Yes")
                            else:
                                bb_arr.append("No")
                        elif value_entry[i] == "up":
                            if close[j] > upper[j] != 0:
                                bb_arr.append("Yes")
                            else:
                                bb_arr.append("No")
                        elif value_entry[i] == "mb":
                            if close[j] > middle[j] != 0:
                                bb_arr.append("Yes")
                            else:
                                bb_arr.append("No")
                    data_reset[col] = bb_arr
                    col += 'a'

                elif parameter_entry[i] == "below":
                    for c in range(20):
                        bb_arr.append("No")
                    for j in range(20, len(close)):
                        if value_entry[i] == "lb":
                            if close[j] < lower[j] != 0:
                                bb_arr.append("Yes")
                            else:
                                bb_arr.append("No")
                        elif value_entry[i] == "up":
                            if close[j] < upper[j] != 0:
                                bb_arr.append("Yes")
                            else:
                                bb_arr.append("No")
                        elif value_entry[i] == "mb":
                            if close[j] < middle[j] != 0:
                                bb_arr.append("Yes")
                            else:
                                bb_arr.append("No")
                    data_reset[col] = bb_arr
                    col += 'a'

            elif indicator_entry[i] == "macd":
                macd_line, signal_line, macd_histogram = data_indic.MACD(close, 12, 26, 9)
                if parameter_entry[i] == "crossover":
                    for c in range(26):
                        macd_arr.append("No")
                    for j in range(26, len(signal_line)):
                        if value_entry[i] == "sig":
                            if macd_line[j - 1] <= signal_line[j] < macd_line[j]:
                                macd_arr.append("Yes")
                            else:
                                macd_arr.append("No")
                        elif value_entry[i] == "zero":
                            if macd_line[j - 1] <= 0 < macd_line[j]:
                                macd_arr.append("Yes")
                            else:
                                macd_arr.append("No")
                    data_reset[col] = macd_arr
                    col += 'a'
                elif parameter_entry[i] == "crossunder":
                    for c in range(26):
                        macd_arr.append("No")
                    for j in range(26, len(signal_line)):
                        if value_entry[i] == "sig":
                            if macd_line[j - 1] >= signal_line[j] > macd_line[j]:
                                macd_arr.append("Yes")
                            else:
                                macd_arr.append("No")
                        elif value_entry[i] == "zero":
                            if macd_line[j - 1] >= 0 > macd_line[j]:
                                macd_arr.append("Yes")
                            else:
                                macd_arr.append("No")
                    data_reset[col] = macd_arr
                    col += 'a'

                elif parameter_entry[i] == "above":
                    for c in range(26):
                        macd_arr.append("No")
                    for j in range(26, len(signal_line)):
                        if value_entry[i] == "sig":
                            if signal_line[j] < macd_line[j]:
                                macd_arr.append("Yes")
                            else:
                                macd_arr.append("No")
                        elif value_entry[i] == "zero":
                            if 0 < macd_line[j]:
                                macd_arr.append("Yes")
                            else:
                                macd_arr.append("No")
                    data_reset[col] = macd_arr
                    col += 'a'

                elif parameter_entry[i] == "below":
                    for c in range(26):
                        macd_arr.append("No")
                    for j in range(26, len(signal_line)):
                        if value_entry[i] == "sig":
                            if signal_line[j] > macd_line[j]:
                                macd_arr.append("Yes")
                            else:
                                macd_arr.append("No")
                        elif value_entry[i] == "zero":
                            if 0 > macd_line[j]:
                                macd_arr.append("Yes")
                            else:
                                macd_arr.append("No")
                    data_reset[col] = macd_arr
                    col += 'a'

            elif indicator_entry[i] == "mfi":
                val = data_indic.MFI(high, low, close, volume, 14)  # aiyan t
                if parameter_entry[i] == "crossover":
                    for c in range(14):
                        mfi_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] <= int(value_entry[i]) < val[j] != 0 and val[j - 1] != 0:
                            mfi_arr.append("Yes")
                        else:
                            mfi_arr.append("No")
                    data_reset[col] = mfi_arr
                    col += 'a'

                elif parameter_entry[i] == "crossunder":
                    for c in range(14):
                        mfi_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] >= int(value_entry[i]) > val[j] != 0 and val[j - 1] != 0:
                            mfi_arr.append("Yes")
                        else:
                            mfi_arr.append("No")
                    data_reset[col] = mfi_arr
                    col += 'a'

                elif parameter_entry[i] == "above":
                    for c in range(14):
                        mfi_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] > int(value_entry[i]):
                            mfi_arr.append("Yes")
                        else:
                            mfi_arr.append("No")
                    data_reset[col] = mfi_arr
                    col += 'a'

                elif parameter_entry[i] == "below":
                    for c in range(14):
                        mfi_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] < int(value_entry[i]):
                            mfi_arr.append("Yes")
                        else:
                            mfi_arr.append("No")
                    data_reset[col] = mfi_arr
                    col += 'a'

            elif indicator_entry[i] == "roc":
                val = data_indic.ROC(close, 9)  # aiyan t
                if parameter_entry[i] == "crossover":
                    for c in range(14):
                        roc_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] <= int(value_entry[i]) < val[j] != 0 and val[j - 1] != 0:
                            roc_arr.append("Yes")
                        else:
                            roc_arr.append("No")
                    data_reset[col] = roc_arr
                    col += 'a'

                elif parameter_entry[i] == "crossunder":
                    for c in range(14):
                        roc_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] >= int(value_entry[i]) > val[j] != 0 and val[j - 1] != 0:
                            roc_arr.append("Yes")
                        else:
                            roc_arr.append("No")
                    data_reset[col] = roc_arr
                    col += 'a'

                elif parameter_entry[i] == "above":
                    for c in range(14):
                        roc_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] > int(value_entry[i]):
                            roc_arr.append("Yes")
                        else:
                            roc_arr.append("No")
                    data_reset[col] = roc_arr
                    col += 'a'

                elif parameter_entry[i] == "below":
                    for c in range(14):
                        roc_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] < int(value_entry[i]):
                            roc_arr.append("Yes")
                        else:
                            roc_arr.append("No")
                    data_reset[col] = roc_arr
                    col += 'a'

            elif indicator_entry[i] == "srsi":
                k_line, d_line = data_indic.S_RSI(close, 14, 3, 3, 14)
                if parameter_entry[i] == "crossover":
                    for c in range(14):
                        srsi_arr.append("No")
                    for j in range(14, len(d_line)):
                        if value_entry[i] == "slow":
                            if k_line[j - 1] <= d_line[j] < k_line[j]:
                                srsi_arr.append("Yes")
                            else:
                                srsi_arr.append("No")
                        else:
                            if k_line[j - 1] <= int(value_entry[i]) < k_line[j]:
                                srsi_arr.append("Yes")
                            else:
                                srsi_arr.append("No")
                    data_reset[col] = srsi_arr
                    col += 'a'
                elif parameter_entry[i] == "crossunder":
                    for c in range(14):
                        srsi_arr.append("No")
                    for j in range(14, len(d_line)):
                        if value_entry[i] == "slow":
                            if k_line[j - 1] >= d_line[j] > k_line[j]:
                                srsi_arr.append("Yes")
                            else:
                                srsi_arr.append("No")
                        else:
                            if k_line[j - 1] >= int(value_entry[i]) > k_line[j]:
                                srsi_arr.append("Yes")
                            else:
                                srsi_arr.append("No")
                    data_reset[col] = srsi_arr
                    col += 'a'

                elif parameter_entry[i] == "above":
                    for c in range(14):
                        srsi_arr.append("No")
                    for j in range(14, len(d_line)):
                        if value_entry[i] == "slow":
                            if d_line[j] < k_line[j]:
                                srsi_arr.append("Yes")
                            else:
                                srsi_arr.append("No")
                        else:
                            if int(value_entry[i]) < k_line[j]:
                                srsi_arr.append("Yes")
                            else:
                                srsi_arr.append("No")
                    data_reset[col] = srsi_arr
                    col += 'a'

                elif parameter_entry[i] == "below":
                    for c in range(14):
                        srsi_arr.append("No")
                    for j in range(14, len(d_line)):
                        if value_entry[i] == "sig":
                            if d_line[j] > k_line[j]:
                                srsi_arr.append("Yes")
                            else:
                                srsi_arr.append("No")
                        else:
                            if int(value_entry[i]) > k_line[j]:
                                srsi_arr.append("Yes")
                            else:
                                srsi_arr.append("No")
                    data_reset[col] = srsi_arr
                    col += 'a'

            elif indicator_entry[i] == "wil":
                val = data_indic.WILLIAM_R(close, 14, high, low)  # aiyan t
                if parameter_entry[i] == "crossover":
                    for c in range(14):
                        wil_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] <= int(value_entry[i]) < val[j] != 0 and val[j - 1] != 0:
                            wil_arr.append("Yes")
                        else:
                            wil_arr.append("No")
                    data_reset[col] = wil_arr
                    col += 'a'

                elif parameter_entry[i] == "crossunder":
                    for c in range(14):
                        wil_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] >= int(value_entry[i]) > val[j] != 0 and val[j - 1] != 0:
                            wil_arr.append("Yes")
                        else:
                            wil_arr.append("No")
                    data_reset[col] = wil_arr
                    col += 'a'

                elif parameter_entry[i] == "above":
                    for c in range(14):
                        wil_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] > int(value_entry[i]):
                            wil_arr.append("Yes")
                        else:
                            wil_arr.append("No")
                    data_reset[col] = wil_arr
                    col += 'a'

                elif parameter_entry[i] == "below":
                    for c in range(14):
                        wil_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] < int(value_entry[i]):
                            wil_arr.append("Yes")
                        else:
                            wil_arr.append("No")
                    data_reset[col] = wil_arr
                    col += 'a'

            elif indicator_entry[i] == "sma10" or indicator_entry[i] == "sma20" or indicator_entry[i] == "sma50" or \
                    indicator_entry[i] == "sma100" or indicator_entry[i] == "sma200" or (indicator_entry[i] == "clo" and (
                    value_entry[i] == "sma10" or value_entry[i] == "sma20" or value_entry[i] == "sma50" or value_entry[
                i] == "sma100" or value_entry[i] == "sma200")):
                val2 = []
                val = []
                if value_entry[i] == "sma20":
                    val2 = data_indic.SMA(close, 20)
                elif value_entry[i] == "sma50":
                    val2 = data_indic.SMA(close, 50)
                elif value_entry[i] == "sma100":
                    val2 = data_indic.SMA(close, 100)
                elif value_entry[i] == "sma200":
                    val2 = data_indic.SMA(close, 200)
                elif value_entry[i] == "sma10":
                    val2 = data_indic.SMA(close, 10)

                if indicator_entry[i] == "sma20":
                    val = data_indic.SMA(close, 20)
                elif indicator_entry[i] == "sma50":
                    val = data_indic.SMA(close, 50)
                elif indicator_entry[i] == "sma100":
                    val = data_indic.SMA(close, 100)
                elif indicator_entry[i] == "sma200":
                    val = data_indic.SMA(close, 200)
                elif indicator_entry[i] == "sma10":
                    val = data_indic.SMA(close, 10)
                elif indicator_entry[i] == "clo":
                    val = close

                ty = int(value_entry[i].split("a")[1])

                if parameter_entry[i] == "crossover":
                    for c in range(ty):
                        sma_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val[j - 1] < val2[j] < val[j]:
                            sma_arr.append("Yes")
                        else:
                            sma_arr.append("No")
                    data_reset[col] = sma_arr
                    col += 'a'
                elif parameter_entry[i] == "crossunder":
                    for c in range(ty):
                        sma_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val[j - 1] > val2[j] > val[j]:
                            sma_arr.append("Yes")
                        else:
                            sma_arr.append("No")
                    data_reset[col] = sma_arr
                    col += 'a'
                elif parameter_entry[i] == "above":
                    for c in range(ty):
                        sma_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val2[j] < val[j]:
                            sma_arr.append("Yes")
                        else:
                            sma_arr.append("No")
                    data_reset[col] = sma_arr
                    col += 'a'
                elif parameter_entry[i] == "below":
                    for c in range(ty):
                        sma_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val2[j] > val[j]:
                            sma_arr.append("Yes")
                        else:
                            sma_arr.append("No")
                    data_reset[col] = sma_arr
                    col += 'a'

            elif indicator_entry[i] == "ema10" or indicator_entry[i] == "ema20" or indicator_entry[i] == "ema50" or \
                    indicator_entry[i] == "ema100" or indicator_entry[i] == "ema200" or (indicator_entry[i] == "clo" and (
                    value_entry[i] == "ema10" or value_entry[i] == "ema20" or value_entry[i] == "ema50" or value_entry[
                i] == "ema100" or value_entry[i] == "ema200")):
                val2 = []
                val = []
                if value_entry[i] == "ema20":
                    val2 = data_indic.EMA(close, 20)
                elif value_entry[i] == "ema50":
                    val2 = data_indic.EMA(close, 50)
                elif value_entry[i] == "ema100":
                    val2 = data_indic.EMA(close, 100)
                elif value_entry[i] == "ema200":
                    val2 = data_indic.EMA(close, 200)
                elif value_entry[i] == "ema10":
                    val2 = data_indic.EMA(close, 10)

                if indicator_entry[i] == "ema20":
                    val = data_indic.EMA(close, 20)
                elif indicator_entry[i] == "ema50":
                    val = data_indic.EMA(close, 50)
                elif indicator_entry[i] == "ema100":
                    val = data_indic.EMA(close, 100)
                elif indicator_entry[i] == "ema200":
                    val = data_indic.EMA(close, 200)
                elif indicator_entry[i] == "ema10":
                    val = data_indic.EMA(close, 10)
                elif indicator_entry[i] == "clo":
                    val = close

                ty = int(value_entry[i].split("a")[1])

                if parameter_entry[i] == "crossover":
                    for c in range(ty):
                        ema_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val[j - 1] < val2[j] < val[j]:
                            ema_arr.append("Yes")
                        else:
                            ema_arr.append("No")
                    data_reset[col] = ema_arr
                    col += 'a'
                elif parameter_entry[i] == "crossunder":
                    for c in range(ty):
                        ema_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val[j - 1] > val2[j] > val[j]:
                            ema_arr.append("Yes")
                        else:
                            ema_arr.append("No")
                    data_reset[col] = ema_arr
                    col += 'a'
                elif parameter_entry[i] == "above":
                    for c in range(ty):
                        ema_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val2[j] < val[j]:
                            ema_arr.append("Yes")
                        else:
                            ema_arr.append("No")
                    data_reset[col] = ema_arr
                    col += 'a'
                elif parameter_entry[i] == "below":
                    for c in range(ty):
                        ema_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val2[j] > val[j]:
                            ema_arr.append("Yes")
                        else:
                            ema_arr.append("No")
                    data_reset[col] = ema_arr
                    col += 'a'

            elif indicator_entry[i] == "clo" and (
                    value_entry[i] == "pp" or value_entry[i] == "s1" or value_entry[i] == "s2" or value_entry[i] == "s3" or
                    value_entry[i] == "r1" or value_entry[i] == "r2" or value_entry[i] == "r3"):
                pp, s1, s2, s3, r1, r2, r3 = data_indic.pivot_points(close, high, low, date)
                val = close
                print("entry")
                print(len(pp))
                print(len(r1))
                for jj in range(abs(len(close)-len(pp))):
                    pp.insert(0, 0)
                    s1.insert(0, 0)
                    s2.insert(0, 0)
                    s3.insert(0, 0)
                    r1.insert(0, 0)
                    r2.insert(0, 0)
                    r3.insert(0, 0)
                print(len(pp))
                print(len(r1))
                pp_arr.append("No")
                if parameter_entry[i] == "crossover":
                    if value_entry[i] == "pp":
                        for j in range(1, len(pp)):
                            if val[j - 1] < pp[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "s1":
                        for j in range(1, len(pp)):
                            if val[j - 1] < s1[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "s2":
                        for j in range(1, len(pp)):
                            if val[j - 1] < s2[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "s3":
                        for j in range(1, len(pp)):
                            if val[j - 1] < s3[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "r1":
                        for j in range(1, len(pp)):
                            if val[j - 1] < r1[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "r2":
                        for j in range(1, len(pp)):
                            if val[j - 1] < r1[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "r3":
                        for j in range(1, len(pp)):
                            if val[j - 1] < r3[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'

                elif parameter_entry[i] == "crossunder":
                    if value_entry[i] == "pp":
                        for j in range(1, len(pp)):
                            if val[j - 1] > pp[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "s1":
                        for j in range(1, len(pp)):
                            if val[j - 1] > s1[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "s2":
                        for j in range(1, len(pp)):
                            if val[j - 1] > s2[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "s3":
                        for j in range(1, len(pp)):
                            if val[j - 1] > s3[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "r1":
                        for j in range(1, len(pp)):
                            if val[j - 1] > r1[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "r2":
                        for j in range(1, len(pp)):
                            if val[j - 1] > r2[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "r3":
                        for j in range(1, len(pp)):
                            if val[j - 1] > r3[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'

                elif parameter_entry[i] == "above":
                    if value_entry[i] == "pp":
                        for j in range(1, len(pp)):
                            if pp[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "s1":
                        for j in range(1, len(pp)):
                            if s1[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "s2":
                        for j in range(1, len(pp)):
                            if s2[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "s3":
                        for j in range(1, len(pp)):
                            if s3[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "r1":
                        for j in range(1, len(pp)):
                            if r1[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "r2":
                        for j in range(1, len(pp)):
                            if r2[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "r3":
                        for j in range(1, len(pp)):
                            if r3[j] < val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'

                elif parameter_entry[i] == "below":
                    if value_entry[i] == "pp":
                        for j in range(1, len(pp)):
                            if pp[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "s1":
                        for j in range(1, len(pp)):
                            if s1[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "s2":
                        for j in range(1, len(pp)):
                            if s2[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "s3":
                        for j in range(1, len(pp)):
                            if s3[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "r1":
                        for j in range(1, len(pp)):
                            if r1[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "r2":
                        for j in range(1, len(pp)):
                            if r2[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'
                    elif value_entry[i] == "r3":
                        for j in range(1, len(pp)):
                            if r3[j] > val[j]:
                                pp_arr.append("Yes")
                            else:
                                pp_arr.append("No")
                        data_reset[col] = pp_arr
                        col += 'a'

            elif indicator_entry[i] == "cl" or indicator_entry[i] == "ls" or (indicator_entry[i] == "clo" and (
                    value_entry[i] == "cl" or value_entry[i] == "bl" or value_entry[i] == "ls" or value_entry[i] == "gc" or
                    value_entry[i] == "rc")):
                val2 = []
                val = []
                cl, bl, sa, sb, ls = data_indic.Icloud(high, low, close, 9, 26, 52, 26)
                if value_entry[i] == "cl":
                    val2 = cl
                elif value_entry[i] == "bl":
                    val2 = bl
                elif value_entry[i] == "ls":
                    val2 = ls
                elif value_entry[i] == "gc":
                    for jk in range(len(sa)):
                        val2.append(sa[jk]-sb[jk])
                elif value_entry[i] == "rc":
                    for jk in range(len(sa)):
                        val2.append(sb[jk] - sa[jk])

                if indicator_entry[i] == "cl":
                    val = cl
                elif indicator_entry[i] == "ls":
                    val = ls
                elif indicator_entry[i] == "clo":
                    val = close
                ic_arr.append("No")
                if parameter_entry[i] == "crossover":
                    for j in range(1, len(val)):
                        if val[j - 1] < val2[j] < val[j]:
                            ic_arr.append("Yes")
                        else:
                            ic_arr.append("No")
                    data_reset[col] = ic_arr
                    col += 'a'
                elif parameter_entry[i] == "crossunder":
                    for j in range(1, len(val)):
                        if val[j - 1] > val2[j] > val[j]:
                            ic_arr.append("Yes")
                        else:
                            ic_arr.append("No")
                    data_reset[col] = ic_arr
                    col += 'a'
                elif parameter_entry[i] == "above":
                    for j in range(1, len(val2)):
                        if val2[j] < val[j] and val2[j] != -1 and val2[j] != 0:
                            ic_arr.append("Yes")
                        else:
                            ic_arr.append("No")
                    data_reset[col] = ic_arr
                    col += 'a'
                elif parameter_entry[i] == "below":
                    for j in range(1, len(val2)):
                        if val2[j] > val[j]:
                            ic_arr.append("Yes")
                        else:
                            ic_arr.append("No")
                    data_reset[col] = ic_arr
                    col += 'a'

        rsie_arr = []
        bbe_arr = []
        macde_arr = []
        mfie_arr = []
        roce_arr = []
        srsie_arr = []
        wile_arr = []
        smae_arr = []
        emae_arr = []
        ppe_arr = []
        ice_arr = []
        # Exit data code
        for i in range(len(indicator_exit)):
            rsie_arr = []
            bbe_arr = []
            macde_arr = []
            mfie_arr = []
            roce_arr = []
            srsie_arr = []
            wile_arr = []
            smae_arr = []
            emae_arr = []
            ppe_arr = []
            ice_arr = []
            if indicator_exit[i] == "rsi":
                val = data_indic.RSI(close, 14)
                if parameter_exit[i] == "crossover":
                    for c in range(14):
                        rsie_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] <= int(value_exit[i]) < val[j] != 0 and val[j - 1] != 0:
                            rsie_arr.append("Yes")
                        else:
                            rsie_arr.append("No")
                    data_reset[col] = rsie_arr
                    col += 'a'

                elif parameter_exit[i] == "crossunder":
                    for c in range(14):
                        rsie_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] >= int(value_exit[i]) > val[j] != 0 and val[j - 1] != 0:
                            rsie_arr.append("Yes")
                        else:
                            rsie_arr.append("No")
                    data_reset[col] = rsie_arr
                    col += 'a'

                elif parameter_exit[i] == "above":
                    for c in range(14):
                        rsie_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] > int(value_exit[i]):
                            rsie_arr.append("Yes")
                        else:
                            rsie_arr.append("No")
                    data_reset[col] = rsie_arr
                    col += 'a'

                elif parameter_exit[i] == "below":
                    for c in range(14):
                        rsie_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] < int(value_exit[i]):
                            rsie_arr.append("Yes")
                        else:
                            rsie_arr.append("No")
                    data_reset[col] = rsie_arr
                    col += 'a'

            elif indicator_exit[i] == "clo" and (value_exit[i] == "lb" or value_exit[i] == "mb" or value_exit[i] == "up"):
                upper, lower, middle = data_indic.bollinger_band(close, 20, 2)
                if parameter_exit[i] == "crossover":
                    for c in range(20):
                        bbe_arr.append("No")
                    for j in range(20, len(close)):
                        if value_exit[i] == "lb":
                            if close[j - 1] <= lower[j] < close[j] and lower[j] != 0:
                                bbe_arr.append("Yes")
                            else:
                                bbe_arr.append("No")
                        elif value_exit[i] == "up":
                            if close[j - 1] <= upper[j] < close[j] and upper[j] != 0:
                                bbe_arr.append("Yes")
                            else:
                                bbe_arr.append("No")
                        elif value_exit[i] == "mb":
                            if close[j - 1] <= middle[j] < close[j] and middle[j] != 0:
                                bbe_arr.append("Yes")
                            else:
                                bbe_arr.append("No")
                    data_reset[col] = bbe_arr
                    col += 'a'

                elif parameter_exit[i] == "crossunder":
                    for c in range(20):
                        bbe_arr.append("No")
                    for j in range(20, len(close)):
                        if value_exit[i] == "lb":
                            if close[j - 1] >= lower[j] > close[j] and lower[j] != 0:
                                bbe_arr.append("Yes")
                            else:
                                bbe_arr.append("No")
                        elif value_exit[i] == "up":
                            if close[j - 1] >= upper[j] > close[j] and upper[j] != 0:
                                bbe_arr.append("Yes")
                            else:
                                bbe_arr.append("No")
                        elif value_exit[i] == "mb":
                            if close[j - 1] >= middle[j] > close[j] and middle[j] != 0:
                                bbe_arr.append("Yes")
                            else:
                                bbe_arr.append("No")
                    data_reset[col] = bbe_arr
                    col += 'a'

                elif parameter_exit[i] == "above":
                    for c in range(20):
                        bbe_arr.append("No")
                    for j in range(20, len(close)):
                        if value_exit[i] == "lb":
                            if close[j] > lower[j] != 0:
                                bbe_arr.append("Yes")
                            else:
                                bbe_arr.append("No")
                        elif value_exit[i] == "up":
                            if close[j] > upper[j] != 0:
                                bbe_arr.append("Yes")
                            else:
                                bbe_arr.append("No")
                        elif value_exit[i] == "mb":
                            if close[j] > middle[j] != 0:
                                bbe_arr.append("Yes")
                            else:
                                bbe_arr.append("No")
                    data_reset[col] = bbe_arr
                    col += 'a'

                elif parameter_exit[i] == "below":
                    for c in range(20):
                        bbe_arr.append("No")
                    for j in range(20, len(close)):
                        if value_exit[i] == "lb":
                            if close[j] < lower[j] != 0:
                                bbe_arr.append("Yes")
                            else:
                                bbe_arr.append("No")
                        elif value_exit[i] == "up":
                            if close[j] < upper[j] != 0:
                                bbe_arr.append("Yes")
                            else:
                                bbe_arr.append("No")
                        elif value_exit[i] == "mb":
                            if close[j] < middle[j] != 0:
                                bbe_arr.append("Yes")
                            else:
                                bbe_arr.append("No")
                    data_reset[col] = bbe_arr
                    col += 'a'

            elif indicator_exit[i] == "macd":
                macd_line, signal_line, macd_histogram = data_indic.MACD(close, 12, 26, 9)
                if parameter_exit[i] == "crossover":
                    for c in range(26):
                        macde_arr.append("No")
                    for j in range(26, len(signal_line)):
                        if value_exit[i] == "sig":
                            if macd_line[j - 1] <= signal_line[j] < macd_line[j]:
                                macde_arr.append("Yes")
                            else:
                                macde_arr.append("No")
                        elif value_exit[i] == "zero":
                            if macd_line[j - 1] <= 0 < macd_line[j]:
                                macde_arr.append("Yes")
                            else:
                                macde_arr.append("No")
                    data_reset[col] = macde_arr
                    col += 'a'

                elif parameter_exit[i] == "crossunder":
                    for c in range(26):
                        macde_arr.append("No")
                    for j in range(26, len(signal_line)):
                        if value_exit[i] == "sig":
                            if macd_line[j - 1] >= signal_line[j] > macd_line[j]:
                                macde_arr.append("Yes")
                            else:
                                macde_arr.append("No")
                        elif value_exit[i] == "zero":
                            if macd_line[j - 1] >= 0 > macd_line[j]:
                                macde_arr.append("Yes")
                            else:
                                macde_arr.append("No")
                    data_reset[col] = macde_arr
                    col += 'a'

                elif parameter_exit[i] == "above":
                    for c in range(26):
                        macde_arr.append("No")
                    for j in range(26, len(signal_line)):
                        if value_exit[i] == "sig":
                            if signal_line[j] < macd_line[j]:
                                macde_arr.append("Yes")
                            else:
                                macde_arr.append("No")
                        elif value_exit[i] == "zero":
                            if 0 < macd_line[j]:
                                macde_arr.append("Yes")
                            else:
                                macde_arr.append("No")
                    data_reset[col] = macde_arr
                    col += 'a'

                elif parameter_exit[i] == "below":
                    for c in range(26):
                        macde_arr.append("No")
                    for j in range(26, len(signal_line)):
                        if value_exit[i] == "sig":
                            if signal_line[j] > macd_line[j]:
                                macde_arr.append("Yes")
                            else:
                                macde_arr.append("No")
                        elif value_exit[i] == "zero":
                            if 0 > macd_line[j]:
                                macde_arr.append("Yes")
                            else:
                                macde_arr.append("No")
                    data_reset[col] = macde_arr
                    col += 'a'

            elif indicator_exit[i] == "mfi":
                val = data_indic.MFI(high, low, close, volume, 14)  # aiyan t
                if parameter_exit[i] == "crossover":
                    for c in range(14):
                        mfie_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] <= int(value_exit[i]) < val[j] != 0 and val[j - 1] != 0:
                            mfie_arr.append("Yes")
                        else:
                            mfie_arr.append("No")
                    data_reset[col] = mfie_arr
                    col += 'a'

                elif parameter_exit[i] == "crossunder":
                    for c in range(14):
                        mfie_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] >= int(value_exit[i]) > val[j] != 0 and val[j - 1] != 0:
                            mfie_arr.append("Yes")
                        else:
                            mfie_arr.append("No")
                    data_reset[col] = mfie_arr
                    col += 'a'

                elif parameter_exit[i] == "above":
                    for c in range(14):
                        mfie_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] > int(value_exit[i]):
                            mfie_arr.append("Yes")
                        else:
                            mfie_arr.append("No")
                    data_reset[col] = mfie_arr
                    col += 'a'

                elif parameter_exit[i] == "below":
                    for c in range(14):
                        mfie_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] < int(value_exit[i]):
                            mfie_arr.append("Yes")
                        else:
                            mfie_arr.append("No")
                    data_reset[col] = mfie_arr
                    col += 'a'

            elif indicator_exit[i] == "roc":
                val = data_indic.ROC(close, 9)  # aiyan t
                if parameter_exit[i] == "crossover":
                    for c in range(14):
                        roce_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] <= int(value_exit[i]) < val[j] != 0 and val[j - 1] != 0:
                            roce_arr.append("Yes")
                        else:
                            roce_arr.append("No")
                    data_reset[col] = roce_arr
                    col += 'a'

                elif parameter_exit[i] == "crossunder":
                    for c in range(14):
                        roce_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] >= int(value_exit[i]) > val[j] != 0 and val[j - 1] != 0:
                            roce_arr.append("Yes")
                        else:
                            roce_arr.append("No")
                    data_reset[col] = roce_arr
                    col += 'a'

                elif parameter_exit[i] == "above":
                    for c in range(14):
                        roce_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] > int(value_exit[i]):
                            roce_arr.append("Yes")
                        else:
                            roce_arr.append("No")
                    data_reset[col] = roce_arr
                    col += 'a'

                elif parameter_exit[i] == "below":
                    for c in range(14):
                        roce_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] < int(value_exit[i]):
                            roce_arr.append("Yes")
                        else:
                            roce_arr.append("No")
                    data_reset[col] = roce_arr
                    col += 'a'

            elif indicator_exit[i] == "srsi":
                k_line, d_line = data_indic.S_RSI(close, 14, 3, 3, 14)
                if parameter_exit[i] == "crossover":
                    for c in range(14):
                        srsie_arr.append("No")
                    for j in range(14, len(d_line)):
                        if value_exit[i] == "slow":
                            if k_line[j - 1] <= d_line[j] < k_line[j]:
                                srsie_arr.append("Yes")
                            else:
                                srsie_arr.append("No")
                        else:
                            if k_line[j - 1] <= int(value_exit[i]) < k_line[j]:
                                srsie_arr.append("Yes")
                            else:
                                srsie_arr.append("No")
                    data_reset[col] = srsie_arr
                    col += 'a'
                elif parameter_exit[i] == "crossunder":
                    for c in range(14):
                        srsie_arr.append("No")
                    for j in range(14, len(d_line)):
                        if value_exit[i] == "slow":
                            if k_line[j - 1] >= d_line[j] > k_line[j]:
                                srsie_arr.append("Yes")
                            else:
                                srsie_arr.append("No")
                        else:
                            if k_line[j - 1] >= int(value_exit[i]) > k_line[j]:
                                srsie_arr.append("Yes")
                            else:
                                srsie_arr.append("No")
                    data_reset[col] = srsie_arr
                    col += 'a'

                elif parameter_exit[i] == "above":
                    for c in range(14):
                        srsie_arr.append("No")
                    for j in range(14, len(d_line)):
                        if value_exit[i] == "slow":
                            if d_line[j] < k_line[j]:
                                srsie_arr.append("Yes")
                            else:
                                srsie_arr.append("No")
                        else:
                            if int(value_exit[i]) < k_line[j]:
                                srsie_arr.append("Yes")
                            else:
                                srsie_arr.append("No")
                    data_reset[col] = srsie_arr
                    col += 'a'

                elif parameter_exit[i] == "below":
                    for c in range(14):
                        srsie_arr.append("No")
                    for j in range(14, len(d_line)):
                        if value_exit[i] == "slow":
                            if d_line[j] > k_line[j]:
                                srsie_arr.append("Yes")
                            else:
                                srsie_arr.append("No")
                        else:
                            if int(value_exit[i]) > k_line[j]:
                                srsie_arr.append("Yes")
                            else:
                                srsie_arr.append("No")
                    data_reset[col] = srsie_arr
                    col += 'a'

            elif indicator_exit[i] == "wil":
                val = data_indic.WILLIAM_R(close, 14, high, low)  # aiyan t
                if parameter_exit[i] == "crossover":
                    for c in range(14):
                        wile_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] <= int(value_exit[i]) < val[j] != 0 and val[j - 1] != 0:
                            wile_arr.append("Yes")
                        else:
                            wile_arr.append("No")
                    data_reset[col] = wile_arr
                    col += 'a'

                elif parameter_exit[i] == "crossunder":
                    for c in range(14):
                        wile_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j - 1] >= int(value_exit[i]) > val[j] != 0 and val[j - 1] != 0:
                            wile_arr.append("Yes")
                        else:
                            wile_arr.append("No")
                    data_reset[col] = wile_arr
                    col += 'a'

                elif parameter_exit[i] == "above":
                    for c in range(14):
                        wile_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] > int(value_exit[i]):
                            wile_arr.append("Yes")
                        else:
                            wile_arr.append("No")
                    data_reset[col] = wile_arr
                    col += 'a'

                elif parameter_exit[i] == "below":
                    for c in range(14):
                        wile_arr.append("No")
                    for j in range(14, len(val)):
                        if val[j] < int(value_exit[i]):
                            wile_arr.append("Yes")
                        else:
                            wile_arr.append("No")
                    data_reset[col] = wile_arr
                    col += 'a'

            elif indicator_exit[i] == "sma10" or indicator_exit[i] == "sma20" or indicator_exit[i] == "sma50" or \
                    indicator_exit[i] == "sma100" or indicator_exit[i] == "sma200" or (
                    indicator_exit[i] == "clo" and (
                    value_exit[i] == "sma10" or value_exit[i] == "sma20" or value_exit[i] == "sma50" or value_exit[
                i] == "sma100" or value_exit[i] == "sma200")):
                val2 = []
                val = []
                if value_exit[i] == "sma20":
                    val2 = data_indic.SMA(close, 20)
                elif value_exit[i] == "sma50":
                    val2 = data_indic.SMA(close, 50)
                elif value_exit[i] == "sma100":
                    val2 = data_indic.SMA(close, 100)
                elif value_exit[i] == "sma200":
                    val2 = data_indic.SMA(close, 200)
                elif value_exit[i] == "sma10":
                    val2 = data_indic.SMA(close, 10)

                if indicator_exit[i] == "sma20":
                    val = data_indic.SMA(close, 20)
                elif indicator_exit[i] == "sma50":
                    val = data_indic.SMA(close, 50)
                elif indicator_exit[i] == "sma100":
                    val = data_indic.SMA(close, 100)
                elif indicator_exit[i] == "sma200":
                    val = data_indic.SMA(close, 200)
                elif indicator_exit[i] == "sma10":
                    val = data_indic.SMA(close, 10)
                elif indicator_exit[i] == "clo":
                    val = close

                ty = int(value_exit[i].split("a")[1])

                if parameter_exit[i] == "crossover":
                    for c in range(ty):
                        smae_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val[j - 1] < val2[j] < val[j]:
                            smae_arr.append("Yes")
                        else:
                            smae_arr.append("No")
                    data_reset[col] = smae_arr
                    col += 'a'
                elif parameter_exit[i] == "crossunder":
                    for c in range(ty):
                        smae_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val[j - 1] > val2[j] > val[j]:
                            smae_arr.append("Yes")
                        else:
                            smae_arr.append("No")
                    data_reset[col] = smae_arr
                    col += 'a'
                elif parameter_exit[i] == "above":
                    for c in range(ty):
                        smae_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val2[j] < val[j]:
                            smae_arr.append("Yes")
                        else:
                            smae_arr.append("No")
                    data_reset[col] = smae_arr
                    col += 'a'
                elif parameter_exit[i] == "below":
                    for c in range(ty):
                        smae_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val2[j] > val[j]:
                            smae_arr.append("Yes")
                        else:
                            smae_arr.append("No")
                    data_reset[col] = smae_arr
                    col += 'a'

            elif indicator_exit[i] == "ema10" or indicator_exit[i] == "ema20" or indicator_exit[i] == "ema50" or \
                    indicator_exit[i] == "ema100" or indicator_exit[i] == "ema200" or (indicator_exit[i] == "clo" and (
                    value_exit[i] == "ema10" or value_exit[i] == "ema20" or value_exit[i] == "ema50" or value_exit[
                i] == "ema100" or value_exit[i] == "ema200")):
                val2 = []
                val = []
                if value_exit[i] == "ema20":
                    val2 = data_indic.EMA(close, 20)
                elif value_exit[i] == "ema50":
                    val2 = data_indic.EMA(close, 50)
                elif value_exit[i] == "ema100":
                    val2 = data_indic.EMA(close, 100)
                elif value_exit[i] == "ema200":
                    val2 = data_indic.EMA(close, 200)
                elif value_exit[i] == "ema10":
                    val2 = data_indic.EMA(close, 10)

                if indicator_exit[i] == "ema20":
                    val = data_indic.EMA(close, 20)
                elif indicator_exit[i] == "ema50":
                    val = data_indic.EMA(close, 50)
                elif indicator_exit[i] == "ema100":
                    val = data_indic.EMA(close, 100)
                elif indicator_exit[i] == "ema200":
                    val = data_indic.EMA(close, 200)
                elif indicator_exit[i] == "ema10":
                    val = data_indic.EMA(close, 10)
                elif indicator_exit[i] == "clo":
                    val = close

                ty = int(value_exit[i].split("a")[1])

                if parameter_exit[i] == "crossover":
                    for c in range(ty):
                        emae_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val[j - 1] < val2[j] < val[j]:
                            emae_arr.append("Yes")
                        else:
                            emae_arr.append("No")
                    data_reset[col] = emae_arr
                    col += 'a'
                elif parameter_exit[i] == "crossunder":
                    for c in range(ty):
                        emae_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val[j - 1] > val2[j] > val[j]:
                            emae_arr.append("Yes")
                        else:
                            emae_arr.append("No")
                    data_reset[col] = emae_arr
                    col += 'a'
                elif parameter_exit[i] == "above":
                    for c in range(ty):
                        emae_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val2[j] < val[j]:
                            emae_arr.append("Yes")
                        else:
                            emae_arr.append("No")
                    data_reset[col] = emae_arr
                    col += 'a'
                elif parameter_exit[i] == "below":
                    for c in range(ty):
                        emae_arr.append("No")
                    for j in range(ty, len(val2)):
                        if val2[j] > val[j]:
                            emae_arr.append("Yes")
                        else:
                            emae_arr.append("No")
                    data_reset[col] = emae_arr
                    col += 'a'

            elif indicator_exit[i] == "clo" and (
                    value_exit[i] == "pp" or value_exit[i] == "s1" or value_exit[i] == "s2" or value_exit[i] == "s3" or
                    value_exit[i] == "r1" or value_exit[i] == "r2" or value_exit[i] == "r3"):

                pp, s1, s2, s3, r1, r2, r3 = data_indic.pivot_points(close, high, low, date)
                val = close

                for jj in range(abs(len(close)-len(pp))):
                    pp.insert(0, 0)
                    s1.insert(0, 0)
                    s2.insert(0, 0)
                    s3.insert(0, 0)
                    r1.insert(0, 0)
                    r2.insert(0, 0)
                    r3.insert(0, 0)

                ppe_arr.append("No")
                if parameter_exit[i] == "crossover":
                    if value_exit[i] == "pp":
                        for j in range(1, len(pp)):
                            if val[j - 1] < pp[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "s1":
                        for j in range(1, len(pp)):
                            if val[j - 1] < s1[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "s2":
                        for j in range(1, len(pp)):
                            if val[j - 1] < s2[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "s3":
                        for j in range(1, len(pp)):
                            if val[j - 1] < s3[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "r1":
                        for j in range(1, len(pp)):
                            if val[j - 1] < r1[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "r2":
                        for j in range(1, len(pp)):
                            if val[j - 1] < r1[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "r3":
                        for j in range(1, len(pp)):
                            if val[j - 1] < r3[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'

                elif parameter_exit[i] == "crossunder":
                    if value_exit[i] == "pp":
                        for j in range(1, len(pp)):
                            if val[j - 1] > pp[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "s1":
                        for j in range(1, len(pp)):
                            if val[j - 1] > s1[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "s2":
                        for j in range(1, len(pp)):
                            if val[j - 1] > s2[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "s3":
                        for j in range(1, len(pp)):
                            if val[j - 1] > s3[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "r1":
                        print(len(r1))
                        print(len(val))
                        for j in range(1, len(pp)):

                            if val[j - 1] > r1[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "r2":
                        for j in range(1, len(pp)):
                            if val[j - 1] > r2[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "r3":
                        for j in range(1, len(pp)):
                            if val[j - 1] > r3[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'

                elif parameter_exit[i] == "above":
                    if value_exit[i] == "pp":
                        for j in range(1, len(pp)):
                            if pp[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "s1":
                        for j in range(1, len(pp)):
                            if s1[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "s2":
                        for j in range(1, len(pp)):
                            if s2[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "s3":
                        for j in range(1, len(pp)):
                            if s3[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "r1":
                        for j in range(1, len(pp)):
                            if r1[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "r2":
                        for j in range(1, len(pp)):
                            if r2[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "r3":
                        for j in range(1, len(pp)):
                            if r3[j] < val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'

                elif parameter_exit[i] == "below":
                    if value_exit[i] == "pp":
                        for j in range(1, len(pp)):
                            if pp[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "s1":
                        for j in range(1, len(pp)):
                            if s1[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "s2":
                        for j in range(1, len(pp)):
                            if s2[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "s3":
                        for j in range(1, len(pp)):
                            if s3[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "r1":
                        for j in range(1, len(pp)):
                            if r1[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "r2":
                        for j in range(1, len(pp)):
                            if r2[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'
                    elif value_exit[i] == "r3":
                        for j in range(1, len(pp)):
                            if r3[j] > val[j]:
                                ppe_arr.append("Yes")
                            else:
                                ppe_arr.append("No")
                        data_reset[col] = ppe_arr
                        col += 'a'

            elif indicator_exit[i] == "cl" or indicator_exit[i] == "ls" or (indicator_exit[i] == "clo" and (
                    value_exit[i] == "cl" or value_exit[i] == "bl" or value_exit[i] == "ls" or value_exit[i] == "gc" or
                    value_exit[i] == "rc")):
                val2 = []
                val = []
                cl, bl, sa, sb, ls = data_indic.Icloud(high, low, close, 9, 26, 52, 26)
                if value_exit[i] == "cl":
                    val2 = cl
                elif value_exit[i] == "bl":
                    val2 = bl
                elif value_exit[i] == "ls":
                    val2 = ls
                elif value_exit[i] == "gc":
                    for jk in range(len(sa)):
                        val2.append(sa[jk]-sb[jk])
                elif value_exit[i] == "rc":
                    for jk in range(len(sa)):
                        val2.append(sb[jk] - sa[jk])

                if indicator_exit[i] == "cl":
                    val = cl
                elif indicator_exit[i] == "ls":
                    val = ls
                elif indicator_exit[i] == "clo":
                    val = close
                ice_arr.append("No")
                if parameter_exit[i] == "crossover":
                    for j in range(1, len(val)):
                        if val[j - 1] < val2[j] < val[j]:
                            ice_arr.append("Yes")
                        else:
                            ice_arr.append("No")
                    data_reset[col] = ice_arr
                    col += 'a'
                elif parameter_exit[i] == "crossunder":
                    for j in range(1, len(val)):
                        if val[j - 1] > val2[j] > val[j]:
                            ice_arr.append("Yes")
                        else:
                            ice_arr.append("No")
                    data_reset[col] = ice_arr
                    col += 'a'
                elif parameter_exit[i] == "above":
                    for j in range(1, len(val2)):
                        if val2[j] < val[j] and val2[j] != -1 and val2[j] != 0:
                            ice_arr.append("Yes")
                        else:
                            ice_arr.append("No")
                    data_reset[col] = ice_arr
                    col += 'a'
                elif parameter_exit[i] == "below":
                    for j in range(1, len(val2)):
                        if val2[j] > val[j]:
                            ice_arr.append("Yes")
                        else:
                            ice_arr.append("No")
                    data_reset[col] = ice_arr
                    col += 'a'

        entry_dt_points = []
        exit_dt_points = []
        entry_date_points = []
        exit_date_points = []
        entry_close_points = []
        exit_close_points = []
        for i in range(len(close)):
            ce = 0
            if data_reset.iloc[i, 7] == "Yes":
                for j in range(count_entry):
                    if data_reset.iloc[i, 6 + 1 + j] == "Yes":
                        ce += 1
                    if ce == count_entry:
                        entry_dt_points.append(dt[i])
                        entry_date_points.append(date[i])
                        entry_close_points.append(close[i])

            cex = 0
            if data_reset.iloc[i, 6 + count_entry + 1] == "Yes":
                for j in range(count_exit):
                    if data_reset.iloc[i, 6 + count_entry + 1 + j] == "Yes":
                        cex += 1
                    if cex == count_exit:
                        exit_dt_points.append(dt[i])
                        exit_date_points.append(date[i])
                        exit_close_points.append(close[i])

        total = []
        ref = 0
        stoploss_choice = ""
        if stoploss == 1000:
            stoploss_choice = "No"
        else:
            stoploss_choice = "Yes"

        print(entry_date_points)
        print(entry_close_points)
        # data_reset['mc'] = macd_line
        # data_reset['sg'] = signal_line
        # data_reset.to_csv('dekle.csv')
        print(exit_date_points)
        print(exit_close_points)
        date_front = []
        entry_front = []
        exit_front = []
        dateex_front = []
        whatto = 0
        got = 0
        if stoploss_choice == "Yes":
            if len(entry_date_points) == 0 or len(exit_date_points) == 0:
                whatto = 1
                print(" No Match for this Strategy")
            else:
                for i in range(len(entry_date_points)):
                    if entry_dt_points[i] > ref:
                        price_entry = entry_close_points[i]
                        print(entry_date_points[i])
                        stpls_dt_entry = entry_dt_points[i]
                        for j in range(len(exit_date_points)):
                            if exit_dt_points[j] > entry_dt_points[i]:
                                got = 1
                                stpls_dt_exit = exit_dt_points[j]
                                price_exit = exit_close_points[j]
                                stp_start = 0
                                stp_end = 0
                                for xx in range(len(close)):
                                    if dt[xx] == stpls_dt_entry:
                                        stp_start = xx

                                    if dt[xx] == stpls_dt_exit:
                                        stp_end = xx

                                flag = 0
                                for t in range(stp_start, stp_end + 1):
                                    if close[t] <= close[stp_start] * (100 - stoploss) / 100:
                                        if long_short == "golong":
                                            if round(((close[t] - price_entry) / close[t]) * 100, 2) < -stoploss:
                                                total.append(-stoploss)
                                            else:
                                                total.append(round(((close[t] - price_entry) / close[t]) * 100, 2))
                                        else:
                                            if round(((close[t] - price_entry) / close[t]) * 100, 2) < -stoploss:
                                                total.append(-stoploss)
                                            else:
                                                total.append(-round(((close[t] - price_entry) / close[t]) * 100, 2))
                                        print(date[t])
                                        ref = dt[t]
                                        date_front.append(entry_date_points[i])
                                        dateex_front.append(date[t])
                                        entry_front.append(round(price_entry, 2))
                                        exit_front.append(round(close[t], 2))
                                        flag = 1
                                        break
                                if flag == 0:
                                    if long_short == "golong":
                                        if round(((price_exit - price_entry) / price_exit) * 100, 2) < -stoploss:
                                            total.append(-stoploss)
                                        else:
                                            total.append(round(((price_exit - price_entry) / price_exit) * 100, 2))
                                    else:
                                        if round(((price_exit - price_entry) / price_exit) * 100, 2) < -stoploss:
                                            total.append(-stoploss)
                                        else:
                                            total.append(-round(((price_exit - price_entry) / price_exit) * 100, 2))
                                    ref = exit_dt_points[j]
                                    print(exit_date_points[j])
                                    date_front.append(entry_date_points[i])
                                    dateex_front.append(exit_date_points[j])
                                    entry_front.append(round(price_entry, 2))
                                    exit_front.append(round(price_exit, 2))
                                    break
                                break
                # print(f'Your Profit/Loss is {statistics.mean(total):.2f} %')

        else:
            if len(entry_date_points) == 0 or len(exit_date_points) == 0:
                whatto = 1
                print(" No Match for this Strategy")
            else:
                for p in range(len(entry_date_points)):
                    if entry_dt_points[p] > ref:
                        price_entry = entry_close_points[p]
                        for j in range(len(exit_date_points)):
                            if exit_dt_points[j] > entry_dt_points[p]:
                                got = 1
                                price_exit = exit_close_points[j]
                                if long_short == "golong":
                                    total.append(round(((price_exit - price_entry) / price_exit) * 100, 2))
                                else:
                                    total.append(-round(((price_exit - price_entry) / price_exit) * 100, 2))
                                date_front.append(entry_date_points[p])
                                dateex_front.append(exit_date_points[j])
                                entry_front.append(round(price_entry, 2))
                                exit_front.append(round(price_exit, 2))
                                ref = exit_dt_points[j]
                                break
                # print(f'Your Profit/Loss is {statistics.mean(total):.2f} %')
        resultt = 0
        nothing = ""
        flagr = 0
        if whatto == 1 or got == 0:
            nothing = " No Match for this Strategy"
            flagr = 1
        else:
            resultt = float(f'{sum(total):.2f}')

        zipdata = zip(total, entry_front, exit_front, date_front, dateex_front, range(len(date_front) + len(dateex_front)))
        dictb = {'datta': dataa, 'resultt': resultt, 'percentage': total, 'entry': entry_front, 'exit': exit_front,
                 'entry_date': date_front, 'exit_date': dateex_front, 'whatto': whatto, 'zipdata': zipdata,
                 'nothing': nothing, 'flagr': flagr}

        wb = xl.load_workbook('login/users.xlsx')
        ip = get_client_ip(request)
        sheet = wb["Sheet1"]
        for i in range(2, sheet.max_row + 1):
            if (ip == sheet.cell(i, 3).value):
                if (sheet.cell(i, 4).value == "yes"):
                    dictb["email"] = sheet.cell(i, 1).value

        return render(request, 'backtest_detail.html', dictb)
    except Exception as e:
        wb = xl.load_workbook("errors.xlsx")
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.now()
        sheet1.cell(sheet1.max_row , 4).value = request.POST.get('backe', 'default')
        wb.save("errors.xlsx")
        return render(request, "oops.html")

def index(request):
    try:
        wb = xl.load_workbook('login/users.xlsx')
        ip = get_client_ip(request)
        sheet = wb["Sheet1"]
        dictb = {}
        for i in range(2, sheet.max_row + 1):
            if (ip == sheet.cell(i, 3).value):
                if (sheet.cell(i, 4).value == "yes"):
                    dictb["email"] = sheet.cell(i, 1).value
                    return render(request, 'backtestdata.html', dictb)
        response = redirect('/login')
        return response
    except Exception as e:
        wb = xl.load_workbook("errors.xlsx")
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.now()
        wb.save("errors.xlsx")
        return render(request, "oops.html")