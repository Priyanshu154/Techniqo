from django.shortcuts import render
import openpyxl as xl
import os
import requests
from bs4 import BeautifulSoup
import pandas_datareader.data as web
from datetime import datetime
from matplotlib.dates import date2num
from . import data_indis

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip
def shareholding(request):
    all_detail = request.GET.get('tickerdata')
    print(all_detail)
    sarr = all_detail.split(",")
    check = all_detail.split(",")[3]
    sname = all_detail.split(",")[0]
    nse_ticker = ""
    bse_ticker = ""
    print(sarr)
    if len(sarr) == 4:
        nse_ticker = all_detail.split(",")[1]
        bse_ticker = all_detail.split(",")[2]
    else:
        bse_ticker = all_detail.split(", ")[1]
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'market_data_20.xlsx')
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['sheet']
    url3 = ""
    urls4 = ""
    urls5 = ""
    for i in range(2, sheet.max_row + 1):
        if sheet.cell(i, 2).value == sname:
            url3 = sheet.cell(i, 1).value
            urls4 = sheet.cell(i, 9).value
            urls5 = sheet.cell(i, 10).value
    urls3 = url3[:4] + "s" + url3[4:]

    if check == "shareholding":
        r3 = requests.get(urls3)
        htmlcontent3 = r3.content
        soup = BeautifulSoup(htmlcontent3, 'html.parser')
        headd = []
        boddy = []
        boddy2 = []
        boddy3 = []
        boddy4 = []
        boddy5 = []
        for i in range(0,5):
            headd.append(soup.find_all("thead", {"count": "10"})[0].find_all("th")[i].get_text().strip())
        for i in range(0,9):
            boddy.append(soup.find_all("table", {"class": "sharePriceTotalCal"})[0].find("tbody").find_all("tr")[i].find_all("td")[0].get_text().strip())
            boddy2.append(soup.find_all("table", {"class": "sharePriceTotalCal"})[0].find("tbody").find_all("tr")[i].find_all("td")[1].get_text().strip())
            boddy3.append(soup.find_all("table", {"class": "sharePriceTotalCal"})[0].find("tbody").find_all("tr")[i].find_all("td")[2].get_text().strip())
            boddy4.append(soup.find_all("table", {"class": "sharePriceTotalCal"})[0].find("tbody").find_all("tr")[i].find_all("td")[3].get_text().strip())
            boddy5.append(soup.find_all("table", {"class": "sharePriceTotalCal"})[0].find("tbody").find_all("tr")[i].find_all("td")[4].get_text().strip())
        zipb = zip(boddy,boddy2,boddy3,boddy4,boddy5)

        dictts = {'stockn': sname, 'typp': "Shareholdings", 'headd': headd, 'zipb': zipb, 'nse_ticker': nse_ticker,
                  "bse_ticker": bse_ticker, 'flag': 1}
        wb = xl.load_workbook('login/users.xlsx')
        ip = get_client_ip(request)
        sheet = wb["Sheet1"]
        for i in range(2, sheet.max_row + 1):
            if (ip == sheet.cell(i, 3).value):
                if (sheet.cell(i, 4).value == "yes"):
                    print("matched")
                    dictts["email"] = sheet.cell(i, 1).value
        return render(request, 'shareholding.html', dictts)

    elif check == "valuation":
        r3 = requests.get(urls3)
        htmlcontent3 = r3.content
        soup = BeautifulSoup(htmlcontent3, 'html.parser')
        stdn = []
        stdv = []
        conn = []
        conv = []
        ct = 1
        for i in range(3):
            for j in range(4):
                if ct < 12:
                    stdn.append(soup.find_all("div", {"id": "standalone_valuation"})[0].find_all("ul", {"class": "clearfix val_listinner"})[i].find_all("li", {"class": "clearfix"})[j].find_all("div")[0].get_text().strip())
                    stdv.append(soup.find_all("div", {"id": "standalone_valuation"})[0].find_all("ul", {"class": "clearfix val_listinner"})[i].find_all("li", {"class": "clearfix"})[j].find_all("div")[1].get_text().strip())
                    conn.append(soup.find_all("div", {"id": "consolidated_valuation"})[0].find_all("ul", {"class": "clearfix val_listinner"})[i].find_all("li", {"class": "clearfix"})[j].find_all("div")[0].get_text().strip())
                    conv.append(soup.find_all("div", {"id": "consolidated_valuation"})[0].find_all("ul", {"class": "clearfix val_listinner"})[i].find_all("li", {"class": "clearfix"})[j].find_all("div")[1].get_text().strip())
                    ct += 1
        zipa = zip(stdn,stdv)
        zipb = zip(conn,conv)
        dictts = {'zipa': zipa, 'zipb': zipb,'nse_ticker': nse_ticker,
                  "bse_ticker": bse_ticker, "stockn": sname, 'typp': "Valuation Ratios", 'flag': 2}

        wb = xl.load_workbook('login/users.xlsx')
        ip = get_client_ip(request)
        sheet = wb["Sheet1"]
        for i in range(2, sheet.max_row + 1):
            if (ip == sheet.cell(i, 3).value):
                if (sheet.cell(i, 4).value == "yes"):
                    print("matched")
                    dictts["email"] = sheet.cell(i, 1).value
        return render(request, 'shareholding.html', dictts)

    elif check == "income":
        r3 = requests.get(urls3)
        htmlcontent3 = r3.content
        soup = BeautifulSoup(htmlcontent3, 'html.parser')
        ihead = [" "]
        ibody = []
        try:
            for i in range(1, 6):
                    ihead.append(soup.find_all("div", {"id": "IncomeStatement"})[0].find("thead").find_all("th")[i].get_text().strip())
            for i in range(9):
                icbody = []
                for j in range(6):
                    icbody.append(soup.find_all("div", {"id": "IncomeStatement"})[0].find("tbody").find_all("tr")[i].find_all("td")[j].get_text().strip())
                ibody.append(icbody)
        except:
            try:
                for i in range(1, 6):
                    ihead.append(soup.find_all("div", {"id": "SIncomeStatement"})[0].find("thead").find_all("th")[i].get_text().strip())
                for i in range(9):
                    icbody = []
                    for j in range(6):
                        icbody.append(
                            soup.find_all("div", {"id": "SIncomeStatement"})[0].find("tbody").find_all("tr")[i].find_all("td")[j].get_text().strip())
                    ibody.append(icbody)
            except:
                dictts = {'nse_ticker': nse_ticker,
                  "bse_ticker": bse_ticker, "stockn": sname,}

                wb = xl.load_workbook('login/users.xlsx')
                ip = get_client_ip(request)
                sheet = wb["Sheet1"]
                for i in range(2, sheet.max_row + 1):
                    if (ip == sheet.cell(i, 3).value):
                        if (sheet.cell(i, 4).value == "yes"):
                            print("matched")
                            dictts["email"] = sheet.cell(i, 1).value
                return render(request, 'nodata.html',dictts)

        dictts = {'ihead': ihead,'ibody': ibody, 'nse_ticker': nse_ticker,
                  "bse_ticker": bse_ticker, "stockn": sname, 'typp': "Income Statement", 'flag': 3}

        wb = xl.load_workbook('login/users.xlsx')
        ip = get_client_ip(request)
        sheet = wb["Sheet1"]
        for i in range(2, sheet.max_row + 1):
            if (ip == sheet.cell(i, 3).value):
                if (sheet.cell(i, 4).value == "yes"):
                    print("matched")
                    dictts["email"] = sheet.cell(i, 1).value
        return render(request, 'shareholding.html', dictts)

    elif check == "chart":
        if nse_ticker != "":
            dictts = { 'nse_ticker': nse_ticker,
                      "bse_ticker": bse_ticker, "stockn": sname, 'typp': "Chart", 'flag': 5}

            wb = xl.load_workbook('login/users.xlsx')
            ip = get_client_ip(request)
            sheet = wb["Sheet1"]
            for i in range(2, sheet.max_row + 1):
                if (ip == sheet.cell(i, 3).value):
                    if (sheet.cell(i, 4).value == "yes"):
                        print("matched")
                        dictts["email"] = sheet.cell(i, 1).value
            return render(request, 'shareholding.html', dictts)
        else:
            dictts = {'nse_ticker': nse_ticker,
                      "bse_ticker": bse_ticker, "stockn": sname }

            wb = xl.load_workbook('login/users.xlsx')
            ip = get_client_ip(request)
            sheet = wb["Sheet1"]
            for i in range(2, sheet.max_row + 1):
                if (ip == sheet.cell(i, 3).value):
                    if (sheet.cell(i, 4).value == "yes"):
                        print("matched")
                        dictts["email"] = sheet.cell(i, 1).value
            return render(request, 'nodata.html', dictts)

    elif check == "balance":
        try:

            print(urls5)
            r4 = requests.get(urls5)
            htmlcontent4 = r4.content
            soup = BeautifulSoup(htmlcontent4, 'html.parser')
            bhead = ["Financial Year"]
            bbody = []
            bbody1 = []
            bbody2 = []
            bbody3 = []
            bbody4 = []
            bbody5 = []

            for i in range(1,6):
                bhead.append(soup.find_all("table")[3].find_all("th")[i].get_text().strip())

            for i in range(16):
                bbody.append(soup.find_all("table")[3].find("tbody").find_all("tr")[i].find("th").get_text().strip())
                j = 0
                bbody1.append(soup.find_all("table")[3].find("tbody").find_all("tr")[i].find_all("td")[0].get_text().strip())
                bbody2.append(soup.find_all("table")[3].find("tbody").find_all("tr")[i].find_all("td")[1].get_text().strip())
                bbody3.append(soup.find_all("table")[3].find("tbody").find_all("tr")[i].find_all("td")[2].get_text().strip())
                bbody4.append(soup.find_all("table")[3].find("tbody").find_all("tr")[i].find_all("td")[3].get_text().strip())
                bbody5.append(soup.find_all("table")[3].find("tbody").find_all("tr")[i].find_all("td")[4].get_text().strip())


            zipbb = zip(bbody,bbody1,bbody2,bbody3,bbody4,bbody5)
            dictts = {'nse_ticker': nse_ticker, 'zipba':zipbb , 'bheadb': bhead,'typp': "Balance Sheet (in Cr.)", 'flag': 4,
                      "bse_ticker": bse_ticker, "stockn": sname}

            wb = xl.load_workbook('login/users.xlsx')
            ip = get_client_ip(request)
            sheet = wb["Sheet1"]
            for i in range(2, sheet.max_row + 1):
                if (ip == sheet.cell(i, 3).value):
                    if (sheet.cell(i, 4).value == "yes"):
                        print("matched")
                        dictts["email"] = sheet.cell(i, 1).value
            return render(request, 'shareholding.html', dictts)

        except:
            dictts = {'nse_ticker': nse_ticker,
                      "bse_ticker": bse_ticker, "stockn": sname, }

            wb = xl.load_workbook('login/users.xlsx')
            ip = get_client_ip(request)
            sheet = wb["Sheet1"]
            for i in range(2, sheet.max_row + 1):
                if (ip == sheet.cell(i, 3).value):
                    if (sheet.cell(i, 4).value == "yes"):
                        print("matched")
                        dictts["email"] = sheet.cell(i, 1).value
            return render(request, 'nodata.html', dictts)


# Create your views here.
def index(request):
    global soup2
    name = request.GET.get('stock_namee', 'no')
    url = ""
    tick = ""
    url2 = ""
    if name == 'no':

        wb = xl.load_workbook('login/users.xlsx')
        ip = get_client_ip(request)
        sheet = wb["Sheet1"]
        dictts = {}
        for i in range(2, sheet.max_row + 1):
            if (ip == sheet.cell(i, 3).value):
                if (sheet.cell(i, 4).value == "yes"):
                    print("matched")
                    dictts["email"] = sheet.cell(i, 1).value
        return render(request, 'home.html', dictts)
    else:
        workpath = os.path.dirname(os.path.abspath(__file__))
        xx = os.path.join(workpath, 'market_data_20.xlsx')
        wb = xl.load_workbook(xx, data_only=True)
        sheet = wb['sheet']
        sarr = name.split(", ")
        stock_name = name.split(", ")[0]
        nse_ticker = ""
        bse_ticker = ""
        color = []
        bet = []
        if len(sarr) == 3:
            nse_ticker = name.split(", ")[1]
            bse_ticker = name.split(", ")[2]
        else:
            try:
                bse_ticker = name.split(", ")[1]
            except:

                wb = xl.load_workbook('login/users.xlsx')
                ip = get_client_ip(request)
                sheet = wb["Sheet1"]
                dictts = {}
                for i in range(2, sheet.max_row + 1):
                    if (ip == sheet.cell(i, 3).value):
                        if (sheet.cell(i, 4).value == "yes"):
                            print("matched")
                            dictts["email"] = sheet.cell(i, 1).value
                return render(request, 'wrong.html', dictts)
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(i, 2).value == stock_name:
                url = sheet.cell(i, 1).value
                tick = sheet.cell(i, 3).value
                url2 = sheet.cell(i, 8).value
                break
        urls = url[:4] + "s" + url[4:]
        r = requests.get(urls)
        htmlcontent = r.content
        soup = BeautifulSoup(htmlcontent, 'html.parser')
        price = float(soup.find_all('div', {'class': 'pcnsb'})[0].find_all('span')[0].string)
        change = soup.find_all('div', {'class': 'pcnsb'})[0].find_all('span')[2].get_text()
        if float(soup.find_all('div', {'class': 'pcnsb'})[0].find_all('span')[2].get_text().split(' ')[0]) < 0:
            color = ["danger", "fas fa-angle-down"]
        else:
            color = ["success", "fas fa-angle-up"]

        tech_count = 0
        senti = []
        senti_color = []
        flag = 0
        onlyma = ""
        onlyst = ""
        values = []
        try:
            start = datetime(2018, 1, 24)
            end = datetime(2020, 10, 4)

            data = web.DataReader(tick + '.NS', 'yahoo', start, end)
            # data.reset_index() will shift the Date from Header column to normal column you can print to check
            data_reset = data.reset_index()
            # This line is compulsory to make Date  column readable to python programme
            data_reset['date_ax'] = data_reset['Date'].apply(lambda date: date2num(date))
            # putting every column in an individual list
            close = data_reset['Close'].to_list()
            high = data_reset['High'].to_list()
            low = data_reset['Low'].to_list()
            openn = data_reset['Open'].to_list()
            date = data_reset['Date'].to_list()
            dt = data_reset['date_ax'].to_list()
            volume = data_reset['Volume'].to_list()

            rsi_val = round(data_indis.RSI(close, 14)[len(data_indis.RSI(close, 14)) - 1], 2)
            macd_val = round(data_indis.MACD(close, 12, 26, 9)[0][len(data_indis.MACD(close, 12, 26, 9)[0]) - 1], 2)
            sig_val = round(data_indis.MACD(close, 12, 26, 9)[1][len(data_indis.MACD(close, 12, 26, 9)[1]) - 1], 2)
            sma_9 = round(data_indis.SMA(close, 9)[len(data_indis.SMA(close, 9)) - 1], 2)
            sma_20 = round(data_indis.SMA(close, 20)[len(data_indis.SMA(close, 20)) - 1], 2)
            sma_50 = round(data_indis.SMA(close, 50)[len(data_indis.SMA(close, 50)) - 1], 2)
            sma_200 = round(data_indis.SMA(close, 200)[len(data_indis.SMA(close, 200)) - 1], 2)
            l = len(data_indis.pivot_points(close, high, low, date)[0])
            pp = round(data_indis.pivot_points(close, high, low, date)[0][l - 1], 2)
            xx = len(data_indis.S_RSI(close, 14, 3, 3, 14)[0])
            stc_k = round(data_indis.S_RSI(close, 14, 3, 3, 14)[0][xx - 1], 2)
            stc_d = round(data_indis.S_RSI(close, 14, 3, 3, 14)[1][len(data_indis.S_RSI(close, 14, 3, 3, 14)[1]) - 1],
                          2)
            roc = round(data_indis.ROC(close, 20)[len(data_indis.ROC(close, 20)) - 1], 2)
            mfi = round(
                data_indis.MFI(high, low, close, volume, 14)[len(data_indis.MFI(high, low, close, volume, 14)) - 1], 2)
            wil = round(data_indis.WILLIAM_R(close, 14, high, low)[len(data_indis.WILLIAM_R(close, 14, high, low)) - 1],
                        2)
            ic_con = round(data_indis.Icloud(high, low, close, 9, 26, 52, 26)[0][
                               len(data_indis.Icloud(high, low, close, 9, 26, 52, 26)[0]) - 1], 2)
            ic_base = round(data_indis.Icloud(high, low, close, 9, 26, 52, 26)[1][
                                len(data_indis.Icloud(high, low, close, 9, 26, 52, 26)[1]) - 1], 2)
            ic_spana = round(data_indis.Icloud(high, low, close, 9, 26, 52, 26)[2][
                                 len(data_indis.Icloud(high, low, close, 9, 26, 52, 26)[2]) - 26], 2)
            ic_spanb = round(data_indis.Icloud(high, low, close, 9, 26, 52, 26)[3][
                                 len(data_indis.Icloud(high, low, close, 9, 26, 52, 26)[3]) - 26], 2)
            values = [rsi_val, macd_val, sma_9, sma_20, sma_50, sma_200, "", "", "", pp, stc_k, roc, mfi, wil, ic_con]

            if rsi_val > 70:
                tech_count += 5
                senti.append("Overbought")
                senti_color.append("dark")
            elif 70 >= rsi_val > 30:
                tech_count += 10
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if macd_val > 0 and macd_val > sig_val:
                tech_count += 10
                senti.append("Bullish")
                onlyma = "Above sig. line"
                senti_color.append("success")
            elif 0 > macd_val > sig_val:
                tech_count += 5
                senti.append("Neutral")
                onlyma = "Above sig. line"
                senti_color.append("dark")
            elif 0 < macd_val < sig_val:
                tech_count += 5
                senti.append("Neutral")
                onlyma = "Below sig. line"
                senti_color.append("dark")
            else:
                senti.append("Bearish")
                onlyma = "Below sig. line"
                senti_color.append("danger")

            if price > sma_9:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")
            if price > sma_20:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")
            if price > sma_50:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")
            if price > sma_200:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")
            if sma_9 > sma_20:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
                bet.append("Above")
            else:
                senti.append("Bearish")
                senti_color.append("danger")
                bet.append("Below")
            if sma_20 > sma_50:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
                bet.append("Above")
            else:
                senti.append("Bearish")
                senti_color.append("danger")
                bet.append("Below")
            if sma_50 > sma_200:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
                bet.append("Above")
            else:
                senti.append("Bearish")
                senti_color.append("danger")
                bet.append("Below")

            if price > pp:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if stc_k > stc_d and stc_k > 70:
                tech_count += 5
                senti.append("Overbought")
                onlyst = "Above slow line"
                senti_color.append("dark")
            elif stc_k > stc_d and 30 <= stc_k <= 70:
                tech_count += 10
                senti.append("Bullish")
                onlyst = "Above slow line"
                senti_color.append("success")
            elif stc_d < stc_k <= 30:
                tech_count += 5
                senti.append("Neutral")
                onlyst = "Above slow line"
                senti_color.append("dark")
            else:
                senti.append("Bearish")
                onlyst = "Below slow line"
                senti_color.append("danger")

            if roc > 0:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if mfi > 70:
                tech_count += 5
                senti.append("Overbought")
                senti_color.append("dark")
            elif 70 >= mfi > 30:
                tech_count += 10
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if wil > -20:
                tech_count += 10
                senti.append("Bullish")
                senti_color.append("success")
            elif -20 >= wil >= -45:
                tech_count += 5
                senti.append("Neutral")
                senti_color.append("dark")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if ic_con > ic_base and ic_con > ic_spana > ic_spanb:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            elif ic_con > ic_base and ic_con > ic_spana <= ic_spanb:
                tech_count += 2.5
                senti.append("Neutral")
                senti_color.append("dark")
            else:
                senti.append("Bearish")
                senti_color.append("danger")
            flag = 1

        except:
            r2 = requests.get(url2)
            htmlcontent2 = r2.content
            soup2 = BeautifulSoup(htmlcontent2, 'html.parser')
            try:
                rsi_val = float(
                    soup2.find_all("table")[4].find_all("tbody")[1].find_all("tr")[0].find_all("td")[1].get_text())
            except:
                rsi_val = "--"
            try:
                macd_val = float(
                    soup2.find_all("table")[4].find_all("tbody")[1].find_all("tr")[1].find_all("td")[1].get_text())
            except:
                macd_val = "--"
            try:
                s = ""
                sma_10 = float(s.join(soup2.find_all("table")[2].find_all("tbody")[0].find_all("tr")[1].find_all("td")[
                                          1].get_text().split(',')))
            except:
                sma_10 = "--"
            try:
                s = ""
                sma_20 = float(s.join(soup2.find_all("table")[2].find_all("tbody")[0].find_all("tr")[2].find_all("td")[
                                          1].get_text().split(',')))
            except:
                sma_20 = "--"
            try:
                s = ""
                sma_50 = float(s.join(soup2.find_all("table")[2].find_all("tbody")[0].find_all("tr")[3].find_all("td")[
                                          1].get_text().split(',')))
            except:
                sma_50 = "--"
            try:
                s = ""
                sma_200 = float(s.join(soup2.find_all("table")[2].find_all("tbody")[0].find_all("tr")[5].find_all("td")[
                                           1].get_text().split(',')))
            except:
                sma_200 = "--"
            try:
                s = ""
                pp = float(s.join(soup2.find_all("table")[5].find_all("tbody")[0].find_all("tr")[0].find_all("td")[
                                      4].get_text().split(',')))
            except:
                pp = "--"
            try:
                stc_k = float(
                    soup2.find_all("table")[4].find_all("tbody")[1].find_all("tr")[2].find_all("td")[1].get_text())
            except:
                stc_k = "--"
            try:
                roc = float(
                    soup2.find_all("table")[4].find_all("tbody")[1].find_all("tr")[3].find_all("td")[1].get_text())
            except:
                roc = "--"
            try:
                mfi = float(
                    soup2.find_all("table")[4].find_all("tbody")[1].find_all("tr")[6].find_all("td")[1].get_text())
            except:
                mfi = "--"
            try:
                wil = float(
                    soup2.find_all("table")[4].find_all("tbody")[1].find_all("tr")[5].find_all("td")[1].get_text())
            except:
                wil = "--"

            values = [rsi_val, macd_val, sma_10, sma_20, sma_50, sma_200, "", "", "", pp, stc_k, roc, mfi, wil]

            if rsi_val == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif rsi_val > 70:
                tech_count += 5
                senti.append("Overbought")
                senti_color.append("dark")
            elif 70 >= rsi_val > 30:
                tech_count += 10
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if macd_val == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif macd_val > 0:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                onlyma = "Below sig. line"
                senti_color.append("danger")

            if sma_10 == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif price > sma_10:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if sma_20 == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif price > sma_20:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if sma_50 == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif price > sma_50:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if sma_200 == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif price > sma_200:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if sma_20 == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif sma_10 > sma_20:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
                bet.append("Above")
            else:
                senti.append("Bearish")
                senti_color.append("danger")
                bet.append("Below")

            if sma_50 == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif sma_20 > sma_50:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
                bet.append("Above")
            else:
                senti.append("Bearish")
                senti_color.append("danger")
                bet.append("Below")

            if sma_200 == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif sma_50 > sma_200:
                tech_count += 10
                senti.append("Bullish")
                senti_color.append("success")
                bet.append("Above")
            else:
                senti.append("Bearish")
                senti_color.append("danger")
                bet.append("Below")

            if pp == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif price > pp:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if stc_k == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif stc_k > 70:
                tech_count += 5
                senti.append("Overbought")
                senti_color.append("dark")
            elif 30 <= stc_k <= 70:
                tech_count += 10
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if roc == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif roc > 0:
                tech_count += 5
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if mfi == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif mfi > 70:
                tech_count += 5
                senti.append("Overbought")
                senti_color.append("dark")
            elif 70 >= mfi > 30:
                tech_count += 10
                senti.append("Bullish")
                senti_color.append("success")
            else:
                senti.append("Bearish")
                senti_color.append("danger")

            if wil == "--":
                senti.append("Not Enough Data")
                senti_color.append("dark")
            elif wil > -20:
                tech_count += 10
                senti.append("Bullish")
                senti_color.append("success")
            elif -20 >= wil >= -45:
                tech_count += 5
                senti.append("Neutral")
                senti_color.append("dark")
            else:
                senti.append("Bearish")
                senti_color.append("danger")
            flag = 0

        # Overview Section Starts From Here
        overview = []
        color_over = []
        senti_over = []
        counto = 0
        countoarr = []
        try:
            intrin = float(soup.find_all("div", {"class": "low_high3"})[4].get_text())
            if price <= intrin*70/100:
                overview.append("Stock price is below Intrinsic Value")
                color_over.append("success")
                senti_over.append("Bullish")
            else:
                overview.append("Stock price is above Intrinsic Value")
                color_over.append("danger")
                senti_over.append("Bearish")
            counto = counto + 1
        except:
            pass

        try:
            pe = float(soup.find_all("ul", {"class": "val_listinner"})[0].find_all("li")[1].find_all("div")[1].get_text())
            spe = float(soup.find_all("ul", {"class": "val_listinner"})[1].find_all("li")[1].find_all("div")[1].get_text())
            if pe < spe:
                overview.append("Stock PE is less than Sector PE")
                color_over.append("success")
                senti_over.append("Bullish")
            else:
                overview.append("Stock PE is more than Sector PE")
                color_over.append("danger")
                senti_over.append("Bearish")
            counto = counto + 1
        except:
            pass

        try:
            fiio = float(soup.find_all("div", {"class": "finance_lft"})[1].find_all("tr")[3].find_all("td")[2].get_text().strip())
            fiic = float(soup.find_all("div", {"class": "finance_lft"})[1].find_all("tr")[3].find_all("td")[1].get_text().strip())
            if fiic > fiio:
                overview.append("FII Shareholding has Increased")
                color_over.append("success")
                senti_over.append("Bullish")
            elif fiic < fiio:
                overview.append("FII Shareholding has Decreased")
                color_over.append("danger")
                senti_over.append("Bearish")
            else:
                overview.append("FII Shareholding is maintained")
                color_over.append("dark")
                senti_over.append("Neutral")
            counto = counto + 1
        except:
            pass

        try:
            diio = float(soup.find_all("div", {"class": "finance_lft"})[1].find_all("tr")[4].find_all("td")[2].get_text().strip())
            diic = float(soup.find_all("div", {"class": "finance_lft"})[1].find_all("tr")[4].find_all("td")[1].get_text().strip())
            if diic > diio:
                overview.append("DII Shareholding has Increased")
                color_over.append("success")
                senti_over.append("Bullish")
            elif diic < diio:
                overview.append("DII Shareholding has Decreased")
                color_over.append("danger")
                senti_over.append("Bearish")
            else:
                overview.append("DII Shareholding is maintained")
                color_over.append("dark")
                senti_over.append("Neutral")
            counto = counto + 1
        except:
            pass

        try:
            piio = float(soup.find_all("div", {"class": "finance_lft"})[1].find_all("tr")[1].find_all("td")[2].get_text().strip())
            piic = float(soup.find_all("div", {"class": "finance_lft"})[1].find_all("tr")[1].find_all("td")[1].get_text().strip())
            if piic > piio:
                overview.append("Promoters Shareholding has Increased")
                color_over.append("success")
                senti_over.append("Bullish")
            elif piic < piio:
                overview.append("Promoters Shareholding has Decreased")
                color_over.append("danger")
                senti_over.append("Bearish")
            else:
                overview.append("Promoters Shareholding is maintained")
                color_over.append("dark")
                senti_over.append("Neutral")
            counto = counto + 1
        except:
            pass

        try:
            ppiio = float(soup.find_all("div", {"class": "finance_lft"})[1].find_all("tr")[2].find_all("td")[2].get_text().strip())
            ppiic = float(soup.find_all("div", {"class": "finance_lft"})[1].find_all("tr")[2].find_all("td")[1].get_text().strip())
            if ppiic > ppiio:
                overview.append("Promoter's Pledging Shareholding has Increased")
                color_over.append("danger")
                senti_over.append("Bearish")
            elif ppiic < ppiio:
                overview.append("Promoter's Pledging Shareholding has Decreased")
                color_over.append("success")
                senti_over.append("Bullish")
            elif ppiic == 0:
                overview.append("Promoter's Pledging is 0%")
                color_over.append("success")
                senti_over.append("Bullish")
            else:
                overview.append("Promoter's Pledging Shareholding is maintained")
                color_over.append("dark")
                senti_over.append("Neutral")
            counto = counto + 1
        except:
            pass

        try:
            miio = float(soup.find_all("div", {"class": "finance_lft"})[1].find_all("tr")[7].find_all("td")[2].get_text().strip())
            miic = float(soup.find_all("div", {"class": "finance_lft"})[1].find_all("tr")[7].find_all("td")[1].get_text().strip())
            if miic > miio:
                overview.append("Mutual Funds Shareholding have Increased")
                color_over.append("success")
                senti_over.append("Bullish")
            elif miic < miio:
                overview.append("Mutual Funds Shareholding have Decreased")
                color_over.append("danger")
                senti_over.append("Bearish")
            else:
                overview.append("Mutual Funds Shareholding is maintained")
                color_over.append("dark")
                senti_over.append("Neutral")
            counto = counto + 1
        except:
            pass

        try:
            stock_return = float(soup.find_all("tbody",{"id": "BSE_history_tbody"})[0].find_all("tr")[3].find("div").get_text())
            if stock_return > 7:
                overview.append("1 Year Return is more than FD")
                color_over.append("success")
                senti_over.append("Bullish")
            else:
                overview.append("1 Year Return is less than FD")
                color_over.append("danger")
                senti_over.append("Bearish")
            counto = counto + 1
        except:
            pass

        try:
            stock_return = float(soup.find_all("tbody",{"id": "BSE_history_tbody"})[0].find_all("tr")[3].find("div").get_text())
            # Manually Change karneka
            nifty_return = 2.2
            if stock_return > nifty_return:
                overview.append("1 Year Return is more than Nifty 50 and Sensex Returns")
                color_over.append("success")
                senti_over.append("Bullish")
            else:
                overview.append("1 Year Return is less than Nifty 50 and Sensex Returns")
                color_over.append("danger")
                senti_over.append("Bearish")
            counto = counto + 1
        except:
            pass

        try:
            s = ""
            q1 = float(s.join(soup.find_all("div", {"id": "IncomeStatement"})[0].find_all("tr")[8].find_all("td")[1].get_text().split(",")))
            q2 = float(s.join(soup.find_all("div", {"id": "IncomeStatement"})[0].find_all("tr")[8].find_all("td")[2].get_text().split(",")))
            q5 = float(s.join(soup.find_all("div", {"id": "IncomeStatement"})[0].find_all("tr")[8].find_all("td")[5].get_text().split(",")))
            if q1 > q2:
                overview.append("QoQ Net Profit is Increased")
                color_over.append("success")
                senti_over.append("Bullish")
            else:
                overview.append("QoQ Net Profit is Decreased")
                color_over.append("danger")
                senti_over.append("Bearish")

            if q1 > q5:
                overview.append("YoY Net Profit is Increased")
                color_over.append("success")
                senti_over.append("Bullish")
            else:
                overview.append("Yoy Net Profit is Decreased")
                color_over.append("danger")
                senti_over.append("Bearish")
            counto = counto + 1
        except:
            pass
        experts = 0
        try:
            experts = float(soup.find_all("ul", {"class": "buy_sellper"})[0].find_all("li")[0].get_text().split("%")[0])
            if experts > 50:
                overview.append("Experts and public sentiments are Bullish")
                color_over.append("success")
                senti_over.append("Bullish")
            else:
                overview.append("Experts and public sentiments are Bearish")
                color_over.append("danger")
                senti_over.append("Bearish")
            counto = counto + 1
        except:
            pass

        hl = []
        high_52 = 100
        low_52 = 1
        try:
            high_52 = float(soup.find_all("div", {"class": "low_high3"})[4].get_text())
            low_52 = float(soup.find_all("div", {"class": "low_high1"})[4].get_text())
            hl.append(high_52)
            hl.append(low_52)
            if price > high_52*93/100:
                overview.append("Stock Price near 52 Week High")
                color_over.append("success")
                senti_over.append("Bullish")
            elif low_52*107/100 > price >= low_52:
                overview.append("Stock Price near 52 Week Low")
                color_over.append("danger")
                senti_over.append("Bearish")
            counto = counto + 1
        except:
            try:
                high_52 = float(soup.find_all("div", {"class": "low_high3"})[1].get_text())
                low_52 = float(soup.find_all("div", {"class": "low_high1"})[1].get_text())
                hl.append(high_52)
                hl.append(low_52)
                if price > high_52 * 93 / 100:
                    overview.append("Stock Price near 52 Week High")
                    color_over.append("success")
                    senti_over.append("Bullish")
                elif low_52 * 107 / 100 > price >= low_52:
                    overview.append("Stock Price near 52 Week Low")
                    color_over.append("danger")
                    senti_over.append("Bearish")
                counto = counto + 1
            except:
                 pass
        if counto != 0:
            scor = 100/counto
        else:
            scor = 0
        total_score = 0
        for i in senti_over:
            if i == "Bullish":
                total_score += scor
            elif i == "Neutral":
                total_score += scor/2
        overview3 = overview[:9]
        color_over3 = color_over[:9]
        senti_over3 = senti_over[:9]
        overview4 = overview[9:]
        color_over4 = color_over[9:]
        senti_over4 = senti_over[9:]
        zipo3 = zip(overview3,color_over3,senti_over3)
        zipo4 = zip(overview4,color_over4,senti_over4)

        #Sentiment Section
        analysis_score = (tech_count + round(total_score,0))/2
        public_senti = 0

        if experts != 0.0 and experts != 100.0:
            public_senti = experts
        elif high_52:
            public_senti = ((100 - ((high_52-price)/high_52)*100) + ((price-low_52)/low_52)*100)/2
        else:
            public_senti = 5

        note = []
        if analysis_score - 10 < public_senti < analysis_score + 10:
            note = ["Neutral","dark"]
        elif analysis_score - 30 < public_senti < analysis_score - 10:
            note = ["Fear", "success"]
        elif analysis_score - 100 < public_senti < analysis_score - 30:
            note = ["Extreme Fear", "success"]
        elif analysis_score + 10 < public_senti < analysis_score + 30:
            note = ["Greed", "danger"]
        else:
            note = ["Extreme Greed", "danger"]

        dictt = {'stockn': stock_name, 'nse_ticker': nse_ticker, "bse_ticker": bse_ticker, 'price': price,
                 'change': change,
                 'colors': color, 'tech_count': tech_count, 'senti': senti, 'scolors': senti_color, 'flag': flag,
                 'values': values, 'onlyma': onlyma, 'onlyst': onlyst, 'bet': bet, 'zipo3': zipo3, 'zipo4': zipo4,
                 'over_count': round(total_score,0),'totall':tech_count + round(total_score,0),'anas': analysis_score,
                 'ps': round(public_senti,0) , 'note': note,'hl': hl}

        wb = xl.load_workbook('login/users.xlsx')
        ip = get_client_ip(request)
        sheet = wb["Sheet1"]
        for i in range(2, sheet.max_row + 1):
            if (ip == sheet.cell(i, 3).value):
                if (sheet.cell(i, 4).value == "yes"):
                    print("matched")
                    dictt["email"] = sheet.cell(i, 1).value
        return render(request, 'stockdata.html', dictt)
