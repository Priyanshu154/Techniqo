import openpyxl as xl
from openpyxl.utils import column_index_from_string


indexs = ['midcap_50.xlsx','midcap_100.xlsx','midcap_150.xlsx','nifty_50.xlsx','nifty_100.xlsx','nifty_200.xlsx','nifty_500.xlsx','nifty_next_50.xlsx','smallcap_50.xlsx','smallcap_250.xlsx']

for a in indexs:
    wb = xl.load_workbook('new.xlsx')
    wb2 = xl.load_workbook(a)

    sheet1 = wb['new']
    sheet2 = wb2['Sheet1']

    for i in range(2,sheet2.max_row+1):
        sheet2.cell(i,35).value = sheet2.cell(i,19).value
        sheet2.cell(i,36).value = sheet2.cell(i,20).value
        sheet2.cell(i,37).value = sheet2.cell(i,21).value
        sheet2.cell(i,38).value = sheet2.cell(i,22).value
        sheet2.cell(i,39).value = sheet2.cell(i,23).value
        sheet2.cell(i,40).value = sheet2.cell(i,24).value

    for i in range(2,sheet2.max_row+1):
        sheet2.cell(i, 19).value = sheet2.cell(i, 2).value
        sheet2.cell(i, 20).value = sheet2.cell(i, 3).value
        sheet2.cell(i, 21).value = sheet2.cell(i, 4).value
        sheet2.cell(i, 22).value = sheet2.cell(i, 5).value
        sheet2.cell(i, 23).value = sheet2.cell(i, 6).value
        sheet2.cell(i, 24).value = sheet2.cell(i, 7).value

    for j in range(2,sheet2.max_row+1):
        for i in range(2,sheet1.max_row+1):
            if (sheet1.cell(i,1).value == sheet2.cell(j,1).value) and sheet1.cell(i,2).value == "EQ":
                sheet2.cell(j,2).value = sheet1.cell(i,3).value
                sheet2.cell(j,3).value = sheet1.cell(i,4).value
                sheet2.cell(j,4).value = sheet1.cell(i,5).value
                sheet2.cell(j,5).value = sheet1.cell(i,6).value
                sheet2.cell(j,6).value = sheet1.cell(i,6).value - sheet1.cell(i,8).value
                sheet2.cell(j,7).value = (sheet1.cell(i,6).value - sheet1.cell(i,8).value)/sheet1.cell(i,8).value*100

    wb.save('new.xlsx')
    wb2.save(a)