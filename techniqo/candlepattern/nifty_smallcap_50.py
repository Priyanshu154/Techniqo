from django.shortcuts import render
from django.http import HttpResponse
import openpyxl as xl
from openpyxl.utils import column_index_from_string
import os


# Create your views here.

def hammerf(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('M'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Hammer' , 'number':num, 'paras': para}
    return render(request, 'result.html', dictt)

def dragonfly_dojif(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('L'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Dragonfly Doji' , 'number':num, 'paras': para}
    return render(request, 'dragonfly_dojih.html', dictt)

def white_marubozuf(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('Q'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'White Marubozu' , 'number':num, 'paras': para}
    return render(request, 'white_marubozuh.html', dictt)

def bullish_engulfingf(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('AE'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Bullish Engulfing' , 'number':num, 'paras': para}
    return render(request, 'bullish_engulfingh.html', dictt)

def bullish_haramif(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('AC'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Bullish Harami' , 'number':num, 'paras': para}
    return render(request, 'bullish_haramih.html', dictt)

def rising_sunf(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('AG'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Rising Sun' , 'number':num, 'paras': para}
    return render(request, 'rising_sunh.html', dictt)

def morning_starf(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('AU'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Morning Star' , 'number':num, 'paras': para}
    return render(request, 'morning_starh.html', dictt)

def three_white_solidersf(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('AW'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Three White Soliders' , 'number':num, 'paras': para}
    return render(request, 'three_white_solidersh.html', dictt)

def inverted_hammerf(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('O'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Inverted Hammer' , 'number':num, 'paras': para}
    return render(request, 'inverted_hammerh.html', dictt)

def gravestone_dojif(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('N'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Gravestone Doji' , 'number':num, 'paras': para}
    return render(request, 'gravestone_dojih.html', dictt)

def black_marubozuf(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('R'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Black Marubozu' , 'number':num, 'paras': para}
    return render(request, 'black_marubozuh.html', dictt)

def bearish_engulfingf(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('AF'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Bearish Engulfing' , 'number':num, 'paras': para}
    return render(request, 'bearish_engulfingh.html', dictt)

def bearish_haramif(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('AD'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Bearish Harami' , 'number':num, 'paras': para}
    return render(request, 'bearish_haramih.html', dictt)

def dark_cloudf(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('AH'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Dark Cloud' , 'number':num, 'paras': para}
    return render(request, 'dark_cloudh.html', dictt)

def evening_starf(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('AV'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Evening Star' , 'number':num, 'paras': para}
    return render(request, 'evening_starh.html', dictt)

def three_black_crowsf(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('AX'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Three Black Crows' , 'number':num, 'paras': para}
    return render(request, 'three_black_crowsh.html', dictt)

def dojif(request):
    #jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
    workpath = os.path.dirname(os.path.abspath(__file__))
    xx = os.path.join(workpath, 'smallcap_50.xlsx') #yaha tak
    wb = xl.load_workbook(xx, data_only=True)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    stocks = []
    opens = []
    high = []
    low = []
    close = []
    num = 0
    for i in range(2 , sheet.max_row + 1):

        cell = sheet.cell(i, column_index_from_string('P'))
        c = cell.value
        if c == "YES":
            num = 1
            stocks.append(sheet.cell(i,1).value)
            opens.append(sheet.cell(i,2).value)
            high.append(sheet.cell(i,3).value)
            low.append(sheet.cell(i,4).value)
            close.append(sheet.cell(i,5).value)
    para = "A hammer is a price pattern in candlestick charting that occurs when a security trades significantly lower than its opening, but rallies within the period to close near opening price. This pattern forms a hammer-shaped candlestick, in which the lower shadow is at least twice the size of the real body. The body of the candlestick represents the difference between the open and closing prices, while the shadow shows the high and low prices for the period."
    zipp = zip(stocks,opens,high,low,close)
    dictt = {'zips': zipp ,'candle' : 'Doji' , 'number':num, 'paras': para}
    return render(request, 'dojih.html', dictt)