import openpyxl as xl
from openpyxl.utils import column_index_from_string
from django.shortcuts import render
from django.http import HttpResponse

wb = xl.load_workbook('CandelStickp.xlsx', data_only=True)
sheet = wb['Sheet1']
cell = sheet['a1']

candle_type = input("Enter Type of Candle Stick: ")

if candle_type == "Hammer":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('M'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Inverted Hammer":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('O'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Dragonfly Doji":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('L'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "GraveStone Doji":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('N'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Doji":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('P'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "White Marabozu":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('Q'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Black Marabozu":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('R'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Bullish Harami":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('AC'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Bearish Harami":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('AD'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Bullish Engulfing":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('AE'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Bearish Engulfing":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('AF'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Rising Sun":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('AG'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Dark Cloud":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('AH'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Morning Star":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('AU'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Evening Star":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('AV'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Three White Soliders":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('AW'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)
elif candle_type == "Three Black Crows":
    for i in range(2 , sheet.max_row + 1):
        #print(candle_type)
        cell = sheet.cell(i, column_index_from_string('AX'))
        c = cell.value
        if c == "YES":
            print(sheet.cell(i,1).value)