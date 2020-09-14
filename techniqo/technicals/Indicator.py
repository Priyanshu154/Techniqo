import pandas_datareader.data as web
from datetime import datetime
from matplotlib.dates import date2num
import math
import openpyxl as xl
from openpyxl.utils import column_index_from_string
import os
import pandas as pd
nifty_500 = ["3MINDIA", "ABB", "ACC", "AIAENG", "APLAPOLLO", "AUBANK", "AARTIIND", "AAVAS", "ABBOTINDIA", "ADANIGAS",
           "ADANIGREEN", "ADANIPORTS", "ADANIPOWER", "ADANITRANS", "ABCAPITAL", "ABFRL", "ADVENZYMES", "AEGISCHEM",
           "AFFLE", "AJANTPHARM", "AKZOINDIA", "APLLTD", "ALKEM", "ALKYLAMINE", "ALLCARGO", "AMARAJABAT", "AMBER",
           "AMBUJACEM", "APOLLOHOSP", "APOLLOTYRE", "ARVINDFASN", "ASAHIINDIA", "ASHOKLEY", "ASHOKA", "ASIANPAINT",
           "ASTERDM", "ASTRAZEN", "ASTRAL", "ATUL", "AUROPHARMA", "AVANTIFEED", "DMART", "AXISBANK", "BASF", "BEML",
           "BSE", "BAJAJ-AUTO", "BAJAJCON", "BAJAJELEC", "BAJFINANCE", "BAJAJFINSV", "BAJAJHLDNG", "BALKRISIND",
           "BALMLAWRIE", "BALRAMCHIN", "BANDHANBNK", "BANKBARODA", "BANKINDIA", "MAHABANK", "BATAINDIA", "BAYERCROP",
           "BERGEPAINT", "BDL", "BEL", "BHARATFORG", "BHEL", "BPCL", "BHARATRAS", "BHARTIARTL", "INFRATEL", "BIOCON",
           "BIRLACORPN", "BSOFT", "BLISSGVS", "BLUEDART", "BLUESTARCO", "BBTC", "BOMDYEING", "BOSCHLTD", "BRIGADE",
           "BRITANNIA", "CARERATING", "CCL", "CESC", "CRISIL", "CSBBANK", "CADILAHC", "CANFINHOME", "CANBK",
           "CAPLIPOINT", "CGCL", "CARBORUNIV", "CASTROLIND", "CEATLTD", "CENTRALBK", "CDSL", "CENTURYPLY", "CENTURYTEX",
           "CERA", "CHAMBLFERT", "CHENNPETRO", "CHOLAHLDNG", "CHOLAFIN", "CIPLA", "CUB", "COALINDIA", "COCHINSHIP",
           "COFORGE", "COLPAL", "CONCOR", "COROMANDEL", "CREDITACC", "CROMPTON", "CUMMINSIND", "CYIENT", "DBCORP",
           "DCBBANK", "DCMSHRIRAM", "DLF", "DABUR", "DALBHARAT", "DEEPAKNTR", "DELTACORP", "DHANUKA", "DBL", "DISHTV",
           "DCAL", "DIVISLAB", "DIXON", "LALPATHLAB", "DRREDDY", "EIDPARRY", "EIHOTEL", "ESABINDIA", "EDELWEISS",
           "EICHERMOT", "ELGIEQUIP", "EMAMILTD", "ENDURANCE", "ENGINERSIN", "EQUITAS", "ERIS", "ESCORTS", "ESSELPACK",
           "EXIDEIND", "FDC", "FEDERALBNK", "FINEORG", "FINCABLES", "FINPIPE", "FSL", "FORTIS", "FCONSUMER", "FRETAIL",
           "GAIL", "GEPIL", "GET&D", "GHCL", "GMMPFAUDLR", "GMRINFRA", "GALAXYSURF", "GRSE", "GARFIBRES", "GICRE",
           "GILLETTE", "GLAXO", "GLENMARK", "GODFRYPHLP", "GODREJAGRO", "GODREJCP", "GODREJIND", "GODREJPROP",
           "GRANULES", "GRAPHITE", "GRASIM", "GESHIP", "GREAVESCOT", "GRINDWELL", "GUJALKALI", "FLUOROCHEM",
           "GUJGASLTD", "GMDCLTD", "GNFC", "GPPL", "GSFC", "GSPL", "GULFOILLUB", "HEG", "HCLTECH", "HDFCAMC",
           "HDFCBANK", "HDFCLIFE", "HFCL", "HATHWAY", "HATSUN", "HAVELLS", "HEIDELBERG", "HERITGFOOD", "HEROMOTOCO",
           "HEXAWARE", "HSCL", "HIMATSEIDE", "HINDALCO", "HAL", "HINDCOPPER", "HINDPETRO", "HINDUNILVR", "HINDZINC",
           "HONAUT", "HUDCO", "HDFC", "ICICIBANK", "ICICIGI", "ICICIPRULI", "ISEC", "ICRA", "IDBI", "IDFCFIRSTB",
           "IDFC", "IFBIND", "IFCI", "IIFL", "IIFLWAM", "IRB", "IRCON", "ITC", "ITI", "INDIACEM", "ITDC", "IBULHSGFIN",
           "IBREALEST", "IBVENTURES", "INDIAMART", "INDIANB", "IEX", "INDHOTEL", "IOC", "IOB", "IRCTC", "INDOSTAR",
           "INDOCO", "IGL", "INDUSINDBK", "INFIBEAM", "NAUKRI", "INFY", "INGERRAND", "INOXLEISUR", "INTELLECT",
           "INDIGO", "IPCALAB", "JBCHEPHARM", "JKCEMENT", "JKLAKSHMI", "JKPAPER", "JKTYRE", "JMFINANCIL", "JSWENERGY",
           "JSWSTEEL", "JAGRAN", "JAICORPLTD", "J&KBANK", "JAMNAAUTO", "JINDALSAW", "JSLHISAR", "JSL", "JINDALSTEL",
           "JCHAC", "JUBLFOOD", "JUBILANT", "JUSTDIAL", "JYOTHYLAB", "KPRMILL", "KEI", "KNRCON", "KPITTECH", "KRBL",
           "KSB", "KAJARIACER", "KALPATPOWR", "KANSAINER", "KTKBANK", "KARURVYSYA", "KSCL", "KEC", "KOLTEPATIL",
           "KOTAKBANK", "L&TFH", "LTTS", "LICHSGFIN", "LAOPALA", "LAXMIMACH", "LTI", "LT", "LAURUSLABS", "LEMONTREE",
           "LINDEINDIA", "LUPIN", "LUXIND", "MASFIN", "MMTC", "MOIL", "MRF", "MGL", "MAHSCOOTER", "MAHSEAMLES",
           "M&MFIN", "M&M", "MAHINDCIE", "MHRIL", "MAHLOG", "MANAPPURAM", "MRPL", "MARICO", "MARUTI", "MFSL",
           "METROPOLIS", "MINDTREE", "MINDACORP", "MINDAIND", "MIDHANI", "MOTHERSUMI", "MOTILALOFS", "MPHASIS", "MCX",
           "MUTHOOTFIN", "NATCOPHARM", "NBCC", "NCC", "NESCO", "NHPC", "NLCINDIA", "NMDC", "NTPC", "NH", "NATIONALUM",
           "NFL", "NBVENTURES", "NAVINFLUOR", "NESTLEIND", "NILKAMAL", "NAM-INDIA", "OBEROIRLTY", "ONGC", "OIL",
           "OMAXE", "OFSS", "ORIENTCEM", "ORIENTELEC", "ORIENTREF", "PIIND", "PNBHOUSING", "PNCINFRA", "PTC", "PVR",
           "PAGEIND", "PERSISTENT", "PETRONET", "PFIZER", "PHILIPCARB", "PHOENIXLTD", "PIDILITIND", "PEL", "POLYMED",
           "POLYCAB", "PFC", "POWERGRID", "PRAJIND", "PRESTIGE", "PRSMJOHNSN", "PGHL", "PGHH", "PNB", "QUESS",
           "RBLBANK", "RECLTD", "RITES", "RADICO", "RVNL", "RAIN", "RAJESHEXPO", "RALLIS", "RCF", "RATNAMANI",
           "RAYMOND", "REDINGTON", "RELAXO", "RELIANCE", "REPCOHOME", "SBICARD", "SBILIFE", "SJVN", "SKFINDIA", "SRF",
           "SADBHAV", "SANOFI", "SCHAEFFLER", "SCHNEIDER", "SIS", "SEQUENT", "SFL", "SCI", "SHOPERSTOP", "SHREECEM",
           "RENUKA", "SHRIRAMCIT", "SRTRANSFIN", "SIEMENS", "SOBHA", "SOLARINDS", "SONATSOFTW", "SOUTHBANK", "SPANDANA",
           "SPICEJET", "STARCEMENT", "SBIN", "SAIL", "SWSOLAR", "STRTECH", "STAR", "SUDARSCHEM", "SUMICHEM", "SPARC",
           "SUNPHARMA", "SUNTV", "SUNDARMFIN", "SUNDRMFAST", "SUNTECK", "SUPRAJIT", "SUPREMEIND", "SUZLON",
           "SWANENERGY", "SYMPHONY", "SYNGENE", "TCIEXP", "TCNSBRANDS", "TTKPRESTIG", "TVTODAY", "TV18BRDCST",
           "TVSMOTOR", "TAKE", "TASTYBITE", "TATACOMM", "TCS", "TATACONSUM", "TATAELXSI", "TATAINVEST", "TATAMTRDVR",
           "TATAMOTORS", "TATAPOWER", "TATASTLBSL", "TATASTEEL", "TEAMLEASE", "TECHM", "NIACL", "RAMCOCEM", "THERMAX",
           "THYROCARE", "TIMETECHNO", "TIMKEN", "TITAN", "TORNTPHARM", "TORNTPOWER", "TRENT", "TRIDENT", "TIINDIA",
           "UCOBANK", "UFLEX", "UPL", "UJJIVAN", "UJJIVANSFB", "ULTRACEMCO", "UNIONBANK", "UBL", "MCDOWELL-N", "VGUARD",
           "VMART", "VIPIND", "VRLLOG", "VSTIND", "VAIBHAVGBL", "VAKRANGEE", "VTL", "VARROC", "VBL", "VENKEYS",
           "VESUVIUS", "VINATIORGA", "IDEA", "VOLTAS", "WABCOINDIA", "WELCORP", "WELSPUNIND", "WESTLIFE", "WHIRLPOOL",
           "WIPRO", "WOCKPHARMA", "ZEEL", "ZENSARTECH", "ZYDUSWELL", "ECLERX"]
start = datetime(2018,1,24)
end = datetime(2020,7,27)

# web.DataReader helps to access data of a particular stock from the site you want from starting date to ending date
# data = web.DataReader('Stock Name', 'Website', starting date, ending date)
# to see how values are stored in data please print to verify
counttt = 0
workpath = os.path.dirname(os.path.abspath(__file__))
wb = xl.load_workbook('technicals_ours.xlsx', data_only=True)
sheet1 = wb['sheet']

for ii in range(len(nifty_500)):
    try:
        st = nifty_500[ii]
        st = st.replace("&", "_")
        data_reset = pd.read_excel(f'D:/college/webend/techniqo/data_new_ticker/{st}.xlsx')
        close = data_reset['Close'].to_list()
        high = data_reset['High'].to_list()
        low = data_reset['Low'].to_list()
        openn = data_reset['Open'].to_list()
        date = data_reset['Date'].to_list()
        dt = data_reset['date_ax'].to_list()
        volume = data_reset['Volume'].to_list()


        def RSI(close, t):
            n = len(close)
            rsi = []
            Ups = 0.0
            Downs = 0.0
            for j in range(t - 1):
                rsi.append(-1)
            # Ye sabse pehla avgU/avgD find karne ke liye simple average vala step
            for i in range(1, t):
                diff = close[i] - close[i - 1]
                if (diff > 0):
                    Ups += diff
                else:
                    Downs += (-diff)

            preU = Ups / t
            preD = Downs / t
            # simple average mil gaya to hamara pehla rsi bi mil gaya
            rs = preU / preD
            rsi.append((100 - (100 / (1 + rs))))
            # yaha se prev_avgUp vala loop
            Ups = 0.0
            Downs = 0.0
            for i in range(t, n):
                diff = close[i] - close[i - 1]
                if (diff > 0):
                    Ups = diff
                    Downs = 0.0
                else:
                    Downs = (-diff)
                    Ups = 0.0
                u = (1 / t) * Ups + ((t - 1) / t) * preU
                d = (1 / t) * Downs + ((t - 1) / t) * preD
                preU = u  # Update previous-Up and previous-Down
                preD = d
                rs = u / d
                rsi.append((100 - (100 / (1 + rs))))  # RSI for a particular date
            return rsi


        # RSI Ends Here

        # SMA starts here
        def SMA(close, t):
            mas = []
            for i in range(t - 1):
                mas.append(-1)
            for i in range(len(close) - t + 1):
                summ = 0
                for j in range(i, t + i):
                    summ = summ + close[j]
                meann = summ / t
                mas.append(meann)
            return mas


        # SMA Ends here

        # Weighted Moving Average(WMA) Starts Here
        # Reference for code is taken from tradingview
        def WMA(close, t):
            wma = []
            for i in range(t - 1):
                wma.append(-1)
            for i in range(t - 1, len(close)):
                norm = 0.0
                summ = 0.0
                for j in range(0, t):
                    weight = (t - j) * t
                    norm = norm + weight
                    summ = summ + (close[i - j] * weight)
                wma.append(summ / norm)
            return wma


        # WMA Ends Here

        # Rolling Moving Average(RMA) Starts here
        def RMA(close, t):
            rma = []
            sma = SMA(close, t)
            for i in range(t):
                rma.append(sma[i])
            for i in range(t, len(close)):
                rma.append((rma[i - 1] * (t - 1) + close[i]) / t)
            return rma


        # RMA Ends here

        # Rate Of Change(ROC) Starts here
        def ROC(close, t):
            roc = []
            for i in range(t - 1):
                roc.append(-1)
            for i in range(t - 1, len(close)):
                sum = 100 * (close[i] - close[i - t]) / close[i - t]
                roc.append(sum)
            return roc


        # ROC Ends here

        # EMA Starts Here
        def EMA(close, t):
            sma = 0.0
            n = len(close)
            for i in range(t):
                sma += close[i]
            sma = sma / (t)
            ema = []
            for j in range(t - 1):
                ema.append(-1)
            ema.append(sma)
            m = 2 / (t + 1)
            for i in range(t, n):
                e = close[i] * m + ema[i - 1] * (1 - m)
                ema.append(e)
            return ema


        # EMA ends here

        # From Here Pivot Points
        final_high = []
        final_low = []
        final_close = []
        final_counts = []


        def assigning(countt, high_maxx, low_minn, closee):
            final_counts.append(countt)
            final_high.append(high_maxx)
            final_low.append(low_minn)
            final_close.append(closee)


        def pivot_points(close, high, low, date):
            flag = 0
            count = 0
            high_max = 0
            low_min = 320000
            final_high.clear()
            final_low.clear()
            final_close.clear()
            final_counts.clear()

            for i in range(len(close)):
                date_st = str(date[i])
                if date_st[3] == "0" and date_st[4] == "1":
                    if flag == 12:
                        assigning(count, high_max, low_min, close[i - 1])
                        flag = 0
                        count = 0
                        high_max = 0
                        low_min = 320000
                    else:
                        if high[i] > high_max:
                            high_max = high[i]
                        if low[i] < low_min:
                            low_min = low[i]
                        flag = 1
                        count += 1
                elif date_st[3] == "0" and date_st[4] == "2":
                    if flag == 1:
                        assigning(count, high_max, low_min, close[i - 1])
                        flag = 0
                        count = 0
                        high_max = 0
                        low_min = 320000
                    else:
                        if high[i] > high_max:
                            high_max = high[i]
                        if low[i] < low_min:
                            low_min = low[i]
                        flag = 2
                        count += 1
                elif date_st[3] == "0" and date_st[4] == "3":
                    if flag == 2:
                        assigning(count, high_max, low_min, close[i - 1])
                        flag = 0
                        count = 0
                        high_max = 0
                        low_min = 320000
                    else:
                        if high[i] > high_max:
                            high_max = high[i]
                        if low[i] < low_min:
                            low_min = low[i]
                        flag = 3
                        count += 1
                elif date_st[3] == "0" and date_st[4] == "4":
                    if flag == 3:
                        assigning(count, high_max, low_min, close[i - 1])
                        flag = 0
                        count = 0
                        high_max = 0
                        low_min = 320000
                    else:
                        if high[i] > high_max:
                            high_max = high[i]
                        if low[i] < low_min:
                            low_min = low[i]
                        flag = 4
                        count += 1
                elif date_st[3] == "0" and date_st[4] == "5":
                    if flag == 4:
                        assigning(count, high_max, low_min, close[i - 1])
                        flag = 0
                        count = 0
                        high_max = 0
                        low_min = 320000
                    else:
                        if high[i] > high_max:
                            high_max = high[i]
                        if low[i] < low_min:
                            low_min = low[i]
                        flag = 5
                        count += 1
                elif date_st[3] == "0" and date_st[4] == "6":
                    if flag == 5:
                        assigning(count, high_max, low_min, close[i - 1])
                        flag = 0
                        count = 0
                        high_max = 0
                        low_min = 320000
                    else:
                        if high[i] > high_max:
                            high_max = high[i]
                        if low[i] < low_min:
                            low_min = low[i]
                        flag = 6
                        count += 1
                elif date_st[3] == "0" and date_st[4] == "7":
                    if flag == 6:
                        assigning(count, high_max, low_min, close[i - 1])
                        flag = 0
                        count = 0
                        high_max = 0
                        low_min = 320000
                    else:
                        if high[i] > high_max:
                            high_max = high[i]
                        if low[i] < low_min:
                            low_min = low[i]
                        flag = 7
                        count += 1
                elif date_st[3] == "0" and date_st[4] == "8":
                    if flag == 7:
                        assigning(count, high_max, low_min, close[i - 1])
                        flag = 0
                        count = 0
                        high_max = 0
                        low_min = 320000
                    else:
                        if high[i] > high_max:
                            high_max = high[i]
                        if low[i] < low_min:
                            low_min = low[i]
                        flag = 8
                        count += 1
                elif date_st[3] == "0" and date_st[4] == "9":
                    if flag == 8:
                        assigning(count, high_max, low_min, close[i - 1])
                        flag = 0
                        count = 0
                        high_max = 0
                        low_min = 320000
                    else:
                        if high[i] > high_max:
                            high_max = high[i]
                        if low[i] < low_min:
                            low_min = low[i]
                        flag = 9
                        count += 1
                elif date_st[3] == "1" and date_st[4] == "0":
                    if flag == 9:
                        assigning(count, high_max, low_min, close[i - 1])
                        flag = 0
                        count = 0
                        high_max = 0
                        low_min = 320000
                    else:
                        if high[i] > high_max:
                            high_max = high[i]
                        if low[i] < low_min:
                            low_min = low[i]
                        flag = 10
                        count += 1
                elif date_st[3] == "1" and date_st[4] == "1":
                    if flag == 10:
                        assigning(count, high_max, low_min, close[i - 1])
                        flag = 0
                        count = 0
                        high_max = 0
                        low_min = 320000
                    else:
                        if high[i] > high_max:
                            high_max = high[i]
                        if low[i] < low_min:
                            low_min = low[i]
                        flag = 11
                        count += 1
                elif date_st[3] == "1" and date_st[4] == "2":
                    if flag == 11:
                        assigning(count, high_max, low_min, close[i - 1])
                        flag = 0
                        count = 0
                        high_max = 0
                        low_min = 320000
                    else:
                        if high[i] > high_max:
                            high_max = high[i]
                        if low[i] < low_min:
                            low_min = low[i]
                        flag = 12
                        count += 1

            pivot_point = []
            resistance_1 = []
            resistance_2 = []
            resistance_3 = []
            support_1 = []
            support_2 = []
            support_3 = []
            pivot_point_pr = []
            resistance_1_pr = []
            resistance_2_pr = []
            resistance_3_pr = []
            support_1_pr = []
            support_2_pr = []
            support_3_pr = []

            for i in range(len(final_counts)):
                pivot_point_pr.append((final_high[i] + final_low[i] + final_close[i]) / 3)
                support_1_pr.append((2 * pivot_point_pr[i]) - final_high[i])
                resistance_1_pr.append((2 * pivot_point_pr[i]) - final_low[i])
                support_2_pr.append(pivot_point_pr[i] - final_high[i] + final_low[i])
                resistance_2_pr.append(pivot_point_pr[i] + final_high[i] - final_low[i])
                support_3_pr.append(support_1_pr[i] - final_high[i] + final_low[i])
                resistance_3_pr.append(resistance_1_pr[i] + final_high[i] - final_low[i])
            for i in range(final_counts[0]):
                pivot_point.append(0)
                resistance_1.append(0)
                resistance_2.append(0)
                resistance_3.append(0)
                support_1.append(0)
                support_2.append(0)
                support_3.append(0)
            for i in range(1, len(final_counts)):
                for j in range(final_counts[i]):
                    pivot_point.append(pivot_point_pr[i])
                    resistance_1.append(resistance_1_pr[i])
                    resistance_2.append(resistance_2_pr[i])
                    resistance_3.append(resistance_3_pr[i])
                    support_1.append(support_1_pr[i])
                    support_2.append(support_2_pr[i])
                    support_3.append(support_3_pr[i])
            return pivot_point, support_1, support_2, support_3, resistance_1, resistance_2, resistance_3
        # Pivot Points Ends Here

        # MACD Starts From Here
        def EMA_d(close, t):
            sma = 0.0
            n = len(close)
            for i in range(t):
                sma += close[i]
            sma = sma / (t)
            ema = []
            ema.append(sma)
            m = 2 / (t + 1)
            for i in range(t, n):
                e = close[i] * m + ema[i - t] * (1 - m)
                ema.append(e)
            return ema


        def EMA_MACD(t, macd):
            sma = 0.0
            n = len(macd)
            for i in range(t):
                sma += macd[i]
            sma = sma / (t)
            ema = []
            ema.append(sma)
            m = 2 / (t + 1)
            for i in range(t, n):
                e = macd[i] * m + ema[i - t] * (1 - m)
                ema.append(e)
            return ema


        def MACD(close, x, y, z):
            val_pr = EMA_d(close, x)
            val2_pr = EMA_d(close, y)
            val = []
            val2 = []
            for i in range(x - 1):
                val.append(0)
            for i in range(y - 1):
                val2.append(0)

            for i in range(len(val_pr)):
                val.append(val_pr[i])
            for i in range(len(val2_pr)):
                val2.append(val2_pr[i])

            macd_line = []
            macd_histogram = []
            signal_line = []

            for i in range(len(val)):
                macd_line.append(val[i] - val2[i])

            for i in range(z - 1):
                signal_line.append(0)

            signal_line_pr = EMA_MACD(z, macd_line)

            for i in range(len(signal_line_pr)):
                signal_line.append(signal_line_pr[i])

            for i in range(len(val)):
                macd_histogram.append(macd_line[i] - signal_line[i])

            return macd_line, signal_line, macd_histogram


        # MACD Ends Here

        # Bollinger Band Starts Here
        def bollinger_band(close, n, r):
            up = []
            lo = []
            ma = []
            for i in range(n - 1):
                up.append(0)
                lo.append(0)
                ma.append(0)
            for i in range(len(close) - n + 1):
                sum = 0
                sqr = 0
                for j in range(i, n + i):
                    sum = sum + close[j]
                meann = sum / n
                ma.append(sum / n)
                for z in range(i, n + i):
                    sq = close[z] - meann
                    sqr = sqr + (sq * sq)
                varr = sqr / n
                std = math.sqrt(varr)
                up.append(meann + (r * std))
                lo.append(meann - (r * std))
            return up, lo, ma


        # Bollinger Band Ends here

        # Fibonacci Retracement start here
        def fib_retracement(p1, p2):
            list = [0, 0.236, 0.382, 0.5, 0.618, 0.786, 1, 1.618, 2.618, 3.618, 4.236]
            dict = {}
            dist = p2 - p1
            for val in list:
                dict[str(val)] = (p2 - dist * val)
            return dict


        # Fibonacci Retracement ends here

        # Money Flow Index starts here
        def MFI(high, low, close, volume, t):
            mfi = []  # money flow index
            typ = []  # typical price
            raw_money = []  # raw money flow
            mfr = []  # money flow ratio
            for i in range(t):
                mfi.append(-1)
                mfr.append(-1)
            ind = 1
            typ.append((high[0] + low[0] + close[0]) / 3)
            raw_money.append(typ[0] * volume[0])  # first time assume it is positive

            for i in range(1, len(close)):
                typ.append((high[i] + low[i] + close[i]) / 3)
                if (typ[ind] > typ[ind - 1]):
                    raw_money.append(typ[i] * volume[i])
                else:
                    raw_money.append(-typ[i] * volume[i])
                ind = ind + 1
            for i in range(t, len(close)):
                positive_flows = 0.0
                negative_flows = 0.0
                for j in range(t):
                    if (raw_money[i - j] > 0):
                        positive_flows += raw_money[i - j]
                    else:
                        negative_flows += -raw_money[i - j]
                if (negative_flows != 0):
                    ratio = positive_flows / negative_flows
                else:
                    ratio = positive_flows
                mfr.append(ratio)
                mfi.append((100 - (100 / (1 + ratio))))
            return mfi


        # Money Flow Index ends here

        # Stochastic Rsi Starts ahi thi

        def Rsi_high(high, t):

            rsi_H = []
            for i in range(0, t - 1):
                rsi_H.append(-1)

            i = 0
            for j in range(t, len(high) + 1):
                HIGH = high[i:t]
                rsi_H.append(max(HIGH))
                t += 1
                i += 1

            return rsi_H


        def Rsi_low(low, t):

            rsi_L = []
            for i in range(0, t - 1):
                rsi_L.append(-1)

            i = 0

            for j in range(t, len(low) + 1):
                if low != -1:
                    LOW = low[i:t]
                    rsi_L.append(min(LOW))
                    t += 1
                    i += 1

            return rsi_L


        def stoch(source, high, low, t, rt, close):
            rsi_high = []
            rsi_low = []

            rsi_low = Rsi_low(high, t)
            rsi_high = Rsi_high(low, t)

            count = 0
            for x in rsi_low:
                if (x == -1):
                    count += 1

            Stochastic = []
            for i in range(0, count):
                Stochastic.append(-1)

            cnt = 0
            rsi = RSI(close, rt)
            for i in range(count, (len(source))):
                y = (rsi[i] - rsi_low[i])
                z = (rsi_high[i] - rsi_low[i])
                w = y / z
                Stochastic.append(w * 100)
                cnt += 1

            return Stochastic, count


        def sma(rsi, t, count):
            x = []
            cnt = 0
            for i in range(0, count):
                x.append(-1)
                cnt += 1
            for i in range(t - 1):
                x.append(-1)
                cnt += 1

            cnt += 1
            cnt1 = cnt

            for i in range(cnt, len(rsi) + 1):
                temp = rsi[cnt1 - t:cnt1]
                sum = 0.0000
                for j in temp:
                    sum = sum + j

                sum = sum / t
                cnt1 += 1
                del temp
                x.append(sum)

            return x


        def S_RSI(Close, t, K, D, rt):
            # rt=rsi peroid
            # t=Stochastic Rsi Period
            # K=main line
            # D= moving average of K

            rsi = RSI(Close, rt)
            Stochstic, count = stoch(rsi, rsi, rsi, t, rt, Close)
            k = sma(Stochstic, K, count)
            d = sma(k, D, count)

            return k, d

            # k= blue line on trading view
            # d= orange line on trading view


        # Stochastic Rsi Ends Here

        # Ichimoku Cloud Starts ahi thi

        def IC_high(high, t):

            ic_high = []
            for i in range(0, t - 1):
                ic_high.append(-1)

            i = 0
            for j in range(t, len(high) + 1):
                HIGH = high[i:t]
                ic_high.append(max(HIGH))
                t += 1
                i += 1

            return ic_high


        def IC_low(low, high, t):

            ic_low = []
            for i in range(0, t - 1):
                ic_low.append(-1)

            i = 0
            for j in range(t, len(high) + 1):
                LOW = low[i:t]
                ic_low.append(min(LOW))
                t += 1
                i += 1

            return ic_low


        def average(ic_high, ic_low, high):
            cnt = 0
            cnt1 = 0
            cnt2 = 0
            avg = []
            for i in ic_high:
                if i == -1:
                    cnt1 = cnt1 + 1

            for i in ic_low:
                if i == -1:
                    cnt2 = cnt2 + 1

            if cnt2 > cnt1:
                cnt = cnt2
            else:
                cnt = cnt1

            for i in range(0, cnt):
                avg.append(-1)

            for i in range(cnt, len(high)):
                avg.append((ic_high[i] + ic_low[i]) / 2)

            return avg


        def lag(close, time):
            lag1 = []

            for i in close:
                lag1.append(i)

            return lag1


        def Icloud(high, low, close, c_period, b_period, span_b_period, lag_span_period):

            # c_line is conversion line also known as Tenken-san
            # b_line is base line also known as kijun-san
            # other all are time peroids

            c_high = IC_high(high, c_period)
            c_low = IC_low(low, high, c_period)
            conversion_line = average(c_high, c_low, high)

            b_high = IC_high(high, b_period)
            b_low = IC_low(low, high, b_period)
            base_line = average(b_high, b_low, high)
            span_a = []
            span_b = []
            for jj in range(26):
                span_a.append(-1)
                span_b.append(-1)

            span_a = average(conversion_line, base_line, high)

            span_b_high = IC_high(high, span_b_period)
            span_b_low = IC_low(low, high, span_b_period)
            span_b = average(span_b_high, span_b_low, high)

            lag_span = lag(close, lag_span_period)

            return conversion_line, base_line, span_a, span_b, lag_span
            # the last array of all values is matching with last value on trading view.


        # Ichimoku Cloud Ends Here

        # ATR Starts Ahi Thi

        def tr(high, low, close):
            X = []
            Y = [-1]
            Z = [-1]
            TR = [-1]
            for i in range(len(low)):
                X.append(high[i] - low[i])

            for i in range(1, len(high)):
                Y.append(abs(high[i] - close[i - 1]))

            for i in range(1, len(low)):
                Z.append(abs(low[i] - close[i - 1]))

            for i in range(1, len(low)):
                TR.append(max(X[i], Y[i], Z[i]))

            return TR


        def ATR(source, t):
            # Source Might be EITHER EMA,RMA,SMA OR WMA.
            # At the moment WMA & RMA isn't added so it will return None
            # T Is Time Period
            # take source as a string

            TR = tr()

            source = source.upper()

            if source == "EMA":
                ema = EMA(TR, t)
            elif source == "RMA":
                rma = RMA(TR, t)
            elif source == "WMA":
                wma = WMA(TR, t)
            else:
                sma = SMA(TR, t)

            # for returning
            if source == "EMA":
                return ema
            elif source == "RMA":
                return rma
            elif source == "WMA":
                return wma
            else:
                return sma


        # ATR Ends Here

        # William %R Starts Ahi Thi

        def WILLIAM_R(source, t, high, low):

            W_R = []

            for i in range(0, t - 1):
                W_R.append(-1)

            # hh is highest high
            # ll is lowest low
            hh = Rsi_high(high, t)
            ll = Rsi_low(low, t)

            for i in range(t - 1, len(source)):
                x = source[i] - hh[i]
                y = hh[i] - ll[i]
                z = x / y
                z = z * (100)
                W_R.append(z)

            return W_R


        # William %R Ends Here

        # Super Trend Starts Ahi Thi
        # tx3 uses rma in atr & super trend uses atr so if you want to check use rma in atr in tx3
        def ST(s_atr, t_atr, mul, high, low, close):
            # s_atr Is Source for ATR & t_atr is Time Period For ATR
            # mul is multiplier
            up = []
            down = []
            f_down = []
            f_up = []
            st = []
            cnt = 0
            atr = ATR(s_atr, t_atr)
            for i in range(0, t_atr - 1):
                up.append(-1)
                f_up.append(-1)
                down.append(-1)
                f_down.append(-1)
                st.append(-1)
                cnt += 1
            for i in range(cnt, len(high)):
                x = high[i]
                y = low[i]
                z = (x + y) / 2
                w = atr[i] * mul
                up.append(z + w)
                down.append(z - w)
            for i in range(cnt, len(close)):

                if (i != len(close)):
                    if ((up[i] < f_up[i - 1]) or (close[i - 1] > f_up[i - 1])):
                        f_up.append(up[i])
                    else:
                        f_up.append(f_up[i - 1])

                    if ((down[i] > f_down[i - 1]) or (close[i - 1] < f_down[i - 1])):
                        f_down.append(down[i])
                    else:
                        f_down.append(f_down[i - 1])

            for i in range(cnt, len(high)):

                if ((st[i - 1] == f_up[i - 1]) and (close[i] < f_up[i])):
                    st.append(f_up[i])
                elif ((st[i - 1] == f_up[i - 1]) and (close[i] > f_up[i])):
                    st.append(f_down[i])
                elif ((st[i - 1] == f_down[i - 1]) and (close[i] > f_down[i])):
                    st.append(f_down[i])
                elif ((st[i - 1] == f_down[i - 1]) and (close[i] < f_down[i])):
                    st.append(f_up[i])

            return st


        # Super Trend Ends Here():

        macd, sg, mh = MACD(close, 12, 26, 9)
        rsi = RSI(close, 14)
        smaa20 = SMA(close, 20)
        smaa50 = SMA(close, 50)
        smaa100 = SMA(close, 100)
        smaa200 = SMA(close, 200)
        emaa20 = EMA(close, 20)
        emaa50 = EMA(close, 50)
        emaa100 = EMA(close, 100)
        emaa200 = EMA(close, 200)
        pp, s1, s2, s3, r1, r2, r3 = pivot_points(close, high, low, date)
        mf = MFI(high, low, close, volume, 14)
        up, lo, ma = bollinger_band(close, 20, 2)
        valblue, valred = S_RSI(close, 14, 3, 3, 14)
        cl, bl, sa, sb, ls = Icloud(high, low, close, 9, 26, 52, 26)
        wil = WILLIAM_R(close, 14, high, low)
        roc = ROC(close, 9)

        for x in range(2, sheet1.max_row + 1):
            if st == sheet1.cell(x, 2).value:
                sheet1.cell(x, column_index_from_string('E')).value = rsi[len(rsi)-1]
                sheet1.cell(x, column_index_from_string('D')).value = rsi[len(rsi)-2]
                sheet1.cell(x, column_index_from_string('Y')).value = wil[len(wil)-1]
                sheet1.cell(x, column_index_from_string('BN')).value = wil[len(wil)-2]
                sheet1.cell(x, column_index_from_string('X')).value = roc[len(roc)-1]
                sheet1.cell(x, column_index_from_string('BO')).value = roc[len(roc)-2]
                sheet1.cell(x, column_index_from_string('F')).value = mh[len(mh)-1]
                sheet1.cell(x, column_index_from_string('G')).value = s3[len(s3)-1]
                sheet1.cell(x, column_index_from_string('H')).value = s2[len(s2)-1]
                sheet1.cell(x, column_index_from_string('I')).value = s1[len(s1)-1]
                sheet1.cell(x, column_index_from_string('J')).value = pp[len(pp)-1]
                sheet1.cell(x, column_index_from_string('K')).value = r1[len(r1)-1]
                sheet1.cell(x, column_index_from_string('L')).value = r2[len(r2)-1]
                sheet1.cell(x, column_index_from_string('M')).value = r3[len(r3)-1]
                sheet1.cell(x, column_index_from_string('N')).value = emaa20[len(emaa20)-1]
                sheet1.cell(x, column_index_from_string('O')).value = emaa50[len(emaa50)-1]
                sheet1.cell(x, column_index_from_string('P')).value = emaa100[len(emaa100)-1]
                sheet1.cell(x, column_index_from_string('Q')).value = emaa200[len(emaa200)-1]
                sheet1.cell(x, column_index_from_string('R')).value = smaa20[len(smaa20)-1]
                sheet1.cell(x, column_index_from_string('S')).value = smaa50[len(smaa50)-1]
                sheet1.cell(x, column_index_from_string('T')).value = smaa100[len(smaa100)-1]
                sheet1.cell(x, column_index_from_string('U')).value = smaa200[len(smaa200)-1]
                sheet1.cell(x, column_index_from_string('V')).value = mf[len(mf)-2]
                sheet1.cell(x, column_index_from_string('W')).value = mf[len(mf)-1]
                sheet1.cell(x, column_index_from_string('Z')).value = lo[len(lo)-1]
                sheet1.cell(x, column_index_from_string('AA')).value = up[len(up)-1]
                sheet1.cell(x, column_index_from_string('AB')).value = ma[len(ma)-1]
                sheet1.cell(x, column_index_from_string('AC')).value = close[len(close)-1]
                sheet1.cell(x, column_index_from_string('AD')).value = close[len(close)-2]
                sheet1.cell(x, column_index_from_string('AP')).value = cl[len(cl)-1]
                sheet1.cell(x, column_index_from_string('AQ')).value = bl[len(bl)-1]
                sheet1.cell(x, column_index_from_string('AR')).value = sa[len(sa)-26]
                sheet1.cell(x, column_index_from_string('AS')).value = sb[len(sb)-26]
                sheet1.cell(x, column_index_from_string('AT')).value = ls[len(ls)-1]
                sheet1.cell(x, column_index_from_string('AU')).value = valblue[len(valblue)-1]
                sheet1.cell(x, column_index_from_string('AV')).value = valred[len(valred)-1]
                sheet1.cell(x, column_index_from_string('AW')).value = macd[len(macd)-2]
                sheet1.cell(x, column_index_from_string('AX')).value = macd[len(macd)-1]
                sheet1.cell(x, column_index_from_string('AY')).value = sg[len(sg)-1]
                sheet1.cell(x, column_index_from_string('BA')).value = cl[len(cl)-2]
                sheet1.cell(x, column_index_from_string('BB')).value = ls[len(ls)-2]
                sheet1.cell(x, column_index_from_string('BC')).value = emaa20[len(emaa20)-2]
                sheet1.cell(x, column_index_from_string('BD')).value = emaa50[len(emaa50)-2]
                sheet1.cell(x, column_index_from_string('BE')).value = emaa100[len(emaa100)-2]
                sheet1.cell(x, column_index_from_string('BF')).value = emaa200[len(emaa200)-2]
                sheet1.cell(x, column_index_from_string('BG')).value = smaa20[len(smaa20) - 2]
                sheet1.cell(x, column_index_from_string('BH')).value = smaa50[len(smaa50) - 2]
                sheet1.cell(x, column_index_from_string('BI')).value = smaa100[len(smaa100) - 2]
                sheet1.cell(x, column_index_from_string('BJ')).value = smaa200[len(smaa200) - 2]
                sheet1.cell(x, column_index_from_string('BK')).value = bl[len(bl) - 2]
                sheet1.cell(x, column_index_from_string('BL')).value = valblue[len(valblue) - 2]
                sheet1.cell(x, column_index_from_string('BM')).value = valred[len(valred) - 2]

                wb.save('technicals_ours.xlsx')
                print(counttt)
                counttt += 1
                break
    except:
        print(nifty_500[ii])
