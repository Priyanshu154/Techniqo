# Create your views here.

import requests
from bs4 import BeautifulSoup
from django.shortcuts import render
import openpyxl as xl
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip
def check_by_id(ele,id_):
    return (len(ele.find_all(id = id_)) > 0)
def check_by_tag(ele, name):
    return (len(ele.find_all(name)) > 0)
def check_by_class(ele, name ,class_name):
    return (len(ele.find_all(name,class_ = class_name) ) >0)

def fetch_head(add):
    news = []
    time = []
    href = []
    page = requests.get(add)
    soup = BeautifulSoup(page.content, 'html.parser')
    post_elems = soup.find(id="pageContent")
    list_item = post_elems.find("div", class_="featured")
    if (check_by_tag(list_item, "h2")):
        list = list_item.find("h2")
        news.append(list.get_text())
        href.append("https://economictimes.indiatimes.com" + list.find("a")['href'])
        time.append(list_item.find("time").get_text())

    if (check_by_tag(list_item, "h3")):
        list = list_item.find("h3")
        news.append(list.get_text())
        href.append("https://economictimes.indiatimes.com" + list.find("a")['href'])
        time.append(list_item.find_all("time")[1].get_text())
    if (check_by_class(post_elems,"ul", "list1")):
        list_item = post_elems.find("ul",class_ = "list1")
        list_item = list_item.find_all("li")
        for l in list_item:
            if(not l.has_attr('class') ):
                text = (l.find("a")).get_text()
                h = (l.find("a"))['href']
                href.append("https://economictimes.indiatimes.com" + h)
                if (text != ''):        news.append(text)
                t = (l.find("time")).get_text()
                if (t != ''):            time.append(t)

    if(check_by_class(post_elems,"div", "bThumb")):
        list_item = post_elems.find("div", class_ = "bThumb")
        list_item = list_item.find("ul")
        list_item = list_item.find_all("li")
        for l in list_item:
            l2 = l.find_all("a")
            if (l2):
                text = l2[1].get_text()
                h = l2[1]['href']
                news.append(text)
                href.append("https://economictimes.indiatimes.com" + h)
                time.append("")
    news2 = zip(news,href, time)
    return news2

def fetch_article(add):
    news = []
    time = []
    href = []
    para = []
    img = []
    page = requests.get(add)
    soup = BeautifulSoup(page.content, 'html.parser')
    if(check_by_class(soup, "div", "tabdata")):
        list_item = soup.find_all("div", class_ ="eachStory")
        for l in list_item:
            head = l.find("h3")
            head = head.find("a")
            if(not head): continue
            h = head['href']
            head = head.get_text()
            t = l.find("time")
            t = t.get_text()
            content = l.find("p")
            if(not content): continue
            content = content.get_text()
            sc = l.find('img')
            if( not sc): continue
            sc = sc['data-original']
            img.append(sc)
            news.append(head)
            time.append(t)
            href.append("https://economictimes.indiatimes.com"  + h)
            para.append(content)
    news2 = zip(news,href,time,para,img)
    return news2

def industry(request):
    dict = {'news': fetch_head("https://economictimes.indiatimes.com/industry"),
            'type' : 1,
            'topic': "industry"}

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, 'newsh.html', dict)


def auto_news(request):
    add = "https://economictimes.indiatimes.com/industry/auto/auto-news/articlelist/64829342.cms"
    dict = {'type': 2, "topic": "auto"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html", dict)

def auto_cars(request):
    add = "https://economictimes.indiatimes.com/industry/auto/cars-uvs/articlelist/64829336.cms"
    dict = {'type': 2, "topic": "auto"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html", dict)

def auto_two_three(request):
    add = "https://economictimes.indiatimes.com/industry/auto/two-wheelers-three-wheelers/articlelist/64829323.cms"
    dict = {'type': 2, "topic": "auto"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html", dict)

def auto_lcv_hcv(request):
    add = "https://economictimes.indiatimes.com/industry/auto/lcv-hcv/articlelist/64829321.cms"
    dict = {'type': 2, "topic": "auto"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html", dict)

def auto_components(request):
    add = "https://economictimes.indiatimes.com/industry/auto/auto-components/articlelist/64829316.cms"
    dict = {'type': 2, "topic": "auto"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html", dict)

def auto_tyres(request):
    add = "https://economictimes.indiatimes.com/industry/auto/tyres/articlelist/64829311.cms"
    dict = {'type': 2, "topic": "auto"}
    dict['news'] = fetch_article(add)
    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html", dict)



def banking_banking(request):
    add = "https://economictimes.indiatimes.com/industry/banking/finance/banking"
    dict = {'type': 2, "topic": "banking"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def banking_finance(request):
    add = "https://economictimes.indiatimes.com/industry/banking-/-finance/banking"
    dict = {'type': 2, "topic": "banking"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def banking_insure(request):
    add = "https://economictimes.indiatimes.com/industry/banking/finance/insure/articlelist/58456919.cms"
    dict = {'type': 2, "topic": "banking"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)


def cons_durables(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/durables"
    dict = {'type': 2, "topic": "cons"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def cons_electronics(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/electronics"
    dict = {'type': 2, "topic": "cons"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def cons_fmcg(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/fmcg"
    dict = {'type': 2, "topic": "cons"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def cons_food(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/food"
    dict = {'type': 2, "topic": "cons"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def cons_garments_textiles(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/garments-/-textiles"
    dict = {'type': 2, "topic": "cons"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def cons_liquor(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/liquor"
    dict = {'type': 2, "topic": "cons"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def cons_paints(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/paints"
    dict = {'type': 2, "topic": "cons"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def cons_tobacco(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/tobacco"
    dict = {'type': 2, "topic": "cons"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def cons_fas_cos_jew(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/fashion-/-cosmetics-/-jewellery"
    dict = {'type': 2, "topic": "cons"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)

def energy_power(request):
    add = "https://economictimes.indiatimes.com/industry/energy/power"
    dict = {'type': 2, "topic": "energy"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def energy_oil_n_gas(request):
    add = "https://economictimes.indiatimes.com/industry/energy/oil-gas"
    dict = {'type': 2, "topic": "energy"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)


def indgood_cons(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/construction"
    dict = {'type': 2, "topic": "indgood"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def indgood_eng(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/engineering"
    dict = {'type': 2, "topic": "indgood"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def indgood_cement(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/cement"
    dict = {'type': 2, "topic": "indgood"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def indgood_chem_fertilisers(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/chem-/-fertilisers"
    dict = {'type': 2, "topic": "indgood"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def indgood_metals_n_mining(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/metals-mining"
    dict = {'type': 2, "topic": "indgood"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def indgood_pack(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/packaging"
    dict = {'type': 2, "topic": "indgood"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def indgood_pwgpm(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/paper-/-wood-/-glass/-plastic/-marbles"
    dict = {'type': 2, "topic": "indgood"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def indgood_petrochem(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/petrochem"
    dict = {'type': 2, "topic": "indgood"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)

def indgood_steel(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/steel"
    dict = {'type': 2, "topic": "indgood"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)

def health_healthcare(request):
    add = "https://economictimes.indiatimes.com/industry/healthcare/biotech/healthcare"
    dict = {'type': 2, "topic": "health"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def health_bio(request):
    add = "https://economictimes.indiatimes.com/industry/healthcare-/-biotech/biotech"
    dict = {'type': 2, "topic": "health"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def health_pharm(request):
    add = "https://economictimes.indiatimes.com/industry/healthcare/biotech/pharmaceuticals"
    dict = {'type': 2, "topic": "health"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)


def services_advertising(request):
    add = "https://economictimes.indiatimes.com/industry/services/advertising"
    dict = {'type': 2, "topic": "services"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def services_consultancy_audit(request):
    add = "https://economictimes.indiatimes.com/industry/services/consultancy-/-audit"
    dict = {'type': 2, "topic": "services"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def services_education(request):
    add = "https://economictimes.indiatimes.com/industry/services/education"
    dict = {'type': 2, "topic": "services"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def services_hotels_restaurants(request):
    add = "https://economictimes.indiatimes.com/industry/services/hotels-/-restaurants"
    dict = {'type': 2, "topic": "services"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def services_property_cons(request):
    add = "https://economictimes.indiatimes.com/industry/services/property-/-cstruction"
    dict = {'type': 2, "topic": "services"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def services_retail(request):
    add = "https://economictimes.indiatimes.com/industry/services/retail"
    dict = {'type': 2, "topic": "services"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def services_travel(request):
    add = "https://economictimes.indiatimes.com/industry/services/travel"
    dict = {'type': 2, "topic": "services"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)


def more_entertainment(request):
    add = "https://economictimes.indiatimes.com/industry/media-/-entertainment/entertainment"
    dict = {'type': 2, "topic": "more"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def more_media(request):
    add = "https://economictimes.indiatimes.com/industry/media-/-entertainment/media"
    dict = {'type': 2, "topic": "more"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def more_railways(request):
    add = "https://economictimes.indiatimes.com/industry/transportation/railways"
    dict = {'type': 2, "topic": "more"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def more_airlines_aviation(request):
    add = "https://economictimes.indiatimes.com/industry/transportation/airlines-/-aviation"
    dict = {'type': 2, "topic": "more"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def more_shipping_transport(request):
    add = "https://economictimes.indiatimes.com/industry/transportation/shipping-/-transport"
    dict = {'type': 2, "topic": "more"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def more_roadways(request):
    add = "https://economictimes.indiatimes.com/industry/transportation/roadways/articlelist/58456933.cms"
    dict = {'type': 2, "topic": "more"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def more_tel_news(request):
    add = "https://economictimes.indiatimes.com/industry/telecom/telecom-news/articlelist/64256852.cms"
    dict = {'type': 2, "topic": "more"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def more_tel_policy(request):
    add = "https://economictimes.indiatimes.com/industry/telecom/telecom-policy/articlelist/64256834.cms"
    dict = {'type': 2, "topic": "more"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def more_csr_initiatives(request):
    add = "https://economictimes.indiatimes.com/news/india-unlimited/csr/initiatives/articlelist/47068922.cms"
    dict = {'type': 2, "topic": "more"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def more_csr_policy(request):
    add = "https://economictimes.indiatimes.com/news/india-unlimited/csr/policy/articlelist/47068917.cms"
    dict = {'type': 2, "topic": "more"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)
def more_tech(request):
    dict = {'news': fetch_head("https://economictimes.indiatimes.com/tech"),
            'type' : 1,
            'topic': "more"}

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, 'newsh.html', dict)
def more_misc(request):
    add = "https://economictimes.indiatimes.com/industry/miscellaneous/articlelist/58456958.cms"
    dict = {'type': 2, "topic": "more"}
    dict['news'] = fetch_article(add)

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, "newsh.html",dict)

def more_env(request):
    dict = {'news': fetch_head("https://economictimes.indiatimes.com/environment"),
            'type' : 1,
            'topic': "more"}

    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if(ip == sheet.cell(i, 3).value):
            if(sheet.cell(i,4).value == "yes"):
                print("matched")
                dict["email"] = sheet.cell(i,1).value
    return render(request, 'newsh.html', dict)


def index(request):
    return industry(request)


