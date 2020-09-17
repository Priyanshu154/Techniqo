from django.shortcuts import render
from django.http import HttpResponse
import os
import openpyxl as xl
from openpyxl.utils import column_index_from_string
import datetime
workpath = os.path.dirname(os.path.abspath(__file__))
xx = os.path.join(workpath, 'peers_pd.xlsx')
wb = xl.load_workbook(xx, data_only=True)


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def ziping(stock_type):
    sheet = wb['Sheet1']

    eps = []
    pe = []
    pb = []
    ev = []
    spe = []
    roe = []
    yr = []
    mkt = []
    stk = []
    bv = []
    divn = []
    divy = []
    ps = []

    if stock_type == "banks":
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(i, column_index_from_string('B')).value == "BANKS":
                eps.append(sheet.cell(i, column_index_from_string('L')).value)
                pe.append(sheet.cell(i, column_index_from_string('G')).value)
                pb.append(sheet.cell(i, column_index_from_string('N')).value)
                bv.append(sheet.cell(i, column_index_from_string('H')).value)
                spe.append(sheet.cell(i, column_index_from_string('K')).value)
                divy.append(sheet.cell(i, column_index_from_string('O')).value)
                divn.append(sheet.cell(i, column_index_from_string('I')).value)
                mkt.append(sheet.cell(i, column_index_from_string('F')).value)
                stk.append(sheet.cell(i, column_index_from_string('A')).value)

    elif stock_type == "its":
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(i, column_index_from_string('B')).value == "IT":
                eps.append(sheet.cell(i, column_index_from_string('L')).value)
                pe.append(sheet.cell(i, column_index_from_string('G')).value)
                pb.append(sheet.cell(i, column_index_from_string('N')).value)
                bv.append(sheet.cell(i, column_index_from_string('H')).value)
                spe.append(sheet.cell(i, column_index_from_string('K')).value)
                divy.append(sheet.cell(i, column_index_from_string('O')).value)
                divn.append(sheet.cell(i, column_index_from_string('I')).value)
                mkt.append(sheet.cell(i, column_index_from_string('F')).value)
                stk.append(sheet.cell(i, column_index_from_string('A')).value)

    elif stock_type == "fmcgs":
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(i, column_index_from_string('B')).value == "CONSUMER GOODS":
                eps.append(sheet.cell(i, column_index_from_string('L')).value)
                pe.append(sheet.cell(i, column_index_from_string('G')).value)
                pb.append(sheet.cell(i, column_index_from_string('N')).value)
                bv.append(sheet.cell(i, column_index_from_string('H')).value)
                spe.append(sheet.cell(i, column_index_from_string('K')).value)
                divy.append(sheet.cell(i, column_index_from_string('O')).value)
                divn.append(sheet.cell(i, column_index_from_string('I')).value)
                mkt.append(sheet.cell(i, column_index_from_string('F')).value)
                stk.append(sheet.cell(i, column_index_from_string('A')).value)

    elif stock_type == "pharmas":
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(i, column_index_from_string('B')).value == "PHARMA":
                eps.append(sheet.cell(i, column_index_from_string('L')).value)
                pe.append(sheet.cell(i, column_index_from_string('G')).value)
                pb.append(sheet.cell(i, column_index_from_string('N')).value)
                bv.append(sheet.cell(i, column_index_from_string('H')).value)
                spe.append(sheet.cell(i, column_index_from_string('K')).value)
                divy.append(sheet.cell(i, column_index_from_string('O')).value)
                divn.append(sheet.cell(i, column_index_from_string('I')).value)
                mkt.append(sheet.cell(i, column_index_from_string('F')).value)
                stk.append(sheet.cell(i, column_index_from_string('A')).value)

    elif stock_type == "autos":
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(i, column_index_from_string('B')).value == "AUTOMOBILE":
                eps.append(sheet.cell(i, column_index_from_string('L')).value)
                pe.append(sheet.cell(i, column_index_from_string('G')).value)
                pb.append(sheet.cell(i, column_index_from_string('N')).value)
                bv.append(sheet.cell(i, column_index_from_string('H')).value)
                spe.append(sheet.cell(i, column_index_from_string('K')).value)
                divy.append(sheet.cell(i, column_index_from_string('O')).value)
                divn.append(sheet.cell(i, column_index_from_string('I')).value)
                mkt.append(sheet.cell(i, column_index_from_string('F')).value)
                stk.append(sheet.cell(i, column_index_from_string('A')).value)

    elif stock_type == "metals":
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(i, column_index_from_string('B')).value == "METALS":
                eps.append(sheet.cell(i, column_index_from_string('L')).value)
                pe.append(sheet.cell(i, column_index_from_string('G')).value)
                pb.append(sheet.cell(i, column_index_from_string('N')).value)
                bv.append(sheet.cell(i, column_index_from_string('H')).value)
                spe.append(sheet.cell(i, column_index_from_string('K')).value)
                divy.append(sheet.cell(i, column_index_from_string('O')).value)
                divn.append(sheet.cell(i, column_index_from_string('I')).value)
                mkt.append(sheet.cell(i, column_index_from_string('F')).value)
                stk.append(sheet.cell(i, column_index_from_string('A')).value)

    elif stock_type == "finances":
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(i, column_index_from_string('B')).value == "FINANCIAL SERVICES":
                eps.append(sheet.cell(i, column_index_from_string('L')).value)
                pe.append(sheet.cell(i, column_index_from_string('G')).value)
                pb.append(sheet.cell(i, column_index_from_string('N')).value)
                bv.append(sheet.cell(i, column_index_from_string('H')).value)
                spe.append(sheet.cell(i, column_index_from_string('K')).value)
                divy.append(sheet.cell(i, column_index_from_string('O')).value)
                divn.append(sheet.cell(i, column_index_from_string('I')).value)
                mkt.append(sheet.cell(i, column_index_from_string('F')).value)
                stk.append(sheet.cell(i, column_index_from_string('A')).value)

    elif stock_type == "oils":
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(i, column_index_from_string('B')).value == "OIL & GAS":
                eps.append(sheet.cell(i, column_index_from_string('L')).value)
                pe.append(sheet.cell(i, column_index_from_string('G')).value)
                pb.append(sheet.cell(i, column_index_from_string('N')).value)
                bv.append(sheet.cell(i, column_index_from_string('H')).value)
                spe.append(sheet.cell(i, column_index_from_string('K')).value)
                divy.append(sheet.cell(i, column_index_from_string('O')).value)
                divn.append(sheet.cell(i, column_index_from_string('I')).value)
                mkt.append(sheet.cell(i, column_index_from_string('F')).value)
                stk.append(sheet.cell(i, column_index_from_string('A')).value)

    return zip(stk, mkt, eps, pe, spe, pb, divn, divy, bv)


# Create your views here.
def index(request):
    try:
        banks_zip = ziping("banks")
        it_zip = ziping("its")
        fmcg_zip = ziping("fmcgs")
        pharma_zip = ziping("pharmas")
        auto_zip = ziping("autos")
        metal_zip = ziping("metals")
        finance_zip = ziping("finances")
        oil_zip = ziping("oils")

        dictt = {'banks_zips': banks_zip, 'it_zips': it_zip, 'fmcg_zips': fmcg_zip, 'pharma_zips': pharma_zip,
                 'auto_zips': auto_zip, 'metals_zips': metal_zip, 'finance_zips': finance_zip, 'oil_zips': oil_zip,
                 }
        wb = xl.load_workbook('login/users.xlsx')
        ip = get_client_ip(request)
        sheet = wb["Sheet1"]
        for i in range(2, sheet.max_row + 1):
            if (ip == sheet.cell(i, 3).value):
                if (sheet.cell(i, 4).value == "yes"):
                    print("matched")
                    dictt["email"] = sheet.cell(i, 1).value
        return render(request, 'peers.html', dictt)
    except Exception as e:
        wb = xl.load_workbook("errors.xlsx")
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        wb.save("errors.xlsx")
        return render(request, "oops.html")