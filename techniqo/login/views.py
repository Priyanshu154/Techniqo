import openpyxl as xl
from django.shortcuts import render
from django.shortcuts import redirect
from django.core.mail import send_mail
import random
# Create your views here.'
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip
def index(request):
    return render(request, "signin.html")
def generate_otp():
    otp = ""
    for i in range(0, 6):
        n = random.randint(0, 9)
        otp += str(n)
    return otp

def send_email_verify(email):
    subject = 'welcome to Techniqo'
    message = f'Hi ' + email + ',\n Thank you for registering in techniqo.\n Your OTP is : ' + generate_otp()
    email_from = "tempmoney1404@gmail.com"
    recipient_list = [email, ]
    send_mail(subject, message, email_from, recipient_list)
def save(request):
    wb = xl.load_workbook('login/users.xlsx')
    email = request.POST.get("email")
    password = request.POST.get("password")
    sheet = wb["Sheet1"]
    dict = {}
    dict["email"] = email
    flag = 0
    for i in range(2, sheet.max_row + 1):
        if (sheet.cell(i, 1).value == email):
            dict["email"] = "exists"
            print("matched")
            return render(request, "login.html", dict)
    if flag == 0:
        sheet.cell( sheet.max_row + 1 ,1).value = email
        sheet.cell( sheet.max_row ,2).value = password
        sheet.cell( sheet.max_row ,3).value = get_client_ip(request)
        sheet.cell( sheet.max_row ,4).value = "yes"
        sheet.cell( sheet.max_row ,5).value = request.POST.get("security_que")
        sheet.cell( sheet.max_row ,6).value = request.POST.get("ans").lower()
    wb.save("login/users.xlsx")
    response = redirect('/')
    return response
def login_redirect(request):
    return render(request, "login.html")
def login(request):
    wb = xl.load_workbook('login/users.xlsx')
    email = request.POST.get("email")
    password = request.POST.get("password")
    sheet = wb["Sheet1"]
    dict = {}
    dict["email"] = email
    for i in range(2, sheet.max_row + 1):
        if (sheet.cell(i, 1).value == email):
            if(sheet.cell(i, 2).value == password):
                sheet.cell(i,4).value = "yes"
                sheet.cell(i,3).value = get_client_ip(request)
                wb.save("login/users.xlsx")
                response = redirect('/')
                return response
    dict["email"] = "invalid"
    return render(request, "login.html", dict)

def logout(request):
    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
            if(sheet.cell(i, 3).value == ip):
                sheet.cell(i,4).value = "no"
                wb.save("login/users.xlsx")
    response = redirect('/')
    return response
def forgot_redirect(request):
    return render(request, "forgot.html")
def forgot(request):
    type = request.POST.get("security_que")
    email = request.POST.get("email")
    ans = request.POST.get("ans").lower()
    wb = xl.load_workbook('login/users.xlsx')

    sheet = wb["Sheet1"]
    dict = {
        "email": "invalid"
    }

    for i in range(2, sheet.max_row + 1):
        if(sheet.cell(i,1).value == email):
            if(sheet.cell(i,5).value == type):
                if(sheet.cell(i,6).value == ans):
                    response = redirect("/login/reset_redirect")
                    return response
                else:
                    dict = {"email": "invalid_secque"}
                    break
            else:
                dict = {"email": "invalid_secque"}
                break
    return render(request,"login.html", dict)
def reset_redirect(request):
    wb = xl.load_workbook('login/users.xlsx')
    ip = get_client_ip(request)
    sheet = wb["Sheet1"]
    dict = {}
    for i in range(2, sheet.max_row + 1):
        if (sheet.cell(i, 3).value == ip):
            dict['email'] = sheet.cell(i,1).value
    return  render(request, "reset.html", dict)
def reset(request):
    new_pass = request.POST.get("password")
    email = request.POST.get("email")
    wb = xl.load_workbook('login/users.xlsx')
    sheet = wb["Sheet1"]
    for i in range(2, sheet.max_row + 1):
        if (sheet.cell(i, 1).value == email):
            sheet.cell(i,2).value = new_pass
            sheet.cell(i,4).value = "yes"
            wb.save("login/users.xlsx")
            response = redirect("/")
            return response
    dict = {
        "email": email
    }
    return render(request,"homepage.html", dict)  #never gets executed but still for safety