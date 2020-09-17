from django.shortcuts import render
import openpyxl as xl
import datetime
from django.shortcuts import redirect
# Create your views here.

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def index(request):
    try:
        dict = {}
        pattern = request.GET.get("mark", "channel")
        dict["selected"] = pattern
        tit = ""
        if pattern == "channel":
            tit = "Channel Pattern in stocks | Breakout | Trading | Technical analysis | Scanner"
        elif pattern == "channelup":
            tit = "Channel up / Ascending channel Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "channeldown":
            tit = "Channel down / Descending channel Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "doubletop":
            tit = "Double Top Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "doublebottom":
            tit = "Double Bottom Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "multipletop":
            tit = "Multiple Top Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "multiplebottom":
            tit = "Multiple Bottom Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "head_and_shoulder":
            tit = "Head and Shoulder Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "wedge":
            tit = "Wedge Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "wedge up":
            tit = "Wedge Up Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "wedge down":
            tit = "Wedge Down Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "ascendingtriangle":
            tit = "Ascending Triangle Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "descendingtriangle":
            tit = "Descending Triangle Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "trendlinesupport":
            tit = "Trendline Support Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        elif pattern == "trendlineresistance":
            tit = "Trendline Resistance Pattern in stocks | Breakout |  Trading | Technical analysis | Scanner"
        dict["titl"] = tit
        wb = xl.load_workbook('login/users.xlsx')
        ip = get_client_ip(request)
        sheet = wb["Sheet1"]
        for i in range(2, sheet.max_row + 1):
            if ip == sheet.cell(i, 3).value:
                if sheet.cell(i, 4).value == "yes":
                    dict["email"] = sheet.cell(i, 1).value
                    return render(request, "channelh.html", dict)
        response = redirect('/login')
        return response
    except Exception as e:
        wb = xl.load_workbook("errors.xlsx")
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        wb.save("errors.xlsx")
        return render(request, "oops.html")