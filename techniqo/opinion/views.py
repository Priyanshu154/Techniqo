from django.shortcuts import render
import openpyxl as xl
from django.shortcuts import redirect
import datetime
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

# Create your views here.
def index(request):
    try:
        wb = xl.load_workbook('login/users.xlsx')
        ip = get_client_ip(request)
        sheet = wb["Sheet1"]
        dictb = {}
        for i in range(2, sheet.max_row + 1):
            if (ip == sheet.cell(i, 3).value):
                if (sheet.cell(i, 4).value == "yes"):
                    dictb["email"] = sheet.cell(i, 1).value
                    return render(request, 'opinion.html', dictb)
        response = redirect('/login')
        return response
    except Exception as e:
        wb = xl.load_workbook("errors.xlsx")
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        wb.save("errors.xlsx")
        return render(request, "oops.html")