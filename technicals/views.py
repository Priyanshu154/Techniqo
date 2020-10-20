from django.shortcuts import render
import os
import openpyxl as xl
from openpyxl.utils import column_index_from_string
import datetime
import traceback
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


# Create your views here.
def index(request):
    try:
        return render(request, 'technicals.html')
    except Exception as e:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, 'errors.xlsx')
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        wb.save(xx)
        return render(request, "oops.html")

def assign_para(tr):
    if tr == "rsi_below_30":
        return "are considered oversold. This implies that stock may rebound. Techniqo website helps to scan out the stocks having RSI above/greater than or below/less than or crossingover/crossingunder 10 20 30 40 50 60 70 80 "
    elif tr == "rsi_above_70":
        return "are considered overbought. This implies that stock may rebound. Techniqo website helps to scan out the stocks having RSI above/greater than or below/less than or crossingover/crossingunder 10 20 30 40 50 60 70 80 "
    elif tr == "rsi" or tr == "rsi_crossover_30" or tr == "rsi_crossunder_70":
        return "are considered neutral. This implies that stock may have good strength. Techniqo website helps to scan out the stocks having RSI above/greater than or below/less than or crossingover/crossingunder 10 20 30 40 50 60 70 80 "
    elif tr == "macd_above":
        return "is considered bullish . This implies that stock may have good strength. macd line above/greater than/crossing over zero line and signal line shows bullish trend for stock."
    elif tr == "macd_below":
        return "is considered bearish . This implies that stock may have weak strength. macd line below/less than/crossing under zero line and signal line shows bearish trend for stock."
    elif tr == "macd_crossover":
        return "is considered bullish . This implies that stock may have good strength. macd line above/greater than/crossing over zero line and signal line shows bullish trend for stock."
    elif tr == "macd_crossunder":
        return "is considered bearish . This implies that stock may have weak strength. macd line below/less than/crossing under zero line and signal line shows bearish trend for stock."


def assign_cell(cp):
    if cp == 'nifty_50':
        return 'AE'
    elif cp == 'nifty_100':
        return 'AF'
    elif cp == 'nifty_200':
        return 'AG'
    elif cp == 'nifty_next_50':
        return 'AH'
    elif cp == 'nifty_500':
        return 'AZ'
    elif cp == 'midcap_50':
        return 'AI'
    elif cp == 'midcap_100':
        return 'AJ'
    elif cp == 'midcap_150':
        return 'AK'
    elif cp == 'smallcap_50':
        return 'AL'
    elif cp == 'smallcap_100':
        return 'AM'
    elif cp == 'smallcap_250':
        return 'AN'
    return 'AZ'


def swapp(x, y):
    return y, x


def triggers(request):
    try:
        cp = request.GET.get("nifty", "nifty_500")
        tr = request.GET.get("tech", "rsi_below_30")
        # jab bhi directly ecxel sheet read karneki hogi next two points are compulsary
        workpath = os.path.dirname(os.path.abspath(__file__))
        xx = os.path.join(workpath, 'technicals_ours.xlsx')  # yaha tak
        wb = xl.load_workbook(xx, data_only=True)
        sheet = wb['sheet']
        stocks = []
        ticker = []
        value = []
        close = []

        ni = assign_cell(cp)
        num = 0

        if tr == "rsi_below_30":
            for i in range(2, sheet.max_row + 1):
                if 30 >= float(str(sheet.cell(i, column_index_from_string('E')).value)) > 0 and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('E')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Relative Strength Index (RSI) Below 30", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'RSI  Value', 'number': num, 'techh': tr, 'head2': 'Closing price'}



            return render(request, 'result_technicals.html', dictt)

        elif tr == "rsi_above_70":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('E')).value)) >= 70 and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('E')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Relative Strength Index (RSI) Above 70", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'RSI  Value', 'number': num, 'techh': tr, 'head2': 'Closing price'}

            return render(request, 'result_technicals.html', dictt)

        elif tr == "rsi":
            for i in range(2, sheet.max_row + 1):
                if 30 <= float(str(sheet.cell(i, column_index_from_string('E')).value)) <= 70 and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('E')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Relative Strength Index (RSI) Between 30 to 70", 'paras': para,
                     'signals': signal,
                     'nifty': cp, 'head': 'RSI  Value', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)

        elif tr == "rsi_crossover_30":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('D')).value)) <= 30 <= float(
                        str(sheet.cell(i, column_index_from_string('E')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('E')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Relative Strength Index (RSI) Crossing over 30", 'paras': para,
                     'signals': signal,
                     'nifty': cp, 'head': 'RSI  Value', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)

        elif tr == "rsi_crossunder_70":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('D')).value)) >= 70 >= float(
                        str(sheet.cell(i, column_index_from_string('E')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('E')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Relative Strength Index Crossing Under 70", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'RSI  Value', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)

        elif tr == "macd_above":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AX')).value)) >= 0 and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AX')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "MACD Line above 0 line", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'MACD Line', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)

        elif tr == "macd_below":
            for i in range(2, sheet.max_row + 1):
                if -10000 <= float(str(sheet.cell(i, column_index_from_string('AX')).value)) <= 0 and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AX')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "MACD Line below 0 line", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'MACD Line', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)

        elif tr == "macd_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AW')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AY')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AX')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AX')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "MACD Line Crossing over Signal line", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'MACD Line', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)

        elif tr == "macd_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AW')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AY')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AX')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AX')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "MACD Line Crossing under Signal line", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'MACD Line', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)

        elif tr == "above_lb":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('Z')).value)) <= float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('Z')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price above Lower Band", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Lower Band', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)

        elif tr == "mb_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AB')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AB')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over Middle Band", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Middle Band', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)

        elif tr == "mb_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AB')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AB')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under Middle Band", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Middle Band', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "ub_below":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AA')).value)) >= float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AA')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Below Upper bond", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Upper Band', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "ub_above":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AA')).value)) <= float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AA')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Above Upper bond", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Upper Band', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "lb_below":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('Z')).value)) >= float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('Z')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Below Lower bond", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Lower Band', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "pp_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('J')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('J')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over Pivot Point", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Pivot Point', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "pp_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('J')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('J')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing Under Pivot Point", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Pivot Point', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "r1_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('K')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('K')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over Resistance 1", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Resistance 1', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "r2_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('L')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('L')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over Resistance 2", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Resistance 2', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "r3_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('M')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('M')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over Resistance 3", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Resistance 3', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "r1_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('K')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('K')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under Resistance 1", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Resistance 1', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "r2_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('L')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('L')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing Under Resistance 2", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Resistance 2', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "r3_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('M')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('M')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under Resistance 3", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Resistance 3', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "s1_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('I')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('I')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over Support 1", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Support 1', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "s2_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('H')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('H')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over Support 2", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Support 2', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "s3_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('G')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('G')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over Support 3", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Support 3', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "s1_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('I')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('I')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under Support 1", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Support 1', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "s2_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('H')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('H')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing Under Support 2", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Support 2', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "s3_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('G')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('G')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under Support 3", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Support 3', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)

        elif tr == "50sma_200sma":
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BH')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('U')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('S')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BH')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('U')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('S')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('U')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('S')).value)), 2))
                        num = 1
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close)
                dictt = {'zips': zipp, 'trigger': "50 day SMA crossing over 200 day SMA ", 'paras': para, 'signals': signal,
                         'nifty': cp, 'head': 'SMA 200', 'number': num, 'techh': tr, 'head2': 'SMA 50'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "20sma_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('R')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('R')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over 20 Day SMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'SMA 20', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "50sma_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('S')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('S')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over 50 Day SMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'SMA 50', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "100sma_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('T')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('T')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over 100 Day SMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'SMA 100', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "200sma_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('U')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('R')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over 200 Day SMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'SMA 200', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "20sma_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('R')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('R')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under 20 Day SMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'SMA 20', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "50sma_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('S')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('S')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under 50 Day SMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'SMA 50', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "100sma_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('T')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('T')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under 100 Day SMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'SMA 100', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "200sma_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('U')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('U')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under 200 Day SMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'SMA 200', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "20_50sma_100sma":
            temp = []
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BH')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('BG')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('S')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BH')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('T')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('S')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes" and float(
                            str(sheet.cell(i, column_index_from_string('BG')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('T')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('R')).value)):
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('R')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('S')).value)), 2))
                        temp.append(round(float(str(sheet.cell(i, column_index_from_string('T')).value)), 2))
                        num = 2
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                            temp[h], temp[g] = swapp(temp[h], temp[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close, temp)
                dictt = {'zips': zipp, 'trigger': "50 day SMA and 20 day SMA crossing over 100 day SMA ", 'paras': para,
                         'signals': signal,
                         'nifty': cp, 'head': 'SMA 20', 'number': num, 'techh': tr, 'head2': 'SMA 50', 'head3': 'SMA 100'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "20_50_100sma_200sma":
            temp = []
            temp2 = []
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BH')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('BG')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('BI')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BH')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('U')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('S')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes" and float(
                            str(sheet.cell(i, column_index_from_string('BG')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('U')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('R')).value)) and float(
                            str(sheet.cell(i, column_index_from_string('BI')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('U')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('T')).value)):
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('R')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('S')).value)), 2))
                        temp.append(round(float(str(sheet.cell(i, column_index_from_string('T')).value)), 2))
                        temp2.append(round(float(str(sheet.cell(i, column_index_from_string('U')).value)), 2))
                        num = 3
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                            temp[h], temp[g] = swapp(temp[h], temp[g])
                            temp2[h], temp2[g] = swapp(temp2[h], temp2[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close, temp)
                dictt = {'zips': zipp, 'trigger': "50 day SMA and 20 day SMA and 100 day SMA crossing over 200 day SMA ",
                         'paras': para, 'signals': signal,
                         'nifty': cp, 'head': 'SMA 20', 'number': num, 'techh': tr, 'head2': 'SMA 50', 'head3': 'SMA 100',
                         'head4': 'SMA 200'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "20sma_50sma":
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BG')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BG')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('S')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('R')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('R')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('S')).value)), 2))
                        num = 1
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close)
                dictt = {'zips': zipp, 'trigger': "20 day SMA crossing over 50 day SMA ", 'paras': para, 'signals': signal,
                         'nifty': cp, 'head': 'SMA 20', 'number': num, 'techh': tr, 'head2': 'SMA 50'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "20_50sma_100smau":
            temp = []
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BH')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('BG')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('S')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BH')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('T')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('S')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes" and float(
                            str(sheet.cell(i, column_index_from_string('BG')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('T')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('R')).value)):
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('R')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('S')).value)), 2))
                        temp.append(round(float(str(sheet.cell(i, column_index_from_string('T')).value)), 2))
                        num = 2
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                            temp[h], temp[g] = swapp(temp[h], temp[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close, temp)
                dictt = {'zips': zipp, 'trigger': "50 day SMA and 20 day SMA crossing under 100 day SMA ", 'paras': para,
                         'signals': signal,
                         'nifty': cp, 'head': 'SMA 20', 'number': num, 'techh': tr, 'head2': 'SMA 50', 'head3': 'SMA 100'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "20_50_100sma_200smau":
            temp = []
            temp2 = []
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BH')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('BG')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('BI')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BH')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('U')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('S')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes" and float(
                            str(sheet.cell(i, column_index_from_string('BG')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('U')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('R')).value)) and float(
                            str(sheet.cell(i, column_index_from_string('BI')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('U')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('T')).value)):
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('R')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('S')).value)), 2))
                        temp.append(round(float(str(sheet.cell(i, column_index_from_string('T')).value)), 2))
                        temp2.append(round(float(str(sheet.cell(i, column_index_from_string('U')).value)), 2))
                        num = 3
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                            temp[h], temp[g] = swapp(temp[h], temp[g])
                            temp2[h], temp2[g] = swapp(temp2[h], temp2[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close, temp)
                dictt = {'zips': zipp, 'trigger': "50 day SMA and 20 day SMA and 100 day SMA crossing under 200 day SMA ",
                         'paras': para, 'signals': signal,
                         'nifty': cp, 'head': 'SMA 20', 'number': num, 'techh': tr, 'head2': 'SMA 50', 'head3': 'SMA 100',
                         'head4': 'SMA 200'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "20sma_50smau":
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BG')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BG')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('S')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('R')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('R')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('S')).value)), 2))
                        num = 1
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close)
                dictt = {'zips': zipp, 'trigger': "20 day SMA crossing under 50 day SMA ", 'paras': para, 'signals': signal,
                         'nifty': cp, 'head': 'SMA 20', 'number': num, 'techh': tr, 'head2': 'SMA 50'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "50sma_200smau":
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BH')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('U')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('S')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BH')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('U')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('S')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('U')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('S')).value)), 2))
                        num = 1
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close)
                dictt = {'zips': zipp, 'trigger': "50 day SMA crossing under 200 day SMA ", 'paras': para,
                         'signals': signal,
                         'nifty': cp, 'head': 'SMA 200', 'number': num, 'techh': tr, 'head2': 'SMA 50'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "50EMA_200EMA":
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BD')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('Q')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('O')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BD')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('Q')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('O')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('Q')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('O')).value)), 2))
                        num = 1
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close)
                dictt = {'zips': zipp, 'trigger': "50 day EMA crossing over 200 day EMA ", 'paras': para, 'signals': signal,
                         'nifty': cp, 'head': 'EMA 200', 'number': num, 'techh': tr, 'head2': 'EMA 50'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "20EMA_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('N')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('N')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over 20 Day EMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'EMA 20', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "50EMA_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('S')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('S')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over 50 Day EMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'EMA 50', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "100EMA_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('P')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('P')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over 100 Day EMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'EMA 100', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "200EMA_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('Q')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('Q')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing over 200 Day EMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'EMA 200', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "20EMA_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('N')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('N')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under 20 Day EMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'EMA 20', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "50EMA_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('O')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('O')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under 50 Day EMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'EMA 50', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "100EMA_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('P')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('P')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under 100 Day EMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'EMA 100', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "200EMA_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('Q')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('Q')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price Crossing under 200 Day EMA", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'EMA 200', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "20_50EMA_100EMA":
            temp = []
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BD')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('BC')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('O')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BD')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('P')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('O')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes" and float(
                            str(sheet.cell(i, column_index_from_string('BC')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('P')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('N')).value)):
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('N')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('O')).value)), 2))
                        temp.append(round(float(str(sheet.cell(i, column_index_from_string('P')).value)), 2))
                        num = 2
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                            temp[h], temp[g] = swapp(temp[h], temp[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close, temp)
                dictt = {'zips': zipp, 'trigger': "50 day EMA and 20 day EMA crossing over 100 day EMA ", 'paras': para,
                         'signals': signal,
                         'nifty': cp, 'head': 'EMA 20', 'number': num, 'techh': tr, 'head2': 'EMA 50', 'head3': 'EMA 100'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "20_50_100EMA_200EMA":
            temp = []
            temp2 = []
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BD')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('BC')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('BE')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BD')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('Q')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('O')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes" and float(
                            str(sheet.cell(i, column_index_from_string('BC')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('Q')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('N')).value)) and float(
                            str(sheet.cell(i, column_index_from_string('BE')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('Q')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('P')).value)):
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('N')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('O')).value)), 2))
                        temp.append(round(float(str(sheet.cell(i, column_index_from_string('P')).value)), 2))
                        temp2.append(round(float(str(sheet.cell(i, column_index_from_string('Q')).value)), 2))
                        num = 3
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                            temp[h], temp[g] = swapp(temp[h], temp[g])
                            temp2[h], temp2[g] = swapp(temp2[h], temp2[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close, temp)
                dictt = {'zips': zipp, 'trigger': "50 day EMA and 20 day EMA and 100 day EMA crossing over 200 day EMA ",
                         'paras': para, 'signals': signal,
                         'nifty': cp, 'head': 'EMA 20', 'number': num, 'techh': tr, 'head2': 'EMA 50', 'head3': 'EMA 100',
                         'head4': 'EMA 200'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "20EMA_50EMA":
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BC')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BC')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('O')).value)) < float(
                            str(sheet.cell(i, column_index_from_string('N')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('N')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('O')).value)), 2))
                        num = 1
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close)
                dictt = {'zips': zipp, 'trigger': "20 day EMA crossing over 50 day EMA ", 'paras': para, 'signals': signal,
                         'nifty': cp, 'head': 'EMA 20', 'number': num, 'techh': tr, 'head2': 'EMA 50'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "20_50EMA_100EMAu":
            temp = []
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BD')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('BC')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('O')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BD')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('P')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('S')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes" and float(
                            str(sheet.cell(i, column_index_from_string('BC')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('P')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('N')).value)):
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('N')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('O')).value)), 2))
                        temp.append(round(float(str(sheet.cell(i, column_index_from_string('P')).value)), 2))
                        num = 2
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                            temp[h], temp[g] = swapp(temp[h], temp[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close, temp)
                dictt = {'zips': zipp, 'trigger': "50 day EMA and 20 day EMA crossing under 100 day EMA ", 'paras': para,
                         'signals': signal,
                         'nifty': cp, 'head': 'EMA 20', 'number': num, 'techh': tr, 'head2': 'EMA 50', 'head3': 'EMA 100'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "20_50_100EMA_200EMAu":
            temp = []
            temp2 = []
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BD')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('BC')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('BE')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BD')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('Q')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('O')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes" and float(
                            str(sheet.cell(i, column_index_from_string('BC')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('Q')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('N')).value)) and float(
                            str(sheet.cell(i, column_index_from_string('BE')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('Q')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('P')).value)):
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('N')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('O')).value)), 2))
                        temp.append(round(float(str(sheet.cell(i, column_index_from_string('P')).value)), 2))
                        temp2.append(round(float(str(sheet.cell(i, column_index_from_string('Q')).value)), 2))
                        num = 3
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                            temp[h], temp[g] = swapp(temp[h], temp[g])
                            temp2[h], temp2[g] = swapp(temp2[h], temp2[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close, temp)
                dictt = {'zips': zipp, 'trigger': "50 day EMA and 20 day EMA and 100 day EMA crossing under 200 day EMA ",
                         'paras': para, 'signals': signal,
                         'nifty': cp, 'head': 'EMA 20', 'number': num, 'techh': tr, 'head2': 'EMA 50', 'head3': 'EMA 100',
                         'head4': 'EMA 200'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "20EMA_50EMAu":
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BC')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BC')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('O')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('N')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('N')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('O')).value)), 2))
                        num = 1
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close)
                dictt = {'zips': zipp, 'trigger': "20 day EMA crossing under 50 day EMA ", 'paras': para, 'signals': signal,
                         'nifty': cp, 'head': 'EMA 20', 'number': num, 'techh': tr, 'head2': 'EMA 50'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "50EMA_200EMAu":
            for i in range(2, sheet.max_row + 1):
                if str(sheet.cell(i, column_index_from_string('BD')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('Q')).value) is not None and str(
                        sheet.cell(i, column_index_from_string('O')).value) is not None:
                    if float(str(sheet.cell(i, column_index_from_string('BD')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('Q')).value)) > float(
                            str(sheet.cell(i, column_index_from_string('O')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                        stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                        ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                        value.append(round(float(str(sheet.cell(i, column_index_from_string('Q')).value)), 2))
                        close.append(round(float(str(sheet.cell(i, column_index_from_string('O')).value)), 2))
                        num = 1
                for g in range(len(value)):
                    for h in range(g, len(value)):
                        if value[h] < value[g]:
                            value[h], value[g] = swapp(value[h], value[g])
                            stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                            ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                            close[h], close[g] = swapp(close[h], close[g])
                signal = "Please select index by default result shows Nifty 500 stocks"
                para = assign_para(tr)
                zipp = zip(stocks, ticker, value, close)
                dictt = {'zips': zipp, 'trigger': "50 day EMA crossing under 200 day EMA ", 'paras': para,
                         'signals': signal,
                         'nifty': cp, 'head': 'EMA 200', 'number': num, 'techh': tr, 'head2': 'EMA 50'}
    

            return render(request, 'result_technicals.html', dictt)


        elif tr == "MFI_below_30":
            for i in range(2, sheet.max_row + 1):
                if 30 >= float(str(sheet.cell(i, column_index_from_string('W')).value)) > 0 and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('W')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Money Flow Index (MFI) Below 30", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'MFI  Value', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "MFI_above_70":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('W')).value)) >= 70 and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('W')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Money Flow Index (MFI) Above 70", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'MFI  Value', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "MFI":
            for i in range(2, sheet.max_row + 1):
                if 30 <= float(str(sheet.cell(i, column_index_from_string('W')).value)) <= 70 and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('W')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Money Flow Index (MFI) Between 30 to 70", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'MFI  Value', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "MFI_crossover_30":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('V')).value)) <= 30 <= float(
                        str(sheet.cell(i, column_index_from_string('W')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('W')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Money Flow Index (MFI) Crossing over 30", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'MFI  Value', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "MFI_crossunder_70":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('V')).value)) >= 70 >= float(
                        str(sheet.cell(i, column_index_from_string('W')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('W')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Money Flow Index Crossing Under 70", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'MFI  Value', 'number': num, 'techh': tr, 'head2': 'Closing price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "fl_sl_crossover":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('BL')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AV')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AU')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes" and float(
                        str(sheet.cell(i, column_index_from_string('BL')).value)) > 0 and float(
                        str(sheet.cell(i, column_index_from_string('AU')).value)) < 30:
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AV')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Fast Line crossing over Slow Line below 30", 'paras': para,
                     'signals': signal,
                     'nifty': cp, 'head': 'Fast Line Value', 'number': num, 'techh': tr, 'head2': 'Slow Line Value'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "fl_sl_crossunder":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('BL')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AV')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AU')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes" and float(
                        str(sheet.cell(i, column_index_from_string('BL')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AU')).value)) > 70:
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AV')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Fast Line crossing under Slow Line above 70", 'paras': para,
                     'signals': signal,
                     'nifty': cp, 'head': 'Fast Line Value', 'number': num, 'techh': tr, 'head2': 'Slow Line Value'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "fl_sl_crossovern":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('BL')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AV')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AU')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes" and float(
                        str(sheet.cell(i, column_index_from_string('BL')).value)) > 0:
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AV')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Fast Line crossing over Slow Line", 'paras': para, 'signals': signal,
                     'nifty': cp, 'head': 'Fast Line Value', 'number': num, 'techh': tr, 'head2': 'Slow Line Value'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "fl_sl_crossundern":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('BL')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AV')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AU')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes" and float(
                        str(sheet.cell(i, column_index_from_string('BL')).value)) > 0:
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AV')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Fast Line crossing under Slow Line above 70", 'paras': para,
                     'signals': signal,
                     'nifty': cp, 'head': 'Fast Line Value', 'number': num, 'techh': tr, 'head2': 'Slow Line Value'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "fl_below":
            for i in range(2, sheet.max_row + 1):
                if 30 > float(str(sheet.cell(i, column_index_from_string('AU')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AV')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Fast Line crossing under Slow Line above 70", 'paras': para,
                     'signals': signal,
                     'nifty': cp, 'head': 'Fast Line Value', 'number': num, 'techh': tr, 'head2': 'Slow Line Value'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "fl_above":
            for i in range(2, sheet.max_row + 1):
                if 70 < float(str(sheet.cell(i, column_index_from_string('AU')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AV')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Fast Line crossing under Slow Line above 70", 'paras': para,
                     'signals': signal,
                     'nifty': cp, 'head': 'Fast Line Value', 'number': num, 'techh': tr, 'head2': 'Slow Line Value'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "conl_basel_crossover":
            temp = []
            for i in range(2, sheet.max_row + 1):
                if 0 <= float(str(sheet.cell(i, column_index_from_string('BA')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AQ')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AP')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AP')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AQ')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Conversion Line Crossing over Base line", 'paras': para,
                     'signals': signal,
                     'nifty': cp, 'head': 'Conversion Line', 'number': num, 'techh': tr, 'head2': 'Base Line'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "conl_basel_crossunder":
            for i in range(2, sheet.max_row + 1):
                if 0 <= float(str(sheet.cell(i, column_index_from_string('BA')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AQ')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AP')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AP')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AQ')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Conversion Line Crossing Under Base line", 'paras': para,
                     'signals': signal,
                     'nifty': cp, 'head': 'Conversion Line', 'number': num, 'techh': tr, 'head2': 'Base Line'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "ic_bull":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AC')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AR')).value)) > float(
                        str(sheet.cell(i, column_index_from_string('AS')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AR')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price above Green Cloud", 'paras': para,
                     'signals': signal,
                     'nifty': cp, 'head': 'Cloud Top', 'number': num, 'techh': tr, 'head2': 'Closing Price'}


            return render(request, 'result_technicals.html', dictt)


        elif tr == "ic_bear":
            for i in range(2, sheet.max_row + 1):
                if float(str(sheet.cell(i, column_index_from_string('AC')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AR')).value)) < float(
                        str(sheet.cell(i, column_index_from_string('AS')).value)) and str(
                        sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                    stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                    ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                    value.append(round(float(str(sheet.cell(i, column_index_from_string('AR')).value)), 2))
                    close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    num = 1
            for g in range(len(value)):
                for h in range(g, len(value)):
                    if value[h] < value[g]:
                        value[h], value[g] = swapp(value[h], value[g])
                        stocks[h], stocks[g] = swapp(stocks[h], stocks[g])
                        ticker[h], ticker[g] = swapp(ticker[h], ticker[g])
                        close[h], close[g] = swapp(close[h], close[g])
            signal = "Please select index by default result shows Nifty 500 stocks"
            para = assign_para(tr)
            zipp = zip(stocks, ticker, value, close)
            dictt = {'zips': zipp, 'trigger': "Price below Red Cloud", 'paras': para,
                     'signals': signal,
                     'nifty': cp, 'head': 'Cloud Bottom', 'number': num, 'techh': tr, 'head2': 'Closing Price'}


            return render(request, 'result_technicals.html', dictt)
    except Exception as e:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, 'errors.xlsx')
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row , 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")
