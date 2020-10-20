import requests
from bs4 import BeautifulSoup
import openpyxl as xl

wb = xl.load_workbook('market_data_20.xlsx')
sheet = wb["sheet"]
xx = "abcdefghijklmnopqrstuvwxyz"
count = 1
p = 0
for i in range(117):
    url ="https://ticker.finology.in/Directory/Stocks/"+str(i)
    r = requests.get(url)
    htmlcontent = r.content
    soup = BeautifulSoup(htmlcontent, 'html.parser')
    for j in range(0,100):
        try:
            x = soup.find_all("h4")[j].get_text().strip()
            for h in range(2, sheet.max_row+1):
                if sheet.cell(h, 3).value == x or sheet.cell(h, 4).value == x:
                    sheet.cell(h, 10).value = "https://ticker.finology.in"+soup.find_all("div", {"class": "sparkchartwrapper"})[j].find("a",href = True)["href"]
                    print(count)
                    count = count+1
                    break
        except:
            pass
wb.save("market_data_20.xlsx")
