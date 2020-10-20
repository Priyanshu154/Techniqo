from django.shortcuts import render
import openpyxl as xl
from django.contrib.auth.decorators import login_required
import traceback
import datetime
import calendar
import math
import os
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

# Create your views here.
@login_required(login_url="/accounts/login")
def index(request):
    try:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, 'candlepattern/nifty_100.xlsx')  # yaha tak
        wb2 = xl.load_workbook(xx, data_only=True)
        sheet2 = wb2['Sheet1']
        name_ticker = []
        type_opinion = []
        type_opinion_value = []
        target_value = []
        for i in range(2, sheet2.max_row + 1):
            if sheet2.cell(i, 12).value == "YES" and math.ceil(int(sheet2.cell(i, 3).value)) > 200:
                name_ticker.append(sheet2.cell(i, 1).value)
                type_opinion.append("Buy")
                type_opinion_value.append(math.ceil(int(sheet2.cell(i, 3).value)))
                target_value.append(round(int(sheet2.cell(i, 3).value)*100.5/100))
        for i in range(2, sheet2.max_row + 1):
            if sheet2.cell(i, 13).value == "YES" and math.ceil(int(sheet2.cell(i, 3).value)) > 200:
                name_ticker.append(sheet2.cell(i, 1).value)
                type_opinion.append("Buy")
                type_opinion_value.append(math.ceil(int(sheet2.cell(i, 3).value)))
                target_value.append(round(int(sheet2.cell(i, 3).value)*100.5/100))
        for i in range(2, sheet2.max_row + 1):
            if sheet2.cell(i, 14).value == "YES" and math.ceil(int(sheet2.cell(i, 3).value)) > 200:
                name_ticker.append(sheet2.cell(i, 1).value)
                type_opinion.append("Sell")
                type_opinion_value.append(math.ceil(int(sheet2.cell(i, 4).value)))
                target_value.append(round(int(sheet2.cell(i, 4).value)*99.5/100))
        for i in range(2, sheet2.max_row + 1):
            if sheet2.cell(i, 15).value == "YES" and math.ceil(int(sheet2.cell(i, 3).value)) > 200:
                name_ticker.append(sheet2.cell(i, 1).value)
                type_opinion.append("Sell")
                type_opinion_value.append(math.ceil(int(sheet2.cell(i, 4).value)))
                target_value.append(round(int(sheet2.cell(i, 4).value)*99.5/100))
        for i in range(2, sheet2.max_row + 1):
            if sheet2.cell(i, 16).value == "YES" and math.ceil(int(sheet2.cell(i, 3).value)) > 200:
                name_ticker.append(sheet2.cell(i, 1).value)
                type_opinion.append("Buy")
                type_opinion_value.append(math.ceil(int(sheet2.cell(i, 3).value)))
                target_value.append(round(int(sheet2.cell(i, 3).value)*100.5/100))
        xy = os.path.join(workpath, 'data_new_ticker/3MINDIA.xlsx')  # yaha tak
        wb3 = xl.load_workbook(xy, data_only=True)
        sheet3 = wb3['Sheet1']
        date_used = sheet3.cell(sheet3.max_row, 1).value
        date_need = str(date_used).split('-')
        date_daypr = date_need[2] + "-" + date_need[1] + "-" + date_need[0]
        date_op = datetime.datetime.strptime(date_daypr, '%Y-%m-%d') + datetime.timedelta(days=1)
        date_need = str(date_op).split(" ")[0].split('-')
        date_day = date_need[2]+" "+date_need[1]+" "+date_need[0]
        born = datetime.datetime.strptime(date_day, '%d %m %Y').weekday()
        if calendar.day_name[born] == "Saturday":
            date_op = datetime.datetime.strptime(date_daypr, '%Y-%m-%d') + datetime.timedelta(days=3)
        date_op = date_op.strftime("%d-%m-%Y")
        dictb = {'name_ticker': name_ticker, 'type_opinion': type_opinion, 'type_opinion_value': type_opinion_value, 'target_value': target_value,
                 'date_used': date_op , "kitne": len(name_ticker)}
    
        return render(request, 'opinion.html', dictb)
    except Exception as e:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, f'errors.xlsx')  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row, 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row, 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")
