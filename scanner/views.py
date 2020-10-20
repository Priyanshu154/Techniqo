from django.shortcuts import render
import os
import pandas_datareader.data as web
from datetime import datetime
from matplotlib.dates import date2num
import statistics
import openpyxl as xl
from openpyxl.utils import column_index_from_string
import datetime
import traceback
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


# Create your views here.
def comon(a, b):
    a_set = set(a)
    b_set = set(b)
    if (a_set & b_set):
        return list(a_set & b_set)
    else:
        return []


def assign_cell(cp):
    if cp == 'nifty_50':
        return 'AE'
    elif cp == 'nifty_100':
        return 'AF'
    elif cp == 'nifty_200':
        return 'AG'
    elif cp == 'nifty_next_50':
        return 'AH'
    elif cp == 'nifty_500':
        return 'AZ'
    elif cp == 'midcap_50':
        return 'AI'
    elif cp == 'midcap_100':
        return 'AJ'
    elif cp == 'midcap_150':
        return 'AK'
    elif cp == 'smallcap_50':
        return 'AL'
    elif cp == 'smallcap_100':
        return 'AM'
    elif cp == 'smallcap_250':
        return 'AN'
    return 'AZ'


def swapp(x, y):
    return y, x


def techscreen(request):
    try:
        dataa = request.POST.get('backe', 'default')
        inter = dataa.split(",||")
        indicator_entry_str = inter[0]
        parameter_entry_str = inter[1]
        value_entry_str = inter[2]
        cp = inter[3]
        parameter_entry_str = parameter_entry_str[1:]
        indicator_entry = indicator_entry_str.split(",")
        parameter_entry = parameter_entry_str.split(",")
        value_entry = value_entry_str.split(",")

       #workpath = os.path.dirname(os.path.abspath(__file__))
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, 'technicals/technicals_ours.xlsx')  # yaha tak
        wb = xl.load_workbook(xx, data_only=True)
        sheet = wb['sheet']
        stocks = []
        ticker = []
        value = []
        close = []
        stocks2 = []
        ticker2 = []
        value2 = []
        close2 = []
        ni = assign_cell(cp)

        for ii in range(len(indicator_entry)):
            if indicator_entry[ii] == "rsi":
                if parameter_entry[ii] == "crossover":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('D')).value)) < float(value_entry[ii]) < float(
                                str(sheet.cell(i, column_index_from_string('E')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('E')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "crossunder":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('D')).value)) > float(value_entry[ii]) > float(
                                str(sheet.cell(i, column_index_from_string('E')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('E')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "above":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('E')).value)) > float(value_entry[ii]) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('E')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "below":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('E')).value)) < float(value_entry[ii]) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('E')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)

            elif indicator_entry[ii] == "clo" and (
                    value_entry[ii] == "lb" or value_entry[ii] == "mb" or value_entry[ii] == "up"):
                if parameter_entry[ii] == "crossover":
                    if value_entry[ii] == "lb":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                                    str(sheet.cell(i, column_index_from_string('Z')).value)) < float(
                                str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('Z')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    elif value_entry[ii] == "mb":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                                    str(sheet.cell(i, column_index_from_string('AB')).value)) < float(
                                str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AB')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    elif value_entry[ii] == "up":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AD')).value)) < float(
                                    str(sheet.cell(i, column_index_from_string('AA')).value)) < float(
                                str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AA')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                elif parameter_entry[ii] == "crossunder":
                    if value_entry[ii] == "lb":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                                    str(sheet.cell(i, column_index_from_string('Z')).value)) > float(
                                str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('Z')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    elif value_entry[ii] == "mb":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                                    str(sheet.cell(i, column_index_from_string('AB')).value)) > float(
                                str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AB')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    elif value_entry[ii] == "up":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AD')).value)) > float(
                                    str(sheet.cell(i, column_index_from_string('AA')).value)) > float(
                                str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AA')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                elif parameter_entry[ii] == "above":
                    if value_entry[ii] == "lb":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('Z')).value)) < float(
                                    str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('Z')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    elif value_entry[ii] == "mb":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AB')).value)) < float(
                                    str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AB')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    elif value_entry[ii] == "up":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AA')).value)) < float(
                                    str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AA')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                elif parameter_entry[ii] == "below":
                    if value_entry[ii] == "lb":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('Z')).value)) > float(
                                    str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('Z')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    elif value_entry[ii] == "mb":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AB')).value)) > float(
                                    str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AB')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    elif value_entry[ii] == "up":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AA')).value)) > float(
                                    str(sheet.cell(i, column_index_from_string('AC')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AA')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)

            elif indicator_entry[ii] == "macd":
                if parameter_entry[ii] == "crossover":
                    if value_entry[ii] == "sig":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AW')).value)) < float(
                                    str(sheet.cell(i, column_index_from_string('AY')).value)) < float(
                                str(sheet.cell(i, column_index_from_string('AX')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AX')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    elif value_entry[ii] == "zero":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AW')).value)) < 0 < float(
                                    str(sheet.cell(i, column_index_from_string('AX')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AX')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                elif parameter_entry[ii] == "crossunder":
                    if value_entry[ii] == "sig":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AW')).value)) > float(
                                    str(sheet.cell(i, column_index_from_string('AY')).value)) > float(
                                str(sheet.cell(i, column_index_from_string('AX')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AX')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    elif value_entry[ii] == "zero":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AW')).value)) > 0 > float(
                                    str(sheet.cell(i, column_index_from_string('AX')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AX')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                elif parameter_entry[ii] == "above":
                    if value_entry[ii] == "sig":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AY')).value)) < float(
                                    str(sheet.cell(i, column_index_from_string('AX')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AX')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    elif value_entry[ii] == "zero":
                        for i in range(2, sheet.max_row + 1):
                            if 0 < float(str(sheet.cell(i, column_index_from_string('AX')).value)) and str(
                                    sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AX')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                elif parameter_entry[ii] == "below":
                    if value_entry[ii] == "sig":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AY')).value)) > float(
                                    str(sheet.cell(i, column_index_from_string('AX')).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AX')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    elif value_entry[ii] == "zero":
                        for i in range(2, sheet.max_row + 1):
                            if 0 > float(str(sheet.cell(i, column_index_from_string('AX')).value)) and str(
                                    sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AX')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
            elif indicator_entry[ii] == "mfi":
                if parameter_entry[ii] == "crossover":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('V')).value)) < float(value_entry[ii]) < float(
                                str(sheet.cell(i, column_index_from_string('W')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('W')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "crossunder":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('V')).value)) > float(value_entry[ii]) > float(
                                str(sheet.cell(i, column_index_from_string('W')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('W')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "above":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('W')).value)) > float(value_entry[ii]) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('W')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "below":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('W')).value)) < float(value_entry[ii]) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('W')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
            elif indicator_entry[ii] == "roc":
                if parameter_entry[ii] == "crossover":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('BO')).value)) < float(value_entry[ii]) < float(
                                str(sheet.cell(i, column_index_from_string('X')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('X')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "crossunder":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('BO')).value)) > float(value_entry[ii]) > float(
                                str(sheet.cell(i, column_index_from_string('X')).value)) and str(sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('X')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "above":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('X')).value)) > float(value_entry[ii]) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('X')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "below":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('X')).value)) < float(value_entry[ii]) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('X')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
            elif indicator_entry[ii] == "srsi":
                if parameter_entry[ii] == "crossover":
                    if value_entry[ii] == "slow":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('BL')).value)) < float(str(sheet.cell(i, column_index_from_string('AV')).value))\
                                 < float(str(sheet.cell(i, column_index_from_string('AU')).value)) and str(sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    else:
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('BL')).value)) < float(value_entry[ii])\
                                 < float(str(sheet.cell(i, column_index_from_string('AU')).value)) and str(sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                elif parameter_entry[ii] == "crossunder":
                    if value_entry[ii] == "slow":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('BL')).value)) > float(str(sheet.cell(i, column_index_from_string('AV')).value))\
                                 > float(str(sheet.cell(i, column_index_from_string('AU')).value)) and str(sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    else:
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('BL')).value)) > float(value_entry[ii])\
                                 > float(str(sheet.cell(i, column_index_from_string('AU')).value)) and str(sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                elif parameter_entry[ii] == "above":
                    if value_entry[ii] == "slow":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AV')).value))\
                                 < float(str(sheet.cell(i, column_index_from_string('AU')).value)) and str(sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    else:
                        for i in range(2, sheet.max_row + 1):
                            if float(value_entry[ii])\
                                 < float(str(sheet.cell(i, column_index_from_string('AU')).value)) and str(sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                elif parameter_entry[ii] == "below":
                    if value_entry[ii] == "slow":
                        for i in range(2, sheet.max_row + 1):
                            if float(str(sheet.cell(i, column_index_from_string('AV')).value))\
                                 > float(str(sheet.cell(i, column_index_from_string('AU')).value)) and str(sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
                    else:
                        for i in range(2, sheet.max_row + 1):
                            if float(value_entry[ii])\
                                 > float(str(sheet.cell(i, column_index_from_string('AU')).value)) and str(sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string('AU')).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                        if len(stocks2) == 0:
                            stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                            stocks, ticker, value, close = [], [], [], []
                        else:
                            stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                     value2), \
                                                               comon(close, close2)
            elif indicator_entry[ii] == "wil":
                if parameter_entry[ii] == "crossover":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('BN')).value)) < float(value_entry[ii]) < float(
                                str(sheet.cell(i, column_index_from_string('Y')).value)) and str(
                            sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('Y')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "crossunder":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('BN')).value)) > float(value_entry[ii]) > float(
                                str(sheet.cell(i, column_index_from_string('Y')).value)) and str(sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('Y')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "above":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('Y')).value)) > float(value_entry[ii]) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('Y')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "below":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string('Y')).value)) < float(value_entry[ii]) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string('Y')).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
            elif indicator_entry[ii] == "sma10" or indicator_entry[ii] == "sma20" or indicator_entry[ii] == "sma50" or \
                    indicator_entry[ii] == "sma100" or indicator_entry[ii] == "sma200" or (
                    indicator_entry[ii] == "clo" and (
                    value_entry[ii] == "sma10" or value_entry[ii] == "sma20" or value_entry[ii] == "sma50" or value_entry[
                ii] == "sma100" or value_entry[ii] == "sma200")):
                val = ""
                valp = ""
                val2 = ""
                val2p = ""
                if value_entry[ii] == "sma20":
                    val2 = "R"
                    val2p = "BG"
                elif value_entry[ii] == "sma50":
                    val2 = "S"
                    val2p = "BH"
                elif value_entry[ii] == "sma100":
                    val2 = "T"
                    val2p = "BI"
                elif value_entry[ii] == "sma200":
                    val2 = "U"
                    val2p = "BJ"
                elif value_entry[ii] == "sma10":
                    val2 = "BP"
                    val2p = "BQ"

                if indicator_entry[ii] == "sma20":
                    val = "R"
                    valp = "BG"
                elif indicator_entry[ii] == "sma50":
                    val = "S"
                    valp = "BH"
                elif indicator_entry[ii] == "sma100":
                    val = "T"
                    valp = "BI"
                elif indicator_entry[ii] == "sma200":
                    val = "U"
                    valp = "BJ"
                elif indicator_entry[ii] == "sma10":
                    val = "BP"
                    valp = "BQ"
                elif indicator_entry[ii] == "clo":
                    val = "AC"
                    valp = "AD"
                if parameter_entry[ii] == "crossover":
                    for i in range(2, sheet.max_row + 1):
                        if str(sheet.cell(i, column_index_from_string(valp)).value) != "None" and str(
                                sheet.cell(i, column_index_from_string(val2)).value) != "None" and str(
                                sheet.cell(i, column_index_from_string(val)).value) != "None":
                            if float(str(sheet.cell(i, column_index_from_string(valp)).value)) < float(
                                    str(sheet.cell(i, column_index_from_string(val2)).value)) < float(
                                    str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                    sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "crossunder":
                    for i in range(2, sheet.max_row + 1):
                        if str(sheet.cell(i, column_index_from_string(valp)).value) != "None" and str(
                                sheet.cell(i, column_index_from_string(val2)).value) != "None" and str(
                            sheet.cell(i, column_index_from_string(val)).value) != "None":
                            if float(str(sheet.cell(i, column_index_from_string(valp)).value)) > float(
                                    str(sheet.cell(i, column_index_from_string(val2)).value)) > float(
                                    str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                    sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "above":
                    for i in range(2, sheet.max_row + 1):
                        if str(sheet.cell(i, column_index_from_string(val2)).value) != "None" and str(
                            sheet.cell(i, column_index_from_string(val)).value) != "None":
                            if float(str(sheet.cell(i, column_index_from_string(val2)).value)) < float(
                                    str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                    sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "below":
                    for i in range(2, sheet.max_row + 1):
                        if str(sheet.cell(i, column_index_from_string(val2)).value) != "None" and str(
                                sheet.cell(i, column_index_from_string(val)).value) != "None":
                            if float(str(sheet.cell(i, column_index_from_string(val2)).value)) > float(
                                    str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                    sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
            elif indicator_entry[ii] == "ema10" or indicator_entry[ii] == "ema20" or indicator_entry[ii] == "ema50" or \
                    indicator_entry[ii] == "ema100" or indicator_entry[ii] == "ema200" or (
                    indicator_entry[ii] == "clo" and (
                    value_entry[ii] == "ema10" or value_entry[ii] == "ema20" or value_entry[ii] == "ema50" or value_entry[
                ii] == "ema100" or value_entry[ii] == "ema200")):
                val = ""
                valp = ""
                val2 = ""
                val2p = ""
                if value_entry[ii] == "ema20":
                    val2 = "N"
                    val2p = "BC"
                elif value_entry[ii] == "ema50":
                    val2 = "O"
                    val2p = "BD"
                elif value_entry[ii] == "ema100":
                    val2 = "P"
                    val2p = "BE"
                elif value_entry[ii] == "ema200":
                    val2 = "Q"
                    val2p = "BF"
                elif value_entry[ii] == "ema10":
                    val2 = "BR"
                    val2p = "BS"

                if indicator_entry[ii] == "ema20":
                    val = "N"
                    valp = "BC"
                elif indicator_entry[ii] == "ema50":
                    val = "O"
                    valp = "BD"
                elif indicator_entry[ii] == "ema100":
                    val = "P"
                    valp = "BE"
                elif indicator_entry[ii] == "ema200":
                    val = "Q"
                    valp = "BF"
                elif indicator_entry[ii] == "ema10":
                    val = "BR"
                    valp = "BS"
                elif indicator_entry[ii] == "clo":
                    val = "AC"
                    valp = "AD"
                if parameter_entry[ii] == "crossover":
                    for i in range(2, sheet.max_row + 1):
                        if str(sheet.cell(i, column_index_from_string(valp)).value) != "None" and str(
                                sheet.cell(i, column_index_from_string(val2)).value) != "None" and str(
                            sheet.cell(i, column_index_from_string(val)).value) != "None":
                            if float(str(sheet.cell(i, column_index_from_string(valp)).value)) < float(
                                    str(sheet.cell(i, column_index_from_string(val2)).value)) < float(
                                    str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                    sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "crossunder":
                    for i in range(2, sheet.max_row + 1):
                        if str(sheet.cell(i, column_index_from_string(valp)).value) != "None" and str(
                                sheet.cell(i, column_index_from_string(val2)).value) != "None" and str(
                            sheet.cell(i, column_index_from_string(val)).value) != "None":
                            if float(str(sheet.cell(i, column_index_from_string(valp)).value)) > float(
                                    str(sheet.cell(i, column_index_from_string(val2)).value)) > float(
                                    str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                    sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "above":
                    for i in range(2, sheet.max_row + 1):
                        if str(sheet.cell(i, column_index_from_string(val2)).value) != "None" and str(
                                sheet.cell(i, column_index_from_string(val)).value) != "None":
                            if float(str(sheet.cell(i, column_index_from_string(val2)).value)) < float(
                                    str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                    sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "below":
                    for i in range(2, sheet.max_row + 1):
                        if str(sheet.cell(i, column_index_from_string(val2)).value) != "None" and str(
                                sheet.cell(i, column_index_from_string(val)).value) != "None":
                            if float(str(sheet.cell(i, column_index_from_string(val2)).value)) > float(
                                    str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                    sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                                stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                                ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                                value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                                close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
            elif indicator_entry[ii] == "clo" and (
                    value_entry[ii] == "pp" or value_entry[ii] == "s1" or value_entry[ii] == "s2" or value_entry[ii] == "s3" or
                    value_entry[ii] == "r1" or value_entry[ii] == "r2" or value_entry[ii] == "r3"):
                val = "AC"
                valp = "AD"
                val2 = ""
                if value_entry[ii] == "s3":
                    val2 = "G"
                elif value_entry[ii] == "s2":
                    val2 = "H"
                elif value_entry[ii] == "s1":
                    val2 = "I"
                elif value_entry[ii] == "pp":
                    val2 = "J"
                elif value_entry[ii] == "r1":
                    val2 = "K"
                elif value_entry[ii] == "r2":
                    val2 = "L"
                elif value_entry[ii] == "r3":
                    val2 = "M"
                if parameter_entry[ii] == "crossover":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string(valp)).value)) < float(
                                str(sheet.cell(i, column_index_from_string(val2)).value)) < float(
                                str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                               comon(close, close2)
                elif parameter_entry[ii] == "crossunder":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string(valp)).value)) > float(
                                str(sheet.cell(i, column_index_from_string(val2)).value)) > float(
                                str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "above":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string(val2)).value)) < float(
                                str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "below":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string(val2)).value)) > float(
                                str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
            elif indicator_entry[ii] == "cl" or indicator_entry[ii] == "ls" or (indicator_entry[ii] == "clo" and (
                    value_entry[ii] == "cl" or value_entry[ii] == "bl" or value_entry[ii] == "ls" or value_entry[ii] == "gc" or
                    value_entry[ii] == "rc")):
                val = ""
                valp = ""
                val2 = ""
                if value_entry[ii] == "cl":
                    val2 = "AP"
                elif value_entry[ii] == "bl":
                    val2 = "AQ"
                elif value_entry[ii] == "ls":
                    val2 = "AT"
                elif value_entry[ii] == "gc":
                    val2 = "AR"
                elif value_entry[ii] == "rc":
                    val2 = "AS"

                if indicator_entry[ii] == "clo":
                    val = "AC"
                    valp = "AD"
                elif indicator_entry[ii] == "cl":
                    val = "AP"
                    valp = "BA"
                elif indicator_entry[ii] == "ls":
                    val = "AT"
                    valp = "BB"
                if parameter_entry[ii] == "crossover":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string(valp)).value)) < float(
                                str(sheet.cell(i, column_index_from_string(val2)).value)) < float(
                                str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                               comon(close, close2)
                elif parameter_entry[ii] == "crossunder":
                    if value_entry[ii] == "gc":
                        val2 = "AS"
                    elif value_entry[ii] == "rc":
                        val2 = "AR"
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string(valp)).value)) > float(
                                str(sheet.cell(i, column_index_from_string(val2)).value)) > float(
                                str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "above":
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string(val2)).value)) < float(
                                str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)
                elif parameter_entry[ii] == "below":
                    if value_entry[ii] == "gc":
                        val2 = "AS"
                    elif value_entry[ii] == "rc":
                        val2 = "AR"
                    for i in range(2, sheet.max_row + 1):
                        if float(str(sheet.cell(i, column_index_from_string(val2)).value)) > float(
                                str(sheet.cell(i, column_index_from_string(val)).value)) and str(
                                sheet.cell(i, column_index_from_string(ni)).value) == "Yes":
                            stocks.append(sheet.cell(i, column_index_from_string('A')).value)
                            ticker.append(sheet.cell(i, column_index_from_string('B')).value)
                            value.append(round(float(str(sheet.cell(i, column_index_from_string(val)).value)), 2))
                            close.append(round(float(str(sheet.cell(i, column_index_from_string('AC')).value)), 2))
                    if len(stocks2) == 0:
                        stocks2, ticker2, value2, close2 = stocks, ticker, value, close
                        stocks, ticker, value, close = [], [], [], []
                    else:
                        stocks2, ticker2, value2, close2 = comon(stocks, stocks2), comon(ticker, ticker2), comon(value,
                                                                                                                 value2), \
                                                           comon(close, close2)


        zipal = zip(stocks2, ticker2, close2)
        dick = {'zipal': zipal}
        return render(request, 'scanner_tech_res.html', dick)
    except Exception as e:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, "errors.xlsx")
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row , 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")

def index(request):
    try:
        return render(request, 'scanner_tech.html')
    except Exception as e:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, 'errors.xlsx')
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row , 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")
