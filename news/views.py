# Create your views here.
import datetime
import requests
from bs4 import BeautifulSoup
from django.shortcuts import render
import openpyxl as xl
import os
import traceback
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip
def check_by_id(ele,id_):
    return (len(ele.find_all(id = id_)) > 0)
def check_by_tag(ele, name):
    return (len(ele.find_all(name)) > 0)
def check_by_class(ele, name ,class_name):
    return (len(ele.find_all(name,class_ = class_name) ) >0)

def fetch_head(add):
    news = []
    time = []
    href = []
    page = requests.get(add)
    soup = BeautifulSoup(page.content, 'html.parser')
    post_elems = soup.find(id="pageContent")
    list_item = post_elems.find("div", class_="featured")
    if (check_by_tag(list_item, "h2")):
        list = list_item.find("h2")
        news.append(list.get_text())
        href.append("https://economictimes.indiatimes.com" + list.find("a")['href'])
        time.append(list_item.find("time").get_text())

    if (check_by_tag(list_item, "h3")):
        list = list_item.find("h3")
        news.append(list.get_text())
        href.append("https://economictimes.indiatimes.com" + list.find("a")['href'])
        time.append(list_item.find_all("time")[1].get_text())
    if (check_by_class(post_elems,"ul", "list1")):
        list_item = post_elems.find("ul",class_ = "list1")
        list_item = list_item.find_all("li")
        for l in list_item:
            if(not l.has_attr('class') ):
                text = (l.find("a")).get_text()
                h = (l.find("a"))['href']
                href.append("https://economictimes.indiatimes.com" + h)
                if (text != ''):        news.append(text)
                t = (l.find("time")).get_text()
                if (t != ''):            time.append(t)

    if(check_by_class(post_elems,"div", "bThumb")):
        list_item = post_elems.find("div", class_ = "bThumb")
        list_item = list_item.find("ul")
        list_item = list_item.find_all("li")
        for l in list_item:
            l2 = l.find_all("a")
            if (l2):
                text = l2[1].get_text()
                h = l2[1]['href']
                news.append(text)
                href.append("https://economictimes.indiatimes.com" + h)
                time.append("")
    news2 = zip(news,href, time)
    return news2

def fetch_article(add):
    news = []
    time = []
    href = []
    para = []
    img = []
    page = requests.get(add)
    soup = BeautifulSoup(page.content, 'html.parser')
    if(check_by_class(soup, "div", "tabdata")):
        list_item = soup.find_all("div", class_ ="eachStory")
        for l in list_item:
            head = l.find("h3")
            head = head.find("a")
            if(not head): continue
            h = head['href']
            head = head.get_text()
            t = l.find("time")
            t = t.get_text()
            content = l.find("p")
            if(not content): continue
            content = content.get_text()
            sc = l.find('img')
            if( not sc): continue
            sc = sc['data-original']
            img.append(sc)
            news.append(head)
            time.append(t)
            href.append("https://economictimes.indiatimes.com"  + h)
            para.append(content)
    news2 = zip(news,href,time,para,img)
    return news2
def return_head(request, add, topic):
    try:
        dict = {'news': fetch_head(add),
                'type' : 1,
                'topic': topic}
        return render(request, 'newsh.html', dict)
    except Exception as e:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, f'errors.xlsx')  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row , 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")
def return_article(request, add, topic):
    try:
        dict = {'type': 2, "topic":topic}
        dict['news'] = fetch_article(add)
        return render(request, 'newsh.html', dict)
    except Exception as e:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, f'errors.xlsx')  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row , 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")

def industry(request):
    return return_head(request,"https://economictimes.indiatimes.com/industry","industry")

def auto_news(request):
    return return_article(request, "https://economictimes.indiatimes.com/industry/auto/auto-news/articlelist/64829342.cms","auto" )

def auto_cars(request):
    add = "https://economictimes.indiatimes.com/industry/auto/cars-uvs/articlelist/64829336.cms"
    return return_article(request, add, "auto")

def auto_two_three(request):
    add = "https://economictimes.indiatimes.com/industry/auto/two-wheelers-three-wheelers/articlelist/64829323.cms"
    return return_article(request, add, "auto")

def auto_lcv_hcv(request):
    add = "https://economictimes.indiatimes.com/industry/auto/lcv-hcv/articlelist/64829321.cms"
    return return_article(request, add, "auto")

def auto_components(request):
    add = "https://economictimes.indiatimes.com/industry/auto/auto-components/articlelist/64829316.cms"
    return return_article(request, add, "auto")

def auto_tyres(request):
    add = "https://economictimes.indiatimes.com/industry/auto/tyres/articlelist/64829311.cms"
    return return_article(request, add, "auto")

def banking_banking(request):
    add = "https://economictimes.indiatimes.com/industry/banking/finance/banking"
    return return_article(request, add, "banking")

def banking_finance(request):
    add = "https://economictimes.indiatimes.com/industry/banking-/-finance/banking"
    return return_article(request, add, "banking")

def banking_insure(request):
    add = "https://economictimes.indiatimes.com/industry/banking/finance/insure/articlelist/58456919.cms"
    return return_article(request, add, "banking")

def cons_durables(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/durables"
    return return_article(request, add, "cons")

def cons_electronics(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/electronics"
    return return_article(request, add, "cons")

def cons_fmcg(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/fmcg"
    return return_article(request, add, "cons")

def cons_food(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/food"
    return return_article(request, add, "cons")

def cons_garments_textiles(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/garments-/-textiles"
    return return_article(request, add, "cons")

def cons_liquor(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/liquor"
    return return_article(request, add, "cons")

def cons_paints(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/paints"
    return return_article(request, add, "cons")

def cons_tobacco(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/tobacco"
    return return_article(request, add, "cons")

def cons_fas_cos_jew(request):
    add = "https://economictimes.indiatimes.com/industry/cons-products/fashion-/-cosmetics-/-jewellery"
    return return_article(request, add, "cons")

def energy_power(request):
    add = "https://economictimes.indiatimes.com/industry/energy/power"
    return return_article(request, add, "energy")

def energy_oil_n_gas(request):
    add = "https://economictimes.indiatimes.com/industry/energy/oil-gas"
    return return_article(request, add, "energy")

def indgood_cons(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/construction"
    return return_article(request, add, "indgood")

def indgood_eng(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/engineering"
    return return_article(request, add, "indgood")

def indgood_cement(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/cement"
    return return_article(request, add, "indgood")

def indgood_chem_fertilisers(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/chem-/-fertilisers"
    return return_article(request, add, "indgood")

def indgood_metals_n_mining(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/metals-mining"
    return return_article(request, add, "indgood")

def indgood_pack(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/packaging"
    return return_article(request, add, "indgood")

def indgood_pwgpm(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/paper-/-wood-/-glass/-plastic/-marbles"
    return return_article(request, add, "indgood")

def indgood_petrochem(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/petrochem"
    return return_article(request, add, "indgood")

def indgood_steel(request):
    add = "https://economictimes.indiatimes.com/industry/indl-goods/svs/steel"
    return return_article(request, add, "indgood")

def health_healthcare(request):
    add = "https://economictimes.indiatimes.com/industry/healthcare/biotech/healthcare"
    return return_article(request, add, "health")

def health_bio(request):
    add = "https://economictimes.indiatimes.com/industry/healthcare-/-biotech/biotech"
    return return_article(request, add, "health")

def health_pharm(request):
    add = "https://economictimes.indiatimes.com/industry/healthcare/biotech/pharmaceuticals"
    return return_article(request, add, "health")

def services_advertising(request):
    add = "https://economictimes.indiatimes.com/industry/services/advertising"
    return return_article(request, add, "services")

def services_consultancy_audit(request):
    add = "https://economictimes.indiatimes.com/industry/services/consultancy-/-audit"
    return return_article(request, add, "services")

def services_education(request):
    add = "https://economictimes.indiatimes.com/industry/services/education"
    return return_article(request, add, "services")

def services_hotels_restaurants(request):
    add = "https://economictimes.indiatimes.com/industry/services/hotels-/-restaurants"
    return return_article(request, add, "services")

def services_property_cons(request):
    add = "https://economictimes.indiatimes.com/industry/services/property-/-cstruction"
    return return_article(request, add, "services")

def services_retail(request):
    add = "https://economictimes.indiatimes.com/industry/services/retail"
    return return_article(request, add, "services")

def services_travel(request):
    add = "https://economictimes.indiatimes.com/industry/services/travel"
    return return_article(request, add, "services")

def more_entertainment(request):
    add = "https://economictimes.indiatimes.com/industry/media-/-entertainment/entertainment"
    return return_article(request, add, "more")

def more_media(request):
    add = "https://economictimes.indiatimes.com/industry/media-/-entertainment/media"
    return return_article(request, add, "more")

def more_railways(request):
    add = "https://economictimes.indiatimes.com/industry/transportation/railways"
    return return_article(request, add, "more")
def more_airlines_aviation(request):
    add = "https://economictimes.indiatimes.com/industry/transportation/airlines-/-aviation"
    return return_article(request, add, "more")

def more_shipping_transport(request):
    add = "https://economictimes.indiatimes.com/industry/transportation/shipping-/-transport"
    return return_article(request, add, "more")

def more_roadways(request):
    add = "https://economictimes.indiatimes.com/industry/transportation/roadways/articlelist/58456933.cms"
    return return_article(request, add, "more")

def more_tel_news(request):
    add = "https://economictimes.indiatimes.com/industry/telecom/telecom-news/articlelist/64256852.cms"
    return return_article(request, add, "more")

def more_tel_policy(request):
    add = "https://economictimes.indiatimes.com/industry/telecom/telecom-policy/articlelist/64256834.cms"
    return return_article(request, add, "more")

def more_csr_initiatives(request):
    add = "https://economictimes.indiatimes.com/news/india-unlimited/csr/initiatives/articlelist/47068922.cms"
    return return_article(request, add, "more")

def more_csr_policy(request):
    add = "https://economictimes.indiatimes.com/news/india-unlimited/csr/policy/articlelist/47068917.cms"
    return return_article(request, add, "more")

def more_tech(request):
    return return_head(request,"https://economictimes.indiatimes.com/tech","more")

def more_misc(request):
    add = "https://economictimes.indiatimes.com/industry/miscellaneous/articlelist/58456958.cms"
    return return_article(request, add, "more")

def more_env(request):
    return return_head(request,"https://economictimes.indiatimes.com/environment", "more" )

def index(request):
    return industry(request)


