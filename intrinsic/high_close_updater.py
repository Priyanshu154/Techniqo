import openpyxl as xl
import requests
from bs4 import BeautifulSoup
nifty = ["3MINDIA", "ABB", "ACC", "AIAENG", "APLAPOLLO", "AUBANK", "AARTIIND", "AAVAS", "ABBOTINDIA", "ADANIGAS",
         "ADANIGREEN", "ADANIPORTS", "ADANIPOWER", "ADANITRANS", "ABCAPITAL", "ABFRL", "ADVENZYMES", "AEGISCHEM",
         "AFFLE", "AJANTPHARM", "AKZOINDIA", "APLLTD", "ALKEM", "ALKYLAMINE", "ALLCARGO", "AMARAJABAT", "AMBER",
         "AMBUJACEM", "APOLLOHOSP", "APOLLOTYRE", "ARVINDFASN", "ASAHIINDIA", "ASHOKLEY", "ASHOKA", "ASIANPAINT",
         "ASTERDM", "ASTRAZEN", "ASTRAL", "ATUL", "AUROPHARMA", "AVANTIFEED", "DMART", "AXISBANK", "BASF", "BEML",
         "BSE", "BAJAJ-AUTO", "BAJAJCON", "BAJAJELEC", "BAJFINANCE", "BAJAJFINSV", "BAJAJHLDNG", "BALKRISIND",
         "BALMLAWRIE", "BALRAMCHIN", "BANDHANBNK", "BANKBARODA", "BANKINDIA", "MAHABANK", "BATAINDIA", "BAYERCROP",
         "BERGEPAINT", "BDL", "BEL", "BHARATFORG", "BHEL", "BPCL", "BHARATRAS", "BHARTIARTL", "INFRATEL", "BIOCON",
         "BIRLACORPN", "BSOFT", "BLISSGVS", "BLUEDART", "BLUESTARCO", "BBTC", "BOMDYEING", "BOSCHLTD", "BRIGADE",
         "BRITANNIA", "CARERATING", "CCL", "CESC", "CRISIL", "CSBBANK", "CADILAHC", "CANFINHOME", "CANBK", "CAPLIPOINT",
         "CGCL", "CARBORUNIV", "CASTROLIND", "CEATLTD", "CENTRALBK", "CDSL", "CENTURYPLY", "CENTURYTEX", "CERA",
         "CHAMBLFERT", "CHENNPETRO", "CHOLAHLDNG", "CHOLAFIN", "CIPLA", "CUB", "COALINDIA", "COCHINSHIP", "COLPAL",
         "CONCOR", "COROMANDEL", "CREDITACC", "CROMPTON", "CUMMINSIND", "CYIENT", "DBCORP", "DCBBANK", "DCMSHRIRAM",
         "DLF", "DABUR", "DALBHARAT", "DEEPAKNTR", "DELTACORP", "DHANUKA", "DBL", "DISHTV", "DCAL", "DIVISLAB", "DIXON",
         "LALPATHLAB", "DRREDDY", "EIDPARRY", "EIHOTEL", "ESABINDIA", "EDELWEISS", "EICHERMOT", "ELGIEQUIP", "EMAMILTD",
         "ENDURANCE", "ENGINERSIN", "EQUITAS", "ERIS", "ESCORTS", "ESSELPACK", "EXIDEIND", "FDC", "FEDERALBNK",
         "FINEORG", "FINCABLES", "FINPIPE", "FSL", "FORTIS", "FCONSUMER", "FRETAIL", "GAIL", "GEPIL", "GET&D", "GHCL",
         "GMMPFAUDLR", "GMRINFRA", "GALAXYSURF", "GRSE", "GARFIBRES", "GICRE", "GILLETTE", "GLAXO", "GLENMARK",
         "GODFRYPHLP", "GODREJAGRO", "GODREJCP", "GODREJIND", "GODREJPROP", "GRANULES", "GRAPHITE", "GRASIM", "GESHIP",
         "GREAVESCOT", "GRINDWELL", "GUJALKALI", "FLUOROCHEM", "GUJGASLTD", "GMDCLTD", "GNFC", "GPPL", "GSFC", "GSPL",
         "GULFOILLUB", "HEG", "HCLTECH", "HDFCAMC", "HDFCBANK", "HDFCLIFE", "HFCL", "HATHWAY", "HATSUN", "HAVELLS",
         "HEIDELBERG", "HERITGFOOD", "HEROMOTOCO", "HEXAWARE", "HSCL", "HIMATSEIDE", "HINDALCO", "HAL", "HINDCOPPER",
         "HINDPETRO", "HINDUNILVR", "HINDZINC", "HONAUT", "HUDCO", "HDFC", "ICICIBANK", "ICICIGI", "ICICIPRULI", "ISEC",
         "ICRA", "IDBI", "IDFCFIRSTB", "IDFC", "IFBIND", "IFCI", "IIFL", "IIFLWAM", "IRB", "IRCON", "ITC", "ITI",
         "INDIACEM", "ITDC", "IBULHSGFIN", "IBREALEST", "IBVENTURES", "INDIAMART", "INDIANB", "IEX", "INDHOTEL", "IOC",
         "IOB", "IRCTC", "INDOSTAR", "INDOCO", "IGL", "INDUSINDBK", "INFIBEAM", "NAUKRI", "INFY", "INGERRAND",
         "INOXLEISUR", "INTELLECT", "INDIGO", "IPCALAB", "JBCHEPHARM", "JKCEMENT", "JKLAKSHMI", "JKPAPER", "JKTYRE",
         "JMFINANCIL", "JSWENERGY", "JSWSTEEL", "JAGRAN", "JAICORPLTD", "J&KBANK", "JAMNAAUTO", "JINDALSAW", "JSLHISAR",
         "JSL", "JINDALSTEL", "JCHAC", "JUBLFOOD", "JUBILANT", "JUSTDIAL", "JYOTHYLAB", "KPRMILL", "KEI", "KNRCON",
         "KPITTECH", "KRBL", "KSB", "KAJARIACER", "KALPATPOWR", "KANSAINER", "KTKBANK", "KARURVYSYA", "KSCL", "KEC",
         "KOLTEPATIL", "KOTAKBANK", "L&TFH", "LTTS", "LICHSGFIN", "LAOPALA", "LAXMIMACH", "LTI", "LT", "LAURUSLABS",
         "LEMONTREE", "LINDEINDIA", "LUPIN", "LUXIND", "MASFIN", "MMTC", "MOIL", "MRF", "MGL", "MAHSCOOTER",
         "MAHSEAMLES", "M&MFIN", "M&M", "MAHINDCIE", "MHRIL", "MAHLOG", "MANAPPURAM", "MRPL", "MARICO", "MARUTI",
         "MFSL", "METROPOLIS", "MINDTREE", "MINDACORP", "MINDAIND", "MIDHANI", "MOTHERSUMI", "MOTILALOFS", "MPHASIS",
         "MCX", "MUTHOOTFIN", "NATCOPHARM", "NBCC", "NCC", "NESCO", "NHPC", "NIITTECH", "NLCINDIA", "NMDC", "NTPC",
         "NH", "NATIONALUM", "NFL", "NBVENTURES", "NAVINFLUOR", "NESTLEIND", "NILKAMAL", "NAM-INDIA", "OBEROIRLTY",
         "ONGC", "OIL", "OMAXE", "OFSS", "ORIENTCEM", "ORIENTELEC", "ORIENTREF", "PIIND", "PNBHOUSING", "PNCINFRA",
         "PTC", "PVR", "PAGEIND", "PERSISTENT", "PETRONET", "PFIZER", "PHILIPCARB", "PHOENIXLTD", "PIDILITIND", "PEL",
         "POLYMED", "POLYCAB", "PFC", "POWERGRID", "PRAJIND", "PRESTIGE", "PRSMJOHNSN", "PGHL", "PGHH", "PNB", "QUESS",
         "RBLBANK", "RECLTD", "RITES", "RADICO", "RVNL", "RAIN", "RAJESHEXPO", "RALLIS", "RCF", "RATNAMANI", "RAYMOND",
         "REDINGTON", "RELAXO", "RELIANCE", "REPCOHOME", "SBICARD", "SBILIFE", "SJVN", "SKFINDIA", "SRF", "SADBHAV",
         "SANOFI", "SCHAEFFLER", "SCHNEIDER", "SIS", "SEQUENT", "SFL", "SCI", "SHOPERSTOP", "SHREECEM", "RENUKA",
         "SHRIRAMCIT", "SRTRANSFIN", "SIEMENS", "SOBHA", "SOLARINDS", "SONATSOFTW", "SOUTHBANK", "SPANDANA", "SPICEJET",
         "STARCEMENT", "SBIN", "SAIL", "SWSOLAR", "STRTECH", "STAR", "SUDARSCHEM", "SUMICHEM", "SPARC", "SUNPHARMA",
         "SUNTV", "SUNDARMFIN", "SUNDRMFAST", "SUNTECK", "SUPRAJIT", "SUPREMEIND", "SUZLON", "SWANENERGY", "SYMPHONY",
         "SYNGENE", "TCIEXP", "TCNSBRANDS", "TTKPRESTIG", "TVTODAY", "TV18BRDCST", "TVSMOTOR", "TAKE", "TASTYBITE",
         "TATACOMM", "TCS", "TATACONSUM", "TATAELXSI", "TATAINVEST", "TATAMTRDVR", "TATAMOTORS", "TATAPOWER",
         "TATASTLBSL", "TATASTEEL", "TEAMLEASE", "TECHM", "NIACL", "RAMCOCEM", "THERMAX", "THYROCARE", "TIMETECHNO",
         "TIMKEN", "TITAN", "TORNTPHARM", "TORNTPOWER", "TRENT", "TRIDENT", "TIINDIA", "UCOBANK", "UFLEX", "UPL",
         "UJJIVAN", "UJJIVANSFB", "ULTRACEMCO", "UNIONBANK", "UBL", "MCDOWELL-N", "VGUARD", "VMART", "VIPIND", "VRLLOG",
         "VSTIND", "VAIBHAVGBL", "VAKRANGEE", "VTL", "VARROC", "VBL", "VENKEYS", "VESUVIUS", "VINATIORGA", "IDEA",
         "VOLTAS", "WABCOINDIA", "WELCORP", "WELSPUNIND", "WESTLIFE", "WHIRLPOOL", "WIPRO", "WOCKPHARMA", "ZEEL",
         "ZENSARTECH", "ZYDUSWELL", "ECLERX"]

wb = xl.load_workbook('high_close.xlsx')
wb_source = xl.load_workbook("source_urls.xlsx")
sheet = wb["Sheet1"]
source = wb_source["sheet"]
def get_url_name(ticker):
    for i in range(2, source.max_row + 1):
        if(source.cell(i,3).value == ticker):
            return source.cell(i,1).value, source.cell(i,2).value
    return "invalid","invalid"
row_no = 2
for i in range(len(nifty)):
    url,name = get_url_name(nifty[i])
    if(url != "invalid"):
        try:
            page = requests.get(url)
            soup = BeautifulSoup(page.content, 'html.parser')
            wrapper_div = soup.find(id = "div_nse_livebox_wrap")
            inner_div = wrapper_div.find(class_ = "open_lhs1")
            inner_div2 = wrapper_div.find(class_ = "div_live_price_wrap")
            prev_close = inner_div2.find(class_ = "span_price_wrap")
            low_high = inner_div.find(class_ = "week52_lowhigh_wrap")
            low = low_high.find(class_ = "low_high1")
            high = low_high.find(class_ = "low_high3")
            sheet.cell(row_no, 1).value = name
            sheet.cell(row_no, 2).value = nifty[i]
            sheet.cell(row_no, 3).value = prev_close.get_text()
            sheet.cell(row_no, 4).value = high.get_text()
            sheet.cell(row_no, 5).value = low.get_text()
            wb.save("high_close.xlsx")
            row_no = row_no + 1
            print(i)
        except:
            print("scrap_prblm : " + nifty[i])
    else:
        print(nifty[i])