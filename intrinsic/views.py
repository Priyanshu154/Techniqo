from django.shortcuts import render
import openpyxl as xl
import os
import datetime
import traceback
# Create your views here.
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip
def index(request):
    try:
        workpath = os.path.dirname(os.path.abspath(__file__))
        xx = os.path.join(workpath, 'high_close.xlsx')
        wb = xl.load_workbook(xx, data_only=True)
        sheet2 = wb['Sheet1']
        stock = []
        for i in range(2, sheet2.max_row + 1):
            stock.append(sheet2.cell(i,1).value)
        dictt = {'stocks': stock}
        return render(request, 'intrinsich.html', dictt)
    except Exception as e:
        workpath = os.path.dirname( os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, f"errors.xlsx")  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row , 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")
def value(request):
    try:
        name = request.POST.get('stock_name', 'default')
        workpath = os.path.dirname(os.path.abspath(__file__))
        xx = os.path.join(workpath, 'high_close.xlsx')
        wb = xl.load_workbook(xx, data_only=True)
        sheet2 = wb['Sheet1']
        intrinsic_value = 0
        current_value = 0
        sentiment = ''
        f = 0
        name = name.split(" |")[0]
        for i in range(2, sheet2.max_row+1):
            if sheet2.cell(i, 1).value == name:
                intrinsic_value = sheet2.cell(i, 7).value
                current_value = sheet2.cell(i, 3).value
                sentiment = sheet2.cell(i, 8).value
                f = 1
                break
        dictt = {'intrinsic_values': int(intrinsic_value), 'sentiments': sentiment, 'flag': f, 'ltp': current_value, 'name':name}
        return render(request, 'intrinsic_value.html', dictt)
    except Exception as e:
        workpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        xx = os.path.join(workpath, f"errors.xlsx")  # yaha tak
        wb = xl.load_workbook(xx)
        sheet1 = wb["Sheet1"]
        sheet1.cell(sheet1.max_row+1, 1).value = str(e)
        sheet1.cell(sheet1.max_row,  2).value = request.path_info
        sheet1.cell(sheet1.max_row , 3).value = datetime.datetime.now()
        sheet1.cell(sheet1.max_row , 4).value = traceback.format_exc()
        wb.save(xx)
        return render(request, "oops.html")
