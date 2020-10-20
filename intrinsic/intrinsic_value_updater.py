import openpyxl as xl

wb = xl.load_workbook('intrinsicn.xlsx')
wb2 = xl.load_workbook('intrinsic.xlsx')

sheet1 = wb['Overview']
sheet2 = wb2['intrinsic']

for i in range(2, sheet2.max_row + 1):
    for j in range(2, sheet1.max_row + 1):
        if sheet1.cell(j,1).value == sheet2.cell(i,1).value:
            sheet2.cell(i, 2).value = sheet1.cell(j,2).value
            sheet2.cell(i, 3).value = sheet1.cell(j,3).value
            sheet2.cell(i, 4).value = sheet1.cell(j,4).value
            sheet2.cell(i, 6).value = sheet1.cell(j,7).value
            break

wb.save('intrinsicn.xlsx')
wb2.save('intrinsic.xlsx')