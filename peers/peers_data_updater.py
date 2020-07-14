import openpyxl as xl
from openpyxl.utils import column_index_from_string

wb = xl.load_workbook('peers.xlsx')
wb2 = xl.load_workbook('peers_new.xlsx')


sheet1 = wb['Sheet1']
sheet2 = wb2['Overview']


for i in range(2, sheet1.max_row + 1):
    for j in range(2, sheet2.max_row + 1):
        if sheet1.cell(i,1).value == sheet2.cell(j,1).value:
            for z in range(2,14):
                sheet1.cell(i,z+1).value = sheet2.cell(j,z).value

wb.save('peers.xlsx')
wb2.save('peers_new.xlsx')
